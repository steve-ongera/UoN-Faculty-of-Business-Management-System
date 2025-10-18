from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from .models import *


@csrf_protect
@never_cache
def login_view(request):
    """
    Unified login view for all user types.
    Redirects to appropriate dashboard based on user_type.
    """
    # Redirect if already logged in
    if request.user.is_authenticated:
        return redirect_to_dashboard(request.user)
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        remember = request.POST.get('remember')
        
        # Authenticate user
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.is_active and user.is_active_user:
                # Log the user in
                login(request, user)
                
                # Set session expiry
                if not remember:
                    request.session.set_expiry(0)  # Browser close
                else:
                    request.session.set_expiry(1209600)  # 2 weeks
                
                # Success message
                messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
                
                # Redirect to appropriate dashboard
                return redirect_to_dashboard(user)
            else:
                messages.error(request, 'Your account has been deactivated. Please contact the administrator.')
        else:
            messages.error(request, 'Invalid email or password. Please try again.')
    
    return render(request, 'authentication/login.html')


def redirect_to_dashboard(user):
    """
    Redirect user to appropriate dashboard based on user_type.
    """
    user_type = user.user_type
    
    dashboard_urls = {
        'STUDENT': 'student_dashboard',
        'LECTURER': 'lecturer_dashboard',
        'COD': 'cod_dashboard',
        'DEAN': 'dean_dashboard',
        'ICT_ADMIN': 'admin_dashboard',
    }
    
    # If superuser, redirect to admin dashboard
    if user.is_superuser or user.is_staff:
        return redirect('admin_dashboard')
    
    # Get the appropriate dashboard URL
    dashboard_url = dashboard_urls.get(user_type, 'admin_dashboard')
    
    return redirect(dashboard_url)


@login_required
def logout_view(request):
    """
    Logout view for all users.
    """
    user_name = request.user.get_full_name() or request.user.username
    logout(request)
    messages.success(request, f'Goodbye, {user_name}! You have been logged out successfully.')
    return redirect('login')


# ========================
# DASHBOARD VIEWS
# ========================

@login_required
def student_dashboard(request):
    """
    Dashboard for students.
    """
    if request.user.user_type != 'STUDENT':
        messages.error(request, 'Access denied. You do not have permission to view this page.')
        return redirect_to_dashboard(request.user)
    
    try:
        student = request.user.student_profile
        
        # Get student data
        context = {
            'student': student,
            'current_semester': get_current_semester(),
            'enrolled_units': get_student_enrolled_units(student),
            'upcoming_classes': get_student_timetable(student),
            'recent_announcements': get_student_announcements(student),
            'fee_balance': get_student_fee_balance(student),
            'academic_performance': get_student_performance(student),
        }
        
        return render(request, 'dashboards/student_dashboard.html', context)
    
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        return redirect('login')


@login_required
def lecturer_dashboard(request):
    """
    Dashboard for lecturers.
    """
    if request.user.user_type != 'LECTURER':
        messages.error(request, 'Access denied. You do not have permission to view this page.')
        return redirect_to_dashboard(request.user)
    
    try:
        lecturer = request.user.lecturer_profile
        
        # Get lecturer data
        context = {
            'lecturer': lecturer,
            'current_semester': get_current_semester(),
            'allocated_units': get_lecturer_units(lecturer),
            'today_classes': get_lecturer_today_classes(lecturer),
            'pending_marks': get_pending_marks_count(lecturer),
            'student_count': get_lecturer_student_count(lecturer),
            'recent_announcements': get_general_announcements(),
        }
        
        return render(request, 'dashboards/lecturer_dashboard.html', context)
    
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        return redirect('login')


@login_required
def cod_dashboard(request):
    """
    Dashboard for Chairman of Department (COD).
    """
    if request.user.user_type != 'COD':
        messages.error(request, 'Access denied. You do not have permission to view this page.')
        return redirect_to_dashboard(request.user)
    
    try:
        # Get COD's department
        departments = request.user.headed_departments.all()
        
        if not departments.exists():
            messages.warning(request, 'You are not assigned as head of any department.')
            return render(request, 'dashboards/cod_dashboard.html', {'departments': []})
        
        department = departments.first()
        
        # Get department data
        context = {
            'department': department,
            'current_semester': get_current_semester(),
            'total_students': get_department_students_count(department),
            'total_lecturers': get_department_lecturers_count(department),
            'total_programmes': get_department_programmes_count(department),
            'pending_approvals': get_pending_grade_approvals(department),
            'recent_announcements': get_general_announcements(),
        }
        
        return render(request, 'dashboards/cod_dashboard.html', context)
    
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        return redirect('login')


@login_required
def dean_dashboard(request):
    """
    Dashboard for Dean.
    """
    if request.user.user_type != 'DEAN':
        messages.error(request, 'Access denied. You do not have permission to view this page.')
        return redirect_to_dashboard(request.user)
    
    try:
        # Get faculty-wide data
        context = {
            'current_semester': get_current_semester(),
            'total_students': get_all_students_count(),
            'total_lecturers': get_all_lecturers_count(),
            'total_programmes': get_all_programmes_count(),
            'total_departments': get_all_departments_count(),
            'revenue_this_semester': get_semester_revenue(),
            'pending_approvals': get_all_pending_approvals(),
            'recent_events': get_upcoming_events(),
            'recent_announcements': get_general_announcements(),
        }
        
        return render(request, 'dashboards/dean_dashboard.html', context)
    
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        return redirect('login')


@login_required
def admin_dashboard(request):
    """
    Dashboard for ICT Admin and Superusers.
    """
    if request.user.user_type not in ['ICT_ADMIN'] and not (request.user.is_superuser or request.user.is_staff):
        messages.error(request, 'Access denied. You do not have permission to view this page.')
        return redirect_to_dashboard(request.user)
    
    try:
        # Get system-wide statistics
        context = {
            'current_semester': get_current_semester(),
            'total_users': get_total_users(),
            'total_students': get_all_students_count(),
            'total_lecturers': get_all_lecturers_count(),
            'total_programmes': get_all_programmes_count(),
            'total_departments': get_all_departments_count(),
            'active_sessions': get_active_sessions_count(),
            'system_health': get_system_health(),
            'recent_activities': get_recent_system_activities(),
        }
        
        return render(request, 'dashboards/admin_dashboard.html', context)
    
    except Exception as e:
        messages.error(request, f'Error loading dashboard: {str(e)}')
        return redirect('login')


# ========================
# HELPER FUNCTIONS
# ========================

def get_current_semester():
    """Get the current active semester."""
    from .models import Semester
    try:
        return Semester.objects.filter(is_current=True).first()
    except:
        return None


def get_student_enrolled_units(student):
    """Get units enrolled by student in current semester."""
    from .models import UnitEnrollment
    current_semester = get_current_semester()
    if current_semester:
        return UnitEnrollment.objects.filter(
            student=student,
            semester=current_semester,
            status='ENROLLED'
        ).select_related('unit')
    return []


def get_student_timetable(student):
    """Get student's timetable for current semester."""
    from .models import TimetableSlot
    current_semester = get_current_semester()
    if current_semester:
        return TimetableSlot.objects.filter(
            programme=student.programme,
            year_level=student.current_year,
            unit_allocation__semester=current_semester,
            is_active=True
        ).select_related('unit_allocation__unit', 'venue').order_by('day_of_week', 'start_time')[:5]
    return []


def get_student_announcements(student):
    """Get announcements relevant to student."""
    from .models import Announcement
    from django.utils import timezone
    return Announcement.objects.filter(
        is_published=True,
        publish_date__lte=timezone.now()
    ).filter(
        target_programmes=student.programme
    ).order_by('-publish_date')[:5]


def get_student_fee_balance(student):
    """Get student's fee balance."""
    from .models import FeeStatement
    current_semester = get_current_semester()
    if current_semester:
        try:
            statement = FeeStatement.objects.get(student=student, semester=current_semester)
            return {
                'total_billed': statement.total_billed,
                'total_paid': statement.total_paid,
                'balance': statement.balance,
                'can_register': statement.can_register
            }
        except:
            return None
    return None


def get_student_performance(student):
    """Get student's academic performance summary."""
    from .models import FinalGrade
    from django.db.models import Avg, Count
    
    grades = FinalGrade.objects.filter(
        enrollment__student=student,
        is_approved=True
    )
    
    return {
        'total_units': grades.count(),
        'average_grade_point': grades.aggregate(avg=Avg('grade_point'))['avg'] or 0,
        'units_passed': grades.filter(grade__in=['A', 'B', 'C', 'D']).count(),
        'units_failed': grades.filter(grade='F').count(),
    }


def get_lecturer_units(lecturer):
    """Get units allocated to lecturer."""
    from .models import UnitAllocation
    current_semester = get_current_semester()
    if current_semester:
        return UnitAllocation.objects.filter(
            lecturer=lecturer,
            semester=current_semester,
            is_active=True
        ).select_related('unit').prefetch_related('programmes')
    return []


def get_lecturer_today_classes(lecturer):
    """Get lecturer's classes for today."""
    from .models import TimetableSlot
    from datetime import datetime
    current_semester = get_current_semester()
    today = datetime.now().strftime('%A').upper()
    
    if current_semester:
        return TimetableSlot.objects.filter(
            unit_allocation__lecturer=lecturer,
            unit_allocation__semester=current_semester,
            day_of_week=today,
            is_active=True
        ).select_related('venue', 'unit_allocation__unit').order_by('start_time')
    return []


def get_pending_marks_count(lecturer):
    """Get count of pending marks entry."""
    from .models import StudentMarks, UnitEnrollment
    current_semester = get_current_semester()
    if current_semester:
        # Get enrollments for lecturer's units
        enrollments = UnitEnrollment.objects.filter(
            unit__allocations__lecturer=lecturer,
            semester=current_semester,
            status='ENROLLED'
        ).count()
        
        # Get marks already entered
        marks_entered = StudentMarks.objects.filter(
            entered_by=lecturer,
            enrollment__semester=current_semester
        ).values('enrollment').distinct().count()
        
        return enrollments - marks_entered
    return 0


def get_lecturer_student_count(lecturer):
    """Get total students taught by lecturer."""
    from .models import UnitEnrollment
    current_semester = get_current_semester()
    if current_semester:
        return UnitEnrollment.objects.filter(
            unit__allocations__lecturer=lecturer,
            semester=current_semester,
            status='ENROLLED'
        ).values('student').distinct().count()
    return 0


def get_general_announcements():
    """Get general announcements."""
    from .models import Announcement
    from django.utils import timezone
    return Announcement.objects.filter(
        is_published=True,
        publish_date__lte=timezone.now()
    ).order_by('-publish_date')[:5]


def get_department_students_count(department):
    """Get count of students in department."""
    from .models import Student
    return Student.objects.filter(
        programme__department=department,
        is_active=True
    ).count()


def get_department_lecturers_count(department):
    """Get count of lecturers in department."""
    from .models import Lecturer
    return Lecturer.objects.filter(
        department=department,
        is_active=True
    ).count()


def get_department_programmes_count(department):
    """Get count of programmes in department."""
    from .models import Programme
    return Programme.objects.filter(
        department=department,
        is_active=True
    ).count()


def get_pending_grade_approvals(department):
    """Get pending grade approvals for department."""
    from .models import FinalGrade
    return FinalGrade.objects.filter(
        enrollment__unit__department=department,
        is_approved=False
    ).count()


def get_all_students_count():
    """Get total count of active students."""
    from .models import Student
    return Student.objects.filter(is_active=True).count()


def get_all_lecturers_count():
    """Get total count of active lecturers."""
    from .models import Lecturer
    return Lecturer.objects.filter(is_active=True).count()


def get_all_programmes_count():
    """Get total count of active programmes."""
    from .models import Programme
    return Programme.objects.filter(is_active=True).count()


def get_all_departments_count():
    """Get total count of departments."""
    from .models import Department
    return Department.objects.count()


def get_semester_revenue():
    """Get total revenue for current semester."""
    from .models import FeePayment
    from django.db.models import Sum
    current_semester = get_current_semester()
    if current_semester:
        total = FeePayment.objects.filter(
            semester=current_semester
        ).aggregate(total=Sum('amount_paid'))['total']
        return total or 0
    return 0


def get_all_pending_approvals():
    """Get all pending grade approvals."""
    from .models import FinalGrade
    return FinalGrade.objects.filter(is_approved=False).count()


def get_upcoming_events():
    """Get upcoming events."""
    from .models import Event
    from django.utils import timezone
    return Event.objects.filter(
        event_date__gte=timezone.now().date()
    ).order_by('event_date', 'start_time')[:5]


def get_total_users():
    """Get total count of users."""
    from .models import User
    return User.objects.filter(is_active=True).count()


def get_active_sessions_count():
    """Get count of active user sessions."""
    from django.contrib.sessions.models import Session
    from django.utils import timezone
    return Session.objects.filter(expire_date__gte=timezone.now()).count()


def get_system_health():
    """Get system health metrics."""
    # This is a placeholder - implement actual health checks
    return {
        'status': 'healthy',
        'database': 'connected',
        'cache': 'active',
        'uptime': '99.9%'
    }


def get_recent_system_activities():
    """Get recent system activities."""
    # This is a placeholder - implement actual activity logging
    return []

# ========================
#student views.py
# No changes made to this file
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone


@login_required
@never_cache
def student_announcements_list(request):
    """
    List all announcements visible to the student.
    Announcements are filtered based on student's programme.
    """
    if request.user.user_type != 'STUDENT':
        return redirect('student_dashboard')
    
    try:
        student = request.user.student_profile
        
        # Get announcements for student's programme
        announcements = Announcement.objects.filter(
            is_published=True,
            publish_date__lte=timezone.now()
        ).filter(
            Q(target_programmes=student.programme) | Q(target_programmes__isnull=True)
        ).order_by('-publish_date').distinct()
        
        # Search functionality
        search_query = request.GET.get('search', '')
        if search_query:
            announcements = announcements.filter(
                Q(title__icontains=search_query) | Q(content__icontains=search_query)
            )
        
        # Filter by priority
        priority_filter = request.GET.get('priority', '')
        if priority_filter:
            announcements = announcements.filter(priority=priority_filter)
        
        # Pagination
        paginator = Paginator(announcements, 10)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        
        context = {
            'page_obj': page_obj,
            'announcements': page_obj.object_list,
            'search_query': search_query,
            'priority_filter': priority_filter,
            'priorities': [
                ('LOW', 'Low'),
                ('NORMAL', 'Normal'),
                ('HIGH', 'High'),
                ('URGENT', 'Urgent'),
            ],
            'total_announcements': announcements.count(),
        }
        
        return render(request, 'student/announcements/announcements_list.html', context)
    
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Error loading announcements: {str(e)}')
        return redirect('student_dashboard')


@login_required
@never_cache
def student_announcement_detail(request, pk):
    """
    Display detailed view of a single announcement.
    Only shows announcements visible to student's programme.
    """
    if request.user.user_type != 'STUDENT':
        return redirect('student_dashboard')
    
    try:
        student = request.user.student_profile
        
        # Get announcement and verify access
        announcement = get_object_or_404(
            Announcement,
            pk=pk,
            is_published=True,
            publish_date__lte=timezone.now()
        )
        
        # Check if announcement is for student's programme
        if announcement.target_programmes.exists():
            if student.programme not in announcement.target_programmes.all():
                return redirect('student_announcements_list')
        
        # Check if announcement has expired
        if announcement.expiry_date and announcement.expiry_date < timezone.now():
            from django.contrib import messages
            messages.warning(request, 'This announcement has expired.')
            return redirect('student_announcements_list')
        
        # Get related announcements (same programme)
        related_announcements = Announcement.objects.filter(
            is_published=True,
            publish_date__lte=timezone.now(),
            target_programmes=student.programme
        ).exclude(pk=pk).order_by('-publish_date')[:5]
        
        context = {
            'announcement': announcement,
            'related_announcements': related_announcements,
        }
        
        return render(request, 'student/announcements/announcement_detail.html', context)
    
    except Exception as e:
        from django.contrib import messages
        messages.error(request, f'Error loading announcement: {str(e)}')
        return redirect('student_announcements_list')



# views.py - Student Events Views
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from django.contrib import messages
from .models import Event, EventRegistration, Student, Programme

@login_required(login_url='login')
def student_events_list(request):
    """List all upcoming events for students"""
    # Get current student
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    # Get student's programme
    student_programme = student.programme
    
    # Base queryset - events for student's programme or general events
    events = Event.objects.filter(
        Q(target_programmes=student_programme) | Q(target_programmes__isnull=True),
        event_date__gte=timezone.now().date(),
        is_published=True
    ).distinct().order_by('event_date', 'start_time')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        events = events.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query)
        )
    
    # Filter by event type
    event_type_filter = request.GET.get('event_type', '')
    if event_type_filter:
        events = events.filter(event_type=event_type_filter)
    
    # Filter by mandatory status
    mandatory_filter = request.GET.get('mandatory', '')
    if mandatory_filter == 'true':
        events = events.filter(is_mandatory=True)
    elif mandatory_filter == 'false':
        events = events.filter(is_mandatory=False)
    
    # Pagination
    paginator = Paginator(events, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get registered events for current student
    registered_events = EventRegistration.objects.filter(
        student=student
    ).values_list('event_id', flat=True)
    
    # Add registration status to events
    for event in page_obj:
        event.is_registered = event.id in registered_events
        event.can_register = event.max_attendees is None or \
                           event.registrations.count() < event.max_attendees
    
    context = {
        'page_obj': page_obj,
        'events': page_obj.object_list,
        'search_query': search_query,
        'event_type_filter': event_type_filter,
        'mandatory_filter': mandatory_filter,
        'total_events': paginator.count,
        'event_type_choices': Event.EVENT_TYPES,
        'registered_events': registered_events,
    }
    
    return render(request, 'student/events/events_list.html', context)


@login_required(login_url='login')
def student_event_detail(request, event_id):
    """Display event detail and handle registration"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    event = get_object_or_404(Event, id=event_id)
    
    # Check if student is registered
    registration = EventRegistration.objects.filter(
        event=event,
        student=student
    ).first()
    
    # Check if event is full
    is_event_full = (event.max_attendees and 
                     event.registrations.count() >= event.max_attendees)
    
    # Check if registration is required but not done
    can_view = not event.registration_required or registration is not None
    
    context = {
        'event': event,
        'registration': registration,
        'is_registered': registration is not None,
        'is_event_full': is_event_full,
        'attendee_count': event.registrations.count(),
        'remaining_slots': max(0, (event.max_attendees or 0) - event.registrations.count()) if event.max_attendees else None,
        'can_view': can_view,
    }
    
    return render(request, 'student/events/event_detail.html', context)


@login_required(login_url='login')
def register_for_event(request, event_id):
    """Register student for an event"""
    if request.method != 'POST':
        return redirect('student_event_detail', event_id=event_id)
    
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    event = get_object_or_404(Event, id=event_id)
    
    # Check if already registered
    if EventRegistration.objects.filter(event=event, student=student).exists():
        messages.warning(request, "You are already registered for this event.")
        return redirect('student_event_detail', event_id=event_id)
    
    # Check if event is full
    if event.max_attendees and event.registrations.count() >= event.max_attendees:
        messages.error(request, "This event is full and no longer accepting registrations.")
        return redirect('student_event_detail', event_id=event_id)
    
    # Create registration
    EventRegistration.objects.create(
        event=event,
        student=student
    )
    
    messages.success(request, f"Successfully registered for {event.title}!")
    return redirect('student_event_detail', event_id=event_id)


@login_required(login_url='login')
def unregister_from_event(request, event_id):
    """Unregister student from an event"""
    if request.method != 'POST':
        return redirect('student_event_detail', event_id=event_id)
    
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    event = get_object_or_404(Event, id=event_id)
    
    registration = EventRegistration.objects.filter(
        event=event,
        student=student
    ).first()
    
    if registration:
        registration.delete()
        messages.success(request, f"Unregistered from {event.title}.")
    else:
        messages.warning(request, "You are not registered for this event.")
    
    return redirect('student_event_detail', event_id=event_id)

# views.py - Student Units Management Views
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Prefetch
from django.utils import timezone
from datetime import timedelta
from .models import (
    Student, Unit, ProgrammeUnit, UnitEnrollment, SemesterRegistration, 
    Semester, AcademicYear
)


@login_required(login_url='login')
def register_units(request):
    """Register units for current semester"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    # Get current semester
    current_semester = Semester.objects.filter(is_current=True).first()
    if not current_semester:
        messages.error(request, "No active semester found.")
        return redirect('student_dashboard')
    
    # Check if student can register
    registration_deadline = current_semester.registration_deadline
    if timezone.now().date() > registration_deadline:
        messages.error(request, f"Registration deadline has passed ({registration_deadline}).")
        return redirect('student_dashboard')
    
    # Get available units for student's programme and year
    available_units = ProgrammeUnit.objects.filter(
        programme=student.programme,
        year_level=student.current_year,
        semester=current_semester.semester_number
    ).select_related('unit')
    
    # Get already enrolled units for this semester
    enrolled_unit_ids = UnitEnrollment.objects.filter(
        student=student,
        semester=current_semester
    ).values_list('unit_id', flat=True)
    
    # Check if semester registration exists
    sem_registration, created = SemesterRegistration.objects.get_or_create(
        student=student,
        semester=current_semester
    )
    
    if request.method == 'POST':
        selected_units = request.POST.getlist('units')
        
        if not selected_units:
            messages.warning(request, "Please select at least one unit.")
            return redirect('register_units')
        
        # Validate that selected units belong to student's programme
        valid_units = ProgrammeUnit.objects.filter(
            id__in=selected_units,
            programme=student.programme,
            year_level=student.current_year,
            semester=current_semester.semester_number
        ).values_list('unit_id', flat=True)
        
        created_count = 0
        for unit_id in valid_units:
            # Check if already enrolled
            if unit_id not in enrolled_unit_ids:
                UnitEnrollment.objects.create(
                    student=student,
                    unit_id=unit_id,
                    semester=current_semester,
                    status='ENROLLED'
                )
                created_count += 1
        
        # Update semester registration
        sem_registration.units_enrolled = UnitEnrollment.objects.filter(
            student=student,
            semester=current_semester
        ).count()
        sem_registration.status = 'REGISTERED'
        sem_registration.save()
        
        if created_count > 0:
            messages.success(request, f"Successfully registered for {created_count} unit(s).")
        else:
            messages.info(request, "Units already registered.")
        
        return redirect('student_enrollments')
    
    # Separate mandatory and elective units
    mandatory_units = available_units.filter(is_mandatory=True)
    elective_units = available_units.filter(is_mandatory=False)
    
    context = {
        'current_semester': current_semester,
        'mandatory_units': mandatory_units,
        'elective_units': elective_units,
        'enrolled_unit_ids': enrolled_unit_ids,
        'registration_deadline': registration_deadline,
    }
    
    return render(request, 'student/units/register_units.html', context)

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
from .models import Student, UnitEnrollment

@login_required(login_url='login')
def student_enrollments(request):
    """View student enrollments organized by academic year and semester"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    # Get all enrollments - ONLY get unique enrollments per semester
    enrollments = UnitEnrollment.objects.filter(
        student=student
    ).select_related(
        'unit', 
        'semester', 
        'semester__academic_year'
    ).order_by(
        '-semester__academic_year__start_date',
        'semester__semester_number',
        'unit__code'
    )
    
    # Debug: Print to console to see what we're getting
    print("\n=== DEBUGGING ENROLLMENTS ===")
    for e in enrollments:
        print(f"Unit: {e.unit.code} | Semester: {e.semester.semester_number} | Year: {e.semester.academic_year.year_code}")
    print("=== END DEBUG ===\n")
    
    # Organize by academic year and semester
    enrollments_by_year = {}
    
    for enrollment in enrollments:
        year_code = enrollment.semester.academic_year.year_code
        sem_num = enrollment.semester.semester_number
        
        # Initialize year structure if needed
        if year_code not in enrollments_by_year:
            enrollments_by_year[year_code] = {
                'academic_year': enrollment.semester.academic_year,
                'semesters': {}
            }
        
        # Initialize semester structure if needed
        if sem_num not in enrollments_by_year[year_code]['semesters']:
            enrollments_by_year[year_code]['semesters'][sem_num] = {
                'semester': enrollment.semester,
                'semester_number': sem_num,  # Explicitly store semester number
                'units': []
            }
        
        # CRITICAL: Only add enrollment to the semester it ACTUALLY belongs to
        # Check if this enrollment's semester matches the current semester we're adding to
        if enrollment.semester.semester_number == sem_num:
            enrollments_by_year[year_code]['semesters'][sem_num]['units'].append(enrollment)
    
    # Calculate registration dates for drop eligibility
    current_date = timezone.now().date()
    for year_code in enrollments_by_year:
        for sem_num in enrollments_by_year[year_code]['semesters']:
            for enrollment in enrollments_by_year[year_code]['semesters'][sem_num]['units']:
                # Handle both datetime and date objects
                if hasattr(enrollment.enrollment_date, 'date'):
                    registration_date = enrollment.enrollment_date.date()
                else:
                    registration_date = enrollment.enrollment_date
                
                drop_eligible_date = registration_date + timedelta(days=7)
                enrollment.can_drop = current_date >= drop_eligible_date
                days_diff = (drop_eligible_date - current_date).days
                enrollment.days_until_drop = max(0, days_diff)  # Ensure non-negative
    
    # Debug: Print final structure
    print("\n=== FINAL STRUCTURE ===")
    for year_code, year_data in enrollments_by_year.items():
        print(f"Year: {year_code}")
        for sem_num, sem_data in year_data['semesters'].items():
            print(f"  Semester {sem_num}: {len(sem_data['units'])} units")
            for e in sem_data['units']:
                print(f"    - {e.unit.code}")
    print("=== END STRUCTURE ===\n")
    
    context = {
        'enrollments_by_year': enrollments_by_year,
        'total_enrollments': enrollments.count(),
    }
    
    return render(request, 'student/units/student_enrollments.html', context)


@login_required(login_url='login')
def drop_unit(request, enrollment_id):
    """Drop a unit (after 7 days from registration)"""
    if request.method != 'POST':
        return redirect('student_enrollments')
    
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    enrollment = get_object_or_404(
        UnitEnrollment,
        id=enrollment_id,
        student=student
    )
    
    # Check if within 7 days
    registration_date = enrollment.enrollment_date.date()
    current_date = timezone.now().date()
    drop_eligible_date = registration_date + timedelta(days=7)
    
    if current_date < drop_eligible_date:
        days_remaining = (drop_eligible_date - current_date).days
        messages.error(
            request,
            f"You can only drop this unit after {days_remaining} more day(s) "
            f"(eligible from {drop_eligible_date})."
        )
        return redirect('student_enrollments')
    
    unit_name = enrollment.unit.name
    unit_code = enrollment.unit.code
    
    # Mark as dropped
    enrollment.status = 'DROPPED'
    enrollment.save()
    
    # Update semester registration count
    sem_registration = SemesterRegistration.objects.filter(
        student=student,
        semester=enrollment.semester
    ).first()
    
    if sem_registration:
        sem_registration.units_enrolled = UnitEnrollment.objects.filter(
            student=student,
            semester=enrollment.semester,
            status__in=['ENROLLED', 'COMPLETED']
        ).count()
        sem_registration.save()
    
    messages.success(request, f"Successfully dropped {unit_code} - {unit_name}.")
    return redirect('student_enrollments')


# views.py - Student Programme Curriculum View
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import Student, ProgrammeUnit, Unit


@login_required(login_url='login')
def my_programme(request):
    """View the complete curriculum of student's programme"""
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    programme = student.programme
    
    # Get all programme units organized by year and semester
    programme_units = ProgrammeUnit.objects.filter(
        programme=programme
    ).select_related('unit').order_by('year_level', 'semester')
    
    # Organize by year and semester
    curriculum_by_year = {}
    for program_unit in programme_units:
        year_level = program_unit.year_level
        semester = program_unit.semester
        
        if year_level not in curriculum_by_year:
            curriculum_by_year[year_level] = {
                'year_label': f'Year {year_level}',
                'semesters': {}
            }
        
        if semester not in curriculum_by_year[year_level]['semesters']:
            curriculum_by_year[year_level]['semesters'][semester] = {
                'semester_label': f'Semester {semester}',
                'mandatory_units': [],
                'elective_units': [],
                'total_credits': 0
            }
        
        semester_data = curriculum_by_year[year_level]['semesters'][semester]
        
        if program_unit.is_mandatory:
            semester_data['mandatory_units'].append(program_unit)
        else:
            semester_data['elective_units'].append(program_unit)
        
        semester_data['total_credits'] += program_unit.unit.credit_hours
    
    # Calculate programme statistics
    total_units = programme_units.count()
    total_credits = sum(pu.unit.credit_hours for pu in programme_units)
    mandatory_units = programme_units.filter(is_mandatory=True).count()
    elective_units = programme_units.filter(is_mandatory=False).count()
    
    context = {
        'programme': programme,
        'curriculum_by_year': curriculum_by_year,
        'total_units': total_units,
        'total_credits': total_credits,
        'mandatory_units': mandatory_units,
        'elective_units': elective_units,
        'programme_duration': programme.duration_years,
        'semesters_per_year': programme.semesters_per_year,
    }
    
    return render(request, 'student/programme/view_curriculum.html', context)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.db import transaction
from .models import Student, User
from .forms import StudentProfileForm, UserProfileForm

@login_required
def student_profile_view(request):
    """
    Combined view for viewing and editing student profile
    """
    # Ensure user is a student
    if request.user.user_type != 'STUDENT':
        messages.error(request, "Access denied. This page is only for students.")
        return redirect('student_dashboard')
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    # Handle form submission
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_profile':
            user_form = UserProfileForm(request.POST, request.FILES, instance=request.user)
            student_form = StudentProfileForm(request.POST, instance=student)
            
            if user_form.is_valid() and student_form.is_valid():
                try:
                    with transaction.atomic():
                        user_form.save()
                        student_form.save()
                    messages.success(request, "Profile updated successfully!")
                    return redirect('student_profile')
                except Exception as e:
                    messages.error(request, f"Error updating profile: {str(e)}")
            else:
                messages.error(request, "Please correct the errors below.")
        
        elif action == 'change_password':
            password_form = PasswordChangeForm(request.user, request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)  # Keep user logged in
                messages.success(request, "Your password was successfully updated!")
                return redirect('student_profile')
            else:
                messages.error(request, "Please correct the password errors below.")
                # Pass the form with errors to the context
                context = {
                    'student': student,
                    'user_form': UserProfileForm(instance=request.user),
                    'student_form': StudentProfileForm(instance=student),
                    'password_form': password_form,
                    'show_password_modal': True
                }
                return render(request, 'student/profile.html', context)
    
    
    # GET request - display forms
    user_form = UserProfileForm(instance=request.user)
    student_form = StudentProfileForm(instance=student)
    password_form = PasswordChangeForm(request.user)
    
    context = {
        'student': student,
        'user_form': user_form,
        'student_form': student_form,
        'password_form': password_form,
        'show_password_modal': False
    }
    
    return render(request, 'student/profile.html', context)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from django.db import transaction
from .models import Student, Semester, SemesterRegistration, AcademicYear

@login_required
def student_semester_reporting(request):
    """
    View for students to report for the current semester
    """
    # Ensure user is a student
    if request.user.user_type != 'STUDENT':
        messages.error(request, "Access denied. This page is only for students.")
        return redirect('student_dashboard')
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    # Get current semester
    try:
        current_semester = Semester.objects.get(is_current=True)
        current_academic_year = current_semester.academic_year
    except Semester.DoesNotExist:
        current_semester = None
        current_academic_year = None
    
    # Check if student's programme allows reporting
    can_report = False
    is_on_holiday = False
    holiday_message = None
    
    if current_semester and student.programme.semesters_per_year == 2:
        # Programme has 2 semesters per year
        if current_semester.semester_number in [1, 2]:
            can_report = True
        elif current_semester.semester_number == 3:
            is_on_holiday = True
            holiday_message = f"Your programme ({student.programme.name}) has 2 semesters per academic year. Semester 3 is a holiday period for your programme."
    elif current_semester and student.programme.semesters_per_year == 3:
        # Programme has 3 semesters per year - can always report
        can_report = True
    
    # Check if already registered for current semester
    already_registered = False
    registration = None
    if current_semester:
        registration = SemesterRegistration.objects.filter(
            student=student,
            semester=current_semester
        ).first()
        
        if registration:
            already_registered = True
    
    # Check registration deadline
    registration_open = False
    deadline_passed = False
    if current_semester and can_report:
        today = timezone.now().date()
        if today <= current_semester.registration_deadline:
            registration_open = True
        else:
            deadline_passed = True
    
    # Handle form submission
    if request.method == 'POST' and can_report and registration_open and not already_registered:
        action = request.POST.get('action')
        
        if action == 'report_semester':
            try:
                with transaction.atomic():
                    # Create semester registration
                    registration = SemesterRegistration.objects.create(
                        student=student,
                        semester=current_semester,
                        status='REGISTERED',
                        units_enrolled=0  # Will be updated during unit enrollment
                    )
                    
                    messages.success(
                        request, 
                        f"Successfully reported for {current_semester}! You can now proceed to enroll in units."
                    )
                    return redirect('student_semester_reporting')
                    
            except Exception as e:
                messages.error(request, f"Error reporting for semester: {str(e)}")
    
    # Get previous registrations
    previous_registrations = SemesterRegistration.objects.filter(
        student=student
    ).select_related('semester', 'semester__academic_year').order_by('-registration_date')[:5]
    
    context = {
        'student': student,
        'current_semester': current_semester,
        'current_academic_year': current_academic_year,
        'can_report': can_report,
        'is_on_holiday': is_on_holiday,
        'holiday_message': holiday_message,
        'already_registered': already_registered,
        'registration': registration,
        'registration_open': registration_open,
        'deadline_passed': deadline_passed,
        'previous_registrations': previous_registrations,
    }
    
    return render(request, 'student/semester_reporting.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils import timezone
from datetime import datetime, timedelta
import calendar
from .models import (
    Student, Event, Announcement, Semester, 
    AcademicYear, TimetableSlot
)

@login_required
def student_academic_calendar(request):
    """
    Academic calendar view showing events, announcements, and important dates
    """
    # Ensure user is a student
    if request.user.user_type != 'STUDENT':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        return JsonResponse({'error': 'Student profile not found'}, status=404)
    
    # Get current date or requested month/year
    today = timezone.now().date()
    
    # Get month and year from query params or use current
    try:
        month = int(request.GET.get('month', today.month))
        year = int(request.GET.get('year', today.year))
    except (ValueError, TypeError):
        month = today.month
        year = today.year
    
    # Validate month and year
    if month < 1 or month > 12:
        month = today.month
    if year < 2000 or year > 2100:
        year = today.year
    
    # Get calendar data
    cal = calendar.monthcalendar(year, month)
    month_name = calendar.month_name[month]
    
    # Get date range for the month
    first_day = datetime(year, month, 1).date()
    if month == 12:
        last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Get events for this month targeting student's programme
    events = Event.objects.filter(
        event_date__gte=first_day,
        event_date__lte=last_day,
        is_published=True
    ).filter(
        target_programmes=student.programme
    ).select_related('venue', 'organizer')
    
    # Get announcements published in this month
    announcements = Announcement.objects.filter(
        publish_date__date__gte=first_day,
        publish_date__date__lte=last_day,
        is_published=True
    ).filter(
        target_programmes=student.programme
    ).select_related('created_by')
    
    # Get semesters
    semesters = Semester.objects.filter(
        start_date__lte=last_day,
        end_date__gte=first_day
    ).select_related('academic_year')
    
    # Get timetable slots for the month
    timetable_slots = TimetableSlot.objects.filter(
        programme=student.programme,
        year_level=student.current_year,
        is_active=True,
        unit_allocation__semester__start_date__lte=last_day,
        unit_allocation__semester__end_date__gte=first_day
    ).select_related(
        'unit_allocation__unit',
        'unit_allocation__lecturer__user',
        'venue'
    )
    
    # Organize data by date
    calendar_data = {}
    
    # Add events
    for event in events:
        date_key = event.event_date.strftime('%Y-%m-%d')
        if date_key not in calendar_data:
            calendar_data[date_key] = {
                'events': [],
                'announcements': [],
                'semesters': [],
                'classes': []
            }
        calendar_data[date_key]['events'].append({
            'id': event.id,
            'title': event.title,
            'type': event.event_type,
            'time': event.start_time.strftime('%H:%M'),
            'venue': event.venue.name if event.venue else 'TBA',
            'is_mandatory': event.is_mandatory,
            'icon': get_event_icon(event.event_type)
        })
    
    # Add announcements
    for announcement in announcements:
        date_key = announcement.publish_date.date().strftime('%Y-%m-%d')
        if date_key not in calendar_data:
            calendar_data[date_key] = {
                'events': [],
                'announcements': [],
                'semesters': [],
                'classes': []
            }
        calendar_data[date_key]['announcements'].append({
            'id': announcement.id,
            'title': announcement.title,
            'priority': announcement.priority,
            'author': announcement.created_by.get_full_name()
        })
    
    # Add semester start/end dates
    for semester in semesters:
        # Start date
        if first_day <= semester.start_date <= last_day:
            date_key = semester.start_date.strftime('%Y-%m-%d')
            if date_key not in calendar_data:
                calendar_data[date_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'classes': []
                }
            calendar_data[date_key]['semesters'].append({
                'title': f'{semester} - Starts',
                'type': 'start',
                'semester': str(semester)
            })
        
        # End date
        if first_day <= semester.end_date <= last_day:
            date_key = semester.end_date.strftime('%Y-%m-%d')
            if date_key not in calendar_data:
                calendar_data[date_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'classes': []
                }
            calendar_data[date_key]['semesters'].append({
                'title': f'{semester} - Ends',
                'type': 'end',
                'semester': str(semester)
            })
        
        # Registration deadline
        if first_day <= semester.registration_deadline <= last_day:
            date_key = semester.registration_deadline.strftime('%Y-%m-%d')
            if date_key not in calendar_data:
                calendar_data[date_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'classes': []
                }
            calendar_data[date_key]['semesters'].append({
                'title': f'{semester} - Registration Deadline',
                'type': 'deadline',
                'semester': str(semester)
            })
    
    # Add class schedule summary (count of classes per day)
    for slot in timetable_slots:
        # Get all dates in the month that match this day of week
        for week in cal:
            for day in week:
                if day == 0:  # Skip empty days
                    continue
                date_obj = datetime(year, month, day).date()
                day_name = date_obj.strftime('%A').upper()
                
                if day_name == slot.day_of_week:
                    date_key = date_obj.strftime('%Y-%m-%d')
                    if date_key not in calendar_data:
                        calendar_data[date_key] = {
                            'events': [],
                            'announcements': [],
                            'semesters': [],
                            'classes': []
                        }
                    calendar_data[date_key]['classes'].append({
                        'unit': slot.unit_allocation.unit.code,
                        'time': slot.start_time.strftime('%H:%M'),
                        'venue': slot.venue.code
                    })
    
    # Calculate previous and next month
    if month == 1:
        prev_month = 12
        prev_year = year - 1
    else:
        prev_month = month - 1
        prev_year = year
    
    if month == 12:
        next_month = 1
        next_year = year + 1
    else:
        next_month = month + 1
        next_year = year
    
    context = {
        'student': student,
        'calendar': cal,
        'month': month,
        'year': year,
        'month_name': month_name,
        'today': today,
        'calendar_data': calendar_data,
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
    }
    
    return render(request, 'student/academic_calendar.html', context)


def get_event_icon(event_type):
    """Return Bootstrap icon class for event type"""
    icons = {
        'SEMINAR': 'bi-people',
        'WORKSHOP': 'bi-tools',
        'CONFERENCE': 'bi-briefcase',
        'MEETING': 'bi-person-video3',
        'ORIENTATION': 'bi-compass',
        'EXAMINATION': 'bi-clipboard-check',
        'OTHER': 'bi-calendar-event',
    }
    return icons.get(event_type, 'bi-calendar-event')

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q
from datetime import datetime, time
from collections import defaultdict
from .models import Student, TimetableSlot, Semester

@login_required
def student_timetable_view(request):
    """
    View for students to see their personalized timetable
    """
    # Ensure user is a student
    if request.user.user_type != 'STUDENT':
        messages.error(request, "Access denied. This page is only for students.")
        return redirect('student_dashboard')
    
    try:
        student = request.user.student_profile
    except Student.DoesNotExist:
        messages.error(request, "Student profile not found.")
        return redirect('student_dashboard')
    
    # Get current semester
    try:
        current_semester = Semester.objects.get(is_current=True)
    except Semester.DoesNotExist:
        current_semester = None
    
    # Get timetable slots for the student's programme and year level
    timetable_slots = []
    if current_semester:
        timetable_slots = TimetableSlot.objects.filter(
            programme=student.programme,
            year_level=student.current_year,
            unit_allocation__semester=current_semester,
            is_active=True
        ).select_related(
            'unit_allocation__unit',
            'unit_allocation__lecturer__user',
            'venue'
        ).order_by('day_of_week', 'start_time')
    
    # Define days of the week in order
    days_order = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
    
    # Organize slots by day and time
    timetable_by_day = defaultdict(list)
    all_time_slots = set()
    
    for slot in timetable_slots:
        timetable_by_day[slot.day_of_week].append(slot)
        # Create time slot key (e.g., "08:00-10:00")
        time_key = f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"
        all_time_slots.add((slot.start_time, slot.end_time, time_key))
    
    # Sort time slots
    sorted_time_slots = sorted(list(all_time_slots), key=lambda x: x[0])
    
    # Create a structured timetable grid
    timetable_grid = {}
    for day in days_order:
        timetable_grid[day] = {}
        for start_time, end_time, time_key in sorted_time_slots:
            timetable_grid[day][time_key] = None
    
    # Fill in the timetable grid
    for slot in timetable_slots:
        time_key = f"{slot.start_time.strftime('%H:%M')}-{slot.end_time.strftime('%H:%M')}"
        if slot.day_of_week in timetable_grid and time_key in timetable_grid[slot.day_of_week]:
            timetable_grid[slot.day_of_week][time_key] = {
                'unit_code': slot.unit_allocation.unit.code,
                'unit_name': slot.unit_allocation.unit.name,
                'lecturer': slot.unit_allocation.lecturer.user.get_full_name(),
                'venue': slot.venue.code,
                'venue_name': slot.venue.name,
                'start_time': slot.start_time.strftime('%H:%M'),
                'end_time': slot.end_time.strftime('%H:%M'),
            }
    
    # Calculate statistics
    total_classes = len(timetable_slots)
    unique_units = len(set([slot.unit_allocation.unit for slot in timetable_slots]))
    
    # Count classes per day
    classes_per_day = {}
    for day in days_order:
        classes_per_day[day] = len(timetable_by_day.get(day, []))
    
    context = {
        'student': student,
        'current_semester': current_semester,
        'timetable_slots': timetable_slots,
        'timetable_grid': timetable_grid,
        'days_order': days_order,
        'sorted_time_slots': sorted_time_slots,
        'total_classes': total_classes,
        'unique_units': unique_units,
        'classes_per_day': classes_per_day,
    }
    
    return render(request, 'student/timetable.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from django.views.decorators.http import require_http_methods
import json
import traceback

from .models import User, Message, MessageReadStatus

# ========================
# MESSAGING VIEWS
# ========================

@login_required
def messaging_list(request):
    """
    Main messaging page - shows conversation list and chat area
    """
    user = request.user
    
    try:
        # Get all messages involving this user
        all_messages = Message.objects.filter(
            Q(sender=user) | Q(recipients=user)
        ).select_related('sender').prefetch_related('recipients').order_by('-sent_at')
        
        # Get unique conversation partners manually (SQLite compatible)
        conversation_partners = {}
        for msg in all_messages:
            if msg.sender == user:
                partner = msg.recipients.first()
            else:
                partner = msg.sender
            
            if partner and partner.id not in conversation_partners:
                conversation_partners[partner.id] = {
                    'user': partner,
                    'last_message': msg,
                }
        
        # Build conversation list with unread counts
        conversation_list = []
        for partner_id, conv_data in conversation_partners.items():
            partner = conv_data['user']
            last_msg = conv_data['last_message']
            
            unread_count = MessageReadStatus.objects.filter(
                message__sender=partner,
                recipient=user,
                is_read=False
            ).count()
            
            conversation_list.append({
                'user': partner,
                'last_message': last_msg,
                'unread_count': unread_count,
            })
        
        conversation_list.sort(key=lambda x: x['last_message'].sent_at, reverse=True)
        
        context = {
            'conversations': conversation_list,
            'user_type': user.user_type,
        }
        return render(request, 'messaging/message_list.html', context)
    
    except Exception as e:
        print(f"Error in messaging_list: {str(e)}")
        traceback.print_exc()
        return render(request, 'messaging/message_list.html', {'conversations': [], 'error': str(e)})


@login_required
def message_thread(request, user_id):
    """
    Get message thread between current user and specified user
    """
    current_user = request.user
    
    try:
        other_user = get_object_or_404(User, id=user_id)
        
        if current_user == other_user:
            return JsonResponse({'error': 'Cannot message yourself'}, status=400)
        
        # Get all messages between these two users
        messages = Message.objects.filter(
            Q(sender=current_user, recipients=other_user) |
            Q(sender=other_user, recipients=current_user)
        ).select_related('sender').order_by('sent_at')
        
        # Mark messages as read
        MessageReadStatus.objects.filter(
            message__sender=other_user,
            recipient=current_user,
            is_read=False
        ).update(is_read=True, read_at=timezone.now())
        
        # Prepare message data - ensure all values are strings/primitives
        message_data = []
        for msg in messages:
            read_status = MessageReadStatus.objects.filter(
                message=msg,
                recipient=current_user
            ).first()
            
            # Convert datetime to ISO string
            sent_at = msg.sent_at.isoformat() if hasattr(msg.sent_at, 'isoformat') else str(msg.sent_at)
            
            message_data.append({
                'id': int(msg.id),
                'sender': str(msg.sender.get_full_name() or msg.sender.username),
                'sender_id': int(msg.sender.id),
                'body': str(msg.body),
                'sent_at': sent_at,
                'is_own': bool(msg.sender == current_user),
                'is_read': bool(read_status.is_read if read_status else False),
            })
        
        # Build profile info - ensure all strings
        registration_number = ''
        programme = ''
        staff_number = ''
        department = ''
        
        try:
            if other_user.user_type == 'STUDENT' and hasattr(other_user, 'student_profile'):
                registration_number = str(other_user.student_profile.registration_number)
                programme = str(other_user.student_profile.programme)
        except Exception as e:
            print(f"Error getting student profile: {str(e)}")
        
        try:
            if other_user.user_type == 'LECTURER' and hasattr(other_user, 'lecturer_profile'):
                staff_number = str(other_user.lecturer_profile.staff_number)
                department = str(other_user.lecturer_profile.department)
        except Exception as e:
            print(f"Error getting lecturer profile: {str(e)}")
        
        profile_info = {
            'id': int(other_user.id),
            'name': str(other_user.get_full_name() or other_user.username),
            'email': str(other_user.email or ''),
            'phone': str(other_user.phone_number or ''),
            'user_type': str(other_user.get_user_type_display() or 'User'),
            'registration_number': registration_number,
            'programme': programme,
            'staff_number': staff_number,
            'department': department,
        }
        
        return JsonResponse({
            'success': True,
            'other_user': {
                'id': int(other_user.id),
                'name': str(other_user.get_full_name() or other_user.username),
            },
            'profile_info': profile_info,
            'messages': message_data,
        })
    
    except User.DoesNotExist:
        return JsonResponse({'error': 'User not found'}, status=404)
    except Exception as e:
        print(f"Error in message_thread: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
@login_required
@require_http_methods(["POST"])
def send_message(request, user_id):
    """
    Send a message to a user (AJAX)
    """
    current_user = request.user
    
    try:
        recipient = get_object_or_404(User, id=user_id)
        
        if current_user == recipient:
            return JsonResponse({'error': 'Cannot message yourself'}, status=400)
        
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        
        body = data.get('message', '').strip()
        
        if not body:
            return JsonResponse({'error': 'Message cannot be empty'}, status=400)
        
        if len(body) > 5000:
            return JsonResponse({'error': 'Message is too long'}, status=400)
        
        # Create message
        message = Message.objects.create(
            sender=current_user,
            subject='Direct Message',
            body=body,
            message_type='DIRECT'
        )
        message.recipients.add(recipient)
        
        # Create read status
        MessageReadStatus.objects.create(
            message=message,
            recipient=recipient,
            is_read=False
        )
        
        sent_at = message.sent_at.isoformat() if hasattr(message.sent_at, 'isoformat') else str(message.sent_at)
        
        return JsonResponse({
            'success': True,
            'message': {
                'id': int(message.id),
                'sender': str(current_user.get_full_name() or current_user.username),
                'sender_id': int(current_user.id),
                'body': str(message.body),
                'sent_at': sent_at,
                'is_own': True,
                'is_read': False,
            }
        })
    
    except Exception as e:
        print(f"Error in send_message: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def search_users(request):
    """
    Search for users to message (AJAX)
    """
    query = request.GET.get('q', '').strip()
    current_user = request.user
    
    try:
        if not query or len(query) < 2:
            return JsonResponse({'users': []})
        
        # Search by name and email
        users = User.objects.filter(
            ~Q(id=current_user.id),
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        ).select_related('student_profile', 'lecturer_profile').distinct()[:10]
        
        # Also search by registration/staff number
        try:
            student_users = User.objects.filter(
                ~Q(id=current_user.id),
                student_profile__registration_number__icontains=query
            ).select_related('student_profile')[:5]
            
            lecturer_users = User.objects.filter(
                ~Q(id=current_user.id),
                lecturer_profile__staff_number__icontains=query
            ).select_related('lecturer_profile')[:5]
            
            user_ids = set()
            combined_users = []
            
            for u in users:
                if u.id not in user_ids:
                    combined_users.append(u)
                    user_ids.add(u.id)
            
            for u in student_users:
                if u.id not in user_ids:
                    combined_users.append(u)
                    user_ids.add(u.id)
                    
            for u in lecturer_users:
                if u.id not in user_ids:
                    combined_users.append(u)
                    user_ids.add(u.id)
            
            users = combined_users[:10]
        except Exception as e:
            print(f"Error searching by ID: {str(e)}")
        
        users_data = []
        for user in users:
            identifier = ''
            
            try:
                if user.user_type == 'STUDENT' and hasattr(user, 'student_profile'):
                    identifier = str(user.student_profile.registration_number)
                elif user.user_type == 'LECTURER' and hasattr(user, 'lecturer_profile'):
                    identifier = str(user.lecturer_profile.staff_number)
            except Exception as e:
                print(f"Error getting identifier: {str(e)}")
            
            user_info = {
                'id': int(user.id),
                'name': str(user.get_full_name() or user.username),
                'email': str(user.email or ''),
                'user_type': str(user.get_user_type_display() or 'User'),
                'type_code': str(user.user_type or ''),
                'identifier': identifier,
            }
            
            users_data.append(user_info)
        
        return JsonResponse({'users': users_data})
    
    except Exception as e:
        print(f"Error in search_users: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def mark_as_read(request, message_id):
    """
    Mark a message as read (AJAX)
    """
    try:
        message = get_object_or_404(Message, id=message_id)
        
        read_status = MessageReadStatus.objects.filter(
            message=message,
            recipient=request.user
        ).first()
        
        if read_status:
            read_status.is_read = True
            read_status.read_at = timezone.now()
            read_status.save()
            return JsonResponse({'success': True})
        
        return JsonResponse({'error': 'Read status not found'}, status=404)
    except Exception as e:
        print(f"Error in mark_as_read: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_unread_count(request):
    """
    Get total unread message count (AJAX)
    """
    try:
        unread_count = MessageReadStatus.objects.filter(
            recipient=request.user,
            is_read=False
        ).count()
        
        return JsonResponse({'unread_count': int(unread_count)})
    except Exception as e:
        print(f"Error in get_unread_count: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def delete_conversation(request, user_id):
    """
    Delete a conversation (AJAX)
    """
    try:
        current_user = request.user
        other_user = get_object_or_404(User, id=user_id)
        
        # Delete messages between these users
        Message.objects.filter(
            Q(sender=current_user, recipients=other_user) |
            Q(sender=other_user, recipients=current_user)
        ).delete()
        
        return JsonResponse({'success': True})
    except Exception as e:
        print(f"Error in delete_conversation: {str(e)}")
        traceback.print_exc()
        return JsonResponse({'error': str(e)}, status=500)



from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, Count, Q, F
from decimal import Decimal
from collections import defaultdict
from .models import (
    Student, UnitEnrollment, ProgrammeUnit, 
    Semester, StudentMarks, FinalGrade
)

@login_required
def student_grades_view(request):
    """View for displaying student grades organized by year and semester"""
    
    # Get student profile
    try:
        student = request.user.student_profile
    except:
        return render(request, 'error.html', {'message': 'Student profile not found'})
    
    # Get all enrollments with final grades
    enrollments = UnitEnrollment.objects.filter(
        student=student,
        status='COMPLETED'
    ).select_related(
        'unit',
        'semester',
        'semester__academic_year',
        'final_grade'
    ).prefetch_related(
        'marks__assessment_component'
    ).order_by(
        'semester__academic_year__start_date',
        'semester__semester_number'
    )
    
    # Organize grades by year and semester
    grades_by_year = defaultdict(lambda: {
        'year_label': '',
        'semesters': defaultdict(lambda: {
            'semester_label': '',
            'academic_year': '',
            'units': [],
            'total_units': 0,
            'total_credits': 0,
            'semester_gpa': Decimal('0.00'),
            'graded_units': 0
        })
    })
    
    total_credits_earned = 0
    total_grade_points = Decimal('0.00')
    total_units_completed = 0
    
    for enrollment in enrollments:
        semester = enrollment.semester
        academic_year = semester.academic_year
        semester_num = semester.semester_number
        
        # Get the correct year level from ProgrammeUnit
        # Filter by student's programme to avoid multiple results
        try:
            program_unit = ProgrammeUnit.objects.filter(
                programme=student.programme,
                unit=enrollment.unit,
                semester=semester_num
            ).first()  # Use .first() instead of .get()
            
            if program_unit:
                year_level = program_unit.year_level
            else:
                # Try without semester constraint
                program_unit = ProgrammeUnit.objects.filter(
                    programme=student.programme,
                    unit=enrollment.unit
                ).first()
                year_level = program_unit.year_level if program_unit else 1
        except Exception as e:
            # Fallback to Year 1 if any error occurs
            year_level = 1
        
        # Get assessment breakdown
        assessment_breakdown = []
        total_marks = Decimal('0.00')
        
        marks = enrollment.marks.all().select_related('assessment_component')
        for mark in marks:
            component = mark.assessment_component
            # Calculate weighted marks
            if component.max_marks > 0:
                weighted_marks = (mark.marks_obtained / component.max_marks) * component.weight_percentage
            else:
                weighted_marks = Decimal('0.00')
            
            total_marks += weighted_marks
            
            assessment_breakdown.append({
                'name': component.name,
                'type': component.get_component_type_display(),
                'weight': component.weight_percentage,
                'max_marks': component.max_marks,
                'marks_obtained': mark.marks_obtained,
                'weighted_marks': weighted_marks,
                'percentage': (mark.marks_obtained / component.max_marks) * 100 if component.max_marks > 0 else 0
            })
        
        # Get final grade
        unit_data = {
            'unit_code': enrollment.unit.code,
            'unit_name': enrollment.unit.name,
            'credit_hours': enrollment.unit.credit_hours,
            'assessment_breakdown': assessment_breakdown,
            'total_marks': total_marks,
            'grade': None,
            'grade_point': None,
            'has_grade': False
        }
        
        if hasattr(enrollment, 'final_grade'):
            final_grade = enrollment.final_grade
            unit_data.update({
                'grade': final_grade.grade,
                'grade_point': final_grade.grade_point,
                'total_marks': final_grade.total_marks,
                'has_grade': True
            })
            
            # Calculate totals for GPA
            if final_grade.grade_point > 0:
                total_credits_earned += enrollment.unit.credit_hours
                total_grade_points += (final_grade.grade_point * enrollment.unit.credit_hours)
                total_units_completed += 1
        
        # Add to structure based on ACTUAL year and semester
        year_data = grades_by_year[year_level]
        year_data['year_label'] = f'Year {year_level}'
        
        semester_data = year_data['semesters'][semester_num]
        semester_data['semester_label'] = f'Semester {semester_num}'
        semester_data['academic_year'] = academic_year.year_code
        semester_data['units'].append(unit_data)
        semester_data['total_units'] += 1
        semester_data['total_credits'] += enrollment.unit.credit_hours
        
        if unit_data['has_grade']:
            semester_data['graded_units'] += 1
    
    # Calculate semester GPAs
    for year_level, year_data in grades_by_year.items():
        for semester_num, semester_data in year_data['semesters'].items():
            semester_credits = Decimal('0.00')
            semester_points = Decimal('0.00')
            
            for unit in semester_data['units']:
                if unit['has_grade'] and unit['grade_point']:
                    semester_credits += unit['credit_hours']
                    semester_points += (unit['grade_point'] * unit['credit_hours'])
            
            if semester_credits > 0:
                semester_data['semester_gpa'] = round(semester_points / semester_credits, 2)
    
    # Calculate cumulative GPA
    cumulative_gpa = Decimal('0.00')
    if total_credits_earned > 0:
        cumulative_gpa = round(total_grade_points / total_credits_earned, 2)
    
    # Sort the data properly - Year 1 first, then Year 2, etc.
    grades_by_year = dict(sorted(grades_by_year.items(), key=lambda x: x[0]))
    
    # Sort semesters within each year - Semester 1 first, then 2, then 3
    for year_level in grades_by_year:
        grades_by_year[year_level]['semesters'] = dict(
            sorted(grades_by_year[year_level]['semesters'].items(), key=lambda x: x[0])
        )
    
    context = {
        'student': student,
        'programme': student.programme,
        'grades_by_year': grades_by_year,
        'cumulative_gpa': cumulative_gpa,
        'total_credits_earned': total_credits_earned,
        'total_units_completed': total_units_completed,
        'current_year': student.current_year,
    }
    
    return render(request, 'student/my_grades.html', context)


@login_required
def academic_performance_view(request):
    """View for displaying overall academic performance with classification"""
    
    # Get student profile
    try:
        student = request.user.student_profile
    except:
        return render(request, 'error.html', {'message': 'Student profile not found'})
    
    # Get all completed enrollments with grades
    enrollments = UnitEnrollment.objects.filter(
        student=student,
        status='COMPLETED',
        final_grade__isnull=False
    ).select_related(
        'unit',
        'semester',
        'semester__academic_year',
        'final_grade'
    ).order_by(
        'semester__academic_year__start_date',
        'semester__semester_number'
    )
    
    # Calculate overall statistics
    total_credits_earned = 0
    total_grade_points = Decimal('0.00')
    total_units = 0
    grade_distribution = defaultdict(int)
    
    # Performance by year
    performance_by_year = defaultdict(lambda: {
        'year_label': '',
        'semesters': [],
        'total_credits': 0,
        'total_grade_points': Decimal('0.00'),
        'year_gpa': Decimal('0.00'),
        'units_count': 0
    })
    
    # Performance by semester
    semester_performance = []
    
    current_semester_data = {}
    
    for enrollment in enrollments:
        final_grade = enrollment.final_grade
        semester = enrollment.semester
        
        # Get year level
        try:
            program_unit = ProgrammeUnit.objects.get(
                programme=student.programme,
                unit=enrollment.unit
            )
            year_level = program_unit.year_level
        except:
            year_level = student.current_year
        
        credits = enrollment.unit.credit_hours
        grade_point = final_grade.grade_point
        
        # Overall totals
        total_credits_earned += credits
        total_grade_points += (grade_point * credits)
        total_units += 1
        grade_distribution[final_grade.grade] += 1
        
        # Year performance
        year_data = performance_by_year[year_level]
        year_data['year_label'] = f'Year {year_level}'
        year_data['total_credits'] += credits
        year_data['total_grade_points'] += (grade_point * credits)
        year_data['units_count'] += 1
        
        # Semester data
        semester_key = f"{semester.academic_year.year_code}-{semester.semester_number}"
        if semester_key not in current_semester_data:
            current_semester_data[semester_key] = {
                'semester': semester,
                'academic_year': semester.academic_year.year_code,
                'semester_number': semester.semester_number,
                'semester_label': f'Semester {semester.semester_number}',
                'year_level': year_level,
                'credits': 0,
                'grade_points': Decimal('0.00'),
                'gpa': Decimal('0.00'),
                'units_count': 0,
                'grades': []
            }
        
        sem_data = current_semester_data[semester_key]
        sem_data['credits'] += credits
        sem_data['grade_points'] += (grade_point * credits)
        sem_data['units_count'] += 1
        sem_data['grades'].append({
            'unit_code': enrollment.unit.code,
            'unit_name': enrollment.unit.name,
            'grade': final_grade.grade,
            'grade_point': grade_point,
            'credits': credits
        })
    
    # Calculate GPAs
    cumulative_gpa = Decimal('0.00')
    if total_credits_earned > 0:
        cumulative_gpa = round(total_grade_points / total_credits_earned, 2)
    
    # Calculate year GPAs
    for year_level, year_data in performance_by_year.items():
        if year_data['total_credits'] > 0:
            year_data['year_gpa'] = round(
                year_data['total_grade_points'] / year_data['total_credits'],
                2
            )
    
    # Calculate semester GPAs
    for sem_data in current_semester_data.values():
        if sem_data['credits'] > 0:
            sem_data['gpa'] = round(sem_data['grade_points'] / sem_data['credits'], 2)
        semester_performance.append(sem_data)
    
    # Sort semester performance by date
    semester_performance.sort(
        key=lambda x: (x['semester'].academic_year.start_date, x['semester_number'])
    )
    
    # Group semesters by year for display
    for sem_data in semester_performance:
        year_level = sem_data['year_level']
        performance_by_year[year_level]['semesters'].append(sem_data)
    
    # Determine degree classification
    classification = get_degree_classification(cumulative_gpa, student.programme.level)
    
    # Calculate progression metrics
    credits_required = student.programme.duration_years * 30  # Assuming 30 credits per year
    progress_percentage = (total_credits_earned / credits_required * 100) if credits_required > 0 else 0
    
    # Grade distribution for chart
    grade_dist_list = [
        {'grade': grade, 'count': count}
        for grade, count in sorted(grade_distribution.items())
    ]
    
    # Sort performance by year
    performance_by_year = dict(sorted(performance_by_year.items()))
    
    context = {
        'student': student,
        'programme': student.programme,
        'cumulative_gpa': cumulative_gpa,
        'classification': classification,
        'total_credits_earned': total_credits_earned,
        'total_units': total_units,
        'credits_required': credits_required,
        'progress_percentage': round(progress_percentage, 1),
        'performance_by_year': performance_by_year,
        'semester_performance': semester_performance,
        'grade_distribution': grade_dist_list,
        'highest_gpa_semester': max(semester_performance, key=lambda x: x['gpa']) if semester_performance else None,
        'lowest_gpa_semester': min(semester_performance, key=lambda x: x['gpa']) if semester_performance else None,
    }
    
    return render(request, 'student/academic_performance.html', context)


def get_degree_classification(gpa, programme_level):
    """Determine degree classification based on GPA and programme level"""
    
    # Only classify for Bachelor's degree
    if programme_level != 'BACHELORS':
        if gpa >= 3.5:
            return {
                'class': 'Distinction',
                'color': '#16A34A',
                'description': 'Outstanding Performance'
            }
        elif gpa >= 3.0:
            return {
                'class': 'Credit',
                'color': '#2563EB',
                'description': 'Very Good Performance'
            }
        elif gpa >= 2.0:
            return {
                'class': 'Pass',
                'color': '#F59E0B',
                'description': 'Good Performance'
            }
        else:
            return {
                'class': 'Below Pass',
                'color': '#DC2626',
                'description': 'Needs Improvement'
            }
    
    # Bachelor's degree classification
    if gpa >= 3.60:
        return {
            'class': 'First Class Honours',
            'color': '#16A34A',
            'description': 'Exceptional Academic Achievement'
        }
    elif gpa >= 3.00:
        return {
            'class': 'Second Class Honours (Upper Division)',
            'color': '#2563EB',
            'description': 'Very Good Academic Performance'
        }
    elif gpa >= 2.50:
        return {
            'class': 'Second Class Honours (Lower Division)',
            'color': '#7C3AED',
            'description': 'Good Academic Performance'
        }
    elif gpa >= 2.00:
        return {
            'class': 'Pass',
            'color': '#F59E0B',
            'description': 'Satisfactory Performance'
        }
    else:
        return {
            'class': 'Fail',
            'color': '#DC2626',
            'description': 'Below Minimum Requirement'
        }
    
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q, Sum
from django.http import JsonResponse
from django.utils import timezone
from decimal import Decimal, InvalidOperation
from .models import (
    Student, UnitEnrollment, Semester, AcademicYear, 
    Unit, AssessmentComponent, StudentMarks, FinalGrade,
    Lecturer, GradingScheme
)


def is_admin_or_staff(user):
    """Check if user is admin or staff"""
    return user.is_staff or user.is_superuser or user.user_type in ['COD', 'DEAN', 'ICT_ADMIN']


@login_required
@user_passes_test(is_admin_or_staff)
def marks_entry_view(request):
    """Main view for marks entry - search student and select semester"""
    
    # Get all active academic years and semesters
    academic_years = AcademicYear.objects.filter(is_current=True).order_by('-start_date')
    semesters = Semester.objects.filter(
        academic_year__is_current=True
    ).select_related('academic_year').order_by('-academic_year__start_date', '-semester_number')
    
    # Also get past semesters for reference
    past_semesters = Semester.objects.filter(
        is_current=False
    ).select_related('academic_year').order_by('-academic_year__start_date', '-semester_number')[:10]
    
    all_semesters = list(semesters) + list(past_semesters)
    
    context = {
        'academic_years': academic_years,
        'semesters': all_semesters,
    }
    
    return render(request, 'admin/marks_entry.html', context)


@login_required
@user_passes_test(is_admin_or_staff)
def search_student_ajax(request):
    """AJAX endpoint to search for student by registration number"""
    
    if request.method == 'GET':
        reg_number = request.GET.get('reg_number', '').strip()
        
        if not reg_number:
            return JsonResponse({'error': 'Registration number is required'}, status=400)
        
        try:
            student = Student.objects.select_related(
                'user', 'programme'
            ).get(registration_number__iexact=reg_number)
            
            return JsonResponse({
                'success': True,
                'student': {
                    'id': student.user_id,
                    'registration_number': student.registration_number,
                    'full_name': f"{student.first_name} {student.last_name}",
                    'email': student.email or student.user.email,
                    'programme': student.programme.name,
                    'programme_code': student.programme.code,
                    'current_year': student.current_year,
                }
            })
            
        except Student.DoesNotExist:
            return JsonResponse({
                'error': 'Student not found with this registration number'
            }, status=404)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
@user_passes_test(is_admin_or_staff)
def student_units_view(request, student_id, semester_id):
    """View to display and enter marks for student's enrolled units"""
    
    student = get_object_or_404(
        Student.objects.select_related('user', 'programme'),
        user_id=student_id
    )
    
    semester = get_object_or_404(
        Semester.objects.select_related('academic_year'),
        id=semester_id
    )
    
    # Get all enrollments for this student in this semester
    enrollments = UnitEnrollment.objects.filter(
        student=student,
        semester=semester
    ).select_related(
        'unit'
    ).prefetch_related(
        'marks__assessment_component'
    ).order_by('unit__code')
    
    # Get the lecturer (if current user is a lecturer)
    lecturer = None
    try:
        if hasattr(request.user, 'lecturer_profile'):
            lecturer = request.user.lecturer_profile
    except:
        pass
    
    # Prepare enrollment data with assessments
    enrollment_data = []
    for enrollment in enrollments:
        # Get all assessment components for this unit
        assessment_components = AssessmentComponent.objects.filter(
            unit=enrollment.unit
        ).order_by('component_type')
        
        # Get marks for this enrollment
        marks_dict = {}
        for mark in enrollment.marks.all():
            marks_dict[mark.assessment_component.id] = mark
        
        # Calculate total marks
        total_marks = Decimal('0.00')
        total_weight = Decimal('0.00')
        
        assessment_list = []
        for component in assessment_components:
            mark_obj = marks_dict.get(component.id)
            
            # Calculate weighted score
            weighted_score = None
            if mark_obj and mark_obj.marks_obtained is not None:
                percentage = (mark_obj.marks_obtained / component.max_marks) * 100
                weighted_score = (percentage * component.weight_percentage) / 100
                total_marks += weighted_score
            
            total_weight += component.weight_percentage
            
            assessment_list.append({
                'component': component,
                'mark': mark_obj,
                'weighted_score': weighted_score
            })
        
        # Check if final grade exists
        try:
            final_grade = enrollment.final_grade
        except:
            final_grade = None
        
        enrollment_data.append({
            'enrollment': enrollment,
            'assessments': assessment_list,
            'total_marks': round(total_marks, 2),
            'total_weight': total_weight,
            'final_grade': final_grade,
            'has_all_marks': all(a['mark'] is not None and a['mark'].marks_obtained is not None for a in assessment_list)
        })
    
    # Get grading scheme for this programme
    grading_scheme = GradingScheme.objects.filter(
        programme=student.programme
    ).order_by('-min_marks')
    
    context = {
        'student': student,
        'semester': semester,
        'enrollment_data': enrollment_data,
        'grading_scheme': grading_scheme,
        'lecturer': lecturer,
    }
    
    return render(request, 'admin/student_marks_entry.html', context)


@login_required
@user_passes_test(is_admin_or_staff)
def save_marks_ajax(request):
    """AJAX endpoint to save marks for a specific assessment"""
    
    if request.method == 'POST':
        enrollment_id = request.POST.get('enrollment_id')
        component_id = request.POST.get('component_id')
        marks = request.POST.get('marks')
        
        if not all([enrollment_id, component_id]):
            return JsonResponse({'error': 'Missing required fields'}, status=400)
        
        try:
            enrollment = UnitEnrollment.objects.get(id=enrollment_id)
            component = AssessmentComponent.objects.get(id=component_id)
            
            # Handle empty marks (delete existing)
            if marks == '' or marks is None:
                StudentMarks.objects.filter(
                    enrollment=enrollment,
                    assessment_component=component
                ).delete()
                return JsonResponse({
                    'success': True,
                    'message': 'Marks cleared successfully',
                    'weighted_score': None
                })
            
            marks = Decimal(marks)
            
            # Validate marks
            if marks < 0 or marks > component.max_marks:
                return JsonResponse({
                    'error': f'Marks must be between 0 and {component.max_marks}'
                }, status=400)
            
            # Get or create lecturer
            lecturer = None
            try:
                if hasattr(request.user, 'lecturer_profile'):
                    lecturer = request.user.lecturer_profile
                else:
                    # Get any lecturer as fallback
                    lecturer = Lecturer.objects.first()
            except:
                lecturer = Lecturer.objects.first()
            
            if not lecturer:
                return JsonResponse({'error': 'No lecturer found in system'}, status=400)
            
            # Create or update marks
            mark_obj, created = StudentMarks.objects.update_or_create(
                enrollment=enrollment,
                assessment_component=component,
                defaults={
                    'marks_obtained': marks,
                    'entered_by': lecturer
                }
            )
            
            # Calculate weighted score
            percentage = (marks / component.max_marks) * 100
            weighted_score = (percentage * component.weight_percentage) / 100
            
            return JsonResponse({
                'success': True,
                'message': 'Marks saved successfully',
                'weighted_score': float(round(weighted_score, 2)),
                'created': created
            })
            
        except (UnitEnrollment.DoesNotExist, AssessmentComponent.DoesNotExist):
            return JsonResponse({'error': 'Enrollment or Assessment Component not found'}, status=404)
        except (ValueError, InvalidOperation) as e:
            return JsonResponse({'error': f'Invalid marks value: {str(e)}'}, status=400)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
@user_passes_test(is_admin_or_staff)
def calculate_final_grade_ajax(request):
    """AJAX endpoint to calculate and save final grade for an enrollment"""
    
    if request.method == 'POST':
        enrollment_id = request.POST.get('enrollment_id')
        
        if not enrollment_id:
            return JsonResponse({'error': 'Enrollment ID is required'}, status=400)
        
        try:
            enrollment = UnitEnrollment.objects.select_related('unit', 'student__programme').get(id=enrollment_id)
            
            # Get all assessment components for this unit
            components = AssessmentComponent.objects.filter(unit=enrollment.unit)
            
            if not components.exists():
                return JsonResponse({'error': 'No assessment components defined for this unit'}, status=400)
            
            # Check if all components have marks
            marks = StudentMarks.objects.filter(
                enrollment=enrollment,
                assessment_component__in=components
            )
            
            if marks.count() != components.count():
                return JsonResponse({
                    'error': 'Not all assessment components have been graded'
                }, status=400)
            
            # Calculate total marks (weighted)
            total_marks = Decimal('0.00')
            
            for mark in marks:
                component = mark.assessment_component
                percentage = (mark.marks_obtained / component.max_marks) * 100
                weighted_score = (percentage * component.weight_percentage) / 100
                total_marks += weighted_score
            
            # Determine grade based on grading scheme
            grading_scheme = GradingScheme.objects.filter(
                programme=enrollment.student.programme,
                min_marks__lte=total_marks,
                max_marks__gte=total_marks
            ).first()
            
            if not grading_scheme:
                return JsonResponse({'error': 'No matching grade found in grading scheme'}, status=400)
            
            # Get or create lecturer
            lecturer = None
            try:
                if hasattr(request.user, 'lecturer_profile'):
                    lecturer = request.user.lecturer_profile
            except:
                pass
            
            # Create or update final grade
            final_grade, created = FinalGrade.objects.update_or_create(
                enrollment=enrollment,
                defaults={
                    'total_marks': total_marks,
                    'grade': grading_scheme.grade,
                    'grade_point': grading_scheme.grade_point,
                    'approved_by': lecturer,
                    'is_approved': True
                }
            )
            
            # Update enrollment status
            enrollment.status = 'COMPLETED'
            enrollment.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Final grade calculated successfully',
                'total_marks': float(round(total_marks, 2)),
                'grade': grading_scheme.grade,
                'grade_point': float(grading_scheme.grade_point),
                'description': grading_scheme.description
            })
            
        except UnitEnrollment.DoesNotExist:
            return JsonResponse({'error': 'Enrollment not found'}, status=404)
        except Exception as e:
            return JsonResponse({'error': f'Error calculating grade: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)


@login_required
@user_passes_test(is_admin_or_staff)
def bulk_marks_entry_view(request, semester_id, unit_id):
    """View for entering marks for all students in a unit at once"""
    
    semester = get_object_or_404(
        Semester.objects.select_related('academic_year'),
        id=semester_id
    )
    
    unit = get_object_or_404(Unit, id=unit_id)
    
    # Get all enrollments for this unit in this semester
    enrollments = UnitEnrollment.objects.filter(
        semester=semester,
        unit=unit,
        status__in=['ENROLLED', 'COMPLETED']
    ).select_related(
        'student__user',
        'student__programme'
    ).prefetch_related(
        'marks__assessment_component'
    ).order_by('student__registration_number')
    
    # Get assessment components for this unit
    components = AssessmentComponent.objects.filter(
        unit=unit
    ).order_by('component_type')
    
    # Prepare data
    student_data = []
    for enrollment in enrollments:
        marks_dict = {}
        for mark in enrollment.marks.all():
            marks_dict[mark.assessment_component.id] = mark.marks_obtained
        
        # Get final grade if exists
        try:
            final_grade = enrollment.final_grade
        except:
            final_grade = None
        
        student_data.append({
            'enrollment': enrollment,
            'student': enrollment.student,
            'marks': marks_dict,
            'final_grade': final_grade
        })
    
    context = {
        'semester': semester,
        'unit': unit,
        'components': components,
        'student_data': student_data,
    }
    
    return render(request, 'admin/bulk_marks_entry.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib.auth.hashers import make_password
import csv
import json
from datetime import datetime

from .models import (
    Student, User, Programme, Department, Intake, 
    AcademicYear, UnitEnrollment, SemesterRegistration
)

# Helper function to check if user is admin
def is_admin(user):
    return user.is_authenticated and user.user_type in ['ICT_ADMIN', 'DEAN', 'COD']

# ========================
# STUDENT LIST VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def student_list(request):
    """List all students with filtering and pagination"""
    
    # Get all students
    students = Student.objects.select_related(
        'user', 'programme', 'programme__department', 'intake'
    ).all()
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    programme_filter = request.GET.get('programme', '')
    year_filter = request.GET.get('year', '')
    intake_filter = request.GET.get('intake', '')
    status_filter = request.GET.get('status', '')
    
    # Apply filters
    if search_query:
        students = students.filter(
            Q(registration_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(surname__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    if programme_filter:
        students = students.filter(programme_id=programme_filter)
    
    if year_filter:
        students = students.filter(current_year=year_filter)
    
    if intake_filter:
        students = students.filter(intake_id=intake_filter)
    
    if status_filter:
        if status_filter == 'active':
            students = students.filter(is_active=True)
        elif status_filter == 'inactive':
            students = students.filter(is_active=False)
    
    # Get statistics
    total_students = students.count()
    active_students = students.filter(is_active=True).count()
    inactive_students = students.filter(is_active=False).count()
    
    # Statistics by year
    year_stats = students.values('current_year').annotate(count=Count('user')).order_by('current_year')
    
    # Pagination
    paginator = Paginator(students, 20)  # 20 students per page
    page = request.GET.get('page', 1)
    
    try:
        students_page = paginator.page(page)
    except PageNotAnInteger:
        students_page = paginator.page(1)
    except EmptyPage:
        students_page = paginator.page(paginator.num_pages)
    
    # Get filter options
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    intakes = Intake.objects.filter(is_active=True).order_by('-intake_date')[:10]
    year_levels = Student.YEAR_LEVELS
    
    context = {
        'students': students_page,
        'page_obj': students_page,
        'total_students': total_students,
        'active_students': active_students,
        'inactive_students': inactive_students,
        'year_stats': year_stats,
        'programmes': programmes,
        'intakes': intakes,
        'year_levels': year_levels,
        'search_query': search_query,
        'programme_filter': programme_filter,
        'year_filter': year_filter,
        'intake_filter': intake_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'admin/student_list.html', context)


# ========================
# STUDENT DETAIL VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def student_detail(request, registration_number):
    """View detailed information about a student"""
    
    student = get_object_or_404(
        Student.objects.select_related(
            'user', 'programme', 'programme__department', 'intake'
        ),
        registration_number=registration_number
    )
    
    # Get enrollment history
    enrollments = UnitEnrollment.objects.filter(
        student=student
    ).select_related(
        'unit', 'semester', 'semester__academic_year'
    ).order_by('-semester__start_date')
    
    # Get semester registrations
    registrations = SemesterRegistration.objects.filter(
        student=student
    ).select_related('semester', 'semester__academic_year').order_by('-registration_date')
    
    # Get progression history
    progressions = student.progressions.select_related(
        'from_programme', 'to_programme'
    ).order_by('-completion_date')
    
    # Calculate statistics
    total_enrollments = enrollments.count()
    completed_units = enrollments.filter(status='COMPLETED').count()
    failed_units = enrollments.filter(status='FAILED').count()
    current_enrollments = enrollments.filter(status='ENROLLED').count()
    
    context = {
        'student': student,
        'enrollments': enrollments[:10],  # Last 10 enrollments
        'registrations': registrations[:5],  # Last 5 registrations
        'progressions': progressions,
        'total_enrollments': total_enrollments,
        'completed_units': completed_units,
        'failed_units': failed_units,
        'current_enrollments': current_enrollments,
    }
    
    return render(request, 'admin/student_detail.html', context)


# ========================
# CREATE STUDENT VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def student_create(request):
    """Create a new student"""
    
    if request.method == 'POST':
        try:
            # Get user data
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            surname = request.POST.get('surname', '')
            phone_number = request.POST.get('phone_number', '')
            
            # Get student data
            registration_number = request.POST.get('registration_number')
            programme_id = request.POST.get('programme')
            current_year = request.POST.get('current_year')
            intake_id = request.POST.get('intake')
            admission_date = request.POST.get('admission_date')
            
            # Additional details
            phone = request.POST.get('phone', '')
            student_email = request.POST.get('student_email', '')
            date_of_birth = request.POST.get('date_of_birth', None)
            address = request.POST.get('address', '')
            parent_name = request.POST.get('parent_name', '')
            parent_phone = request.POST.get('parent_phone', '')
            guardian_name = request.POST.get('guardian_name', '')
            guardian_phone = request.POST.get('guardian_phone', '')
            
            # Validation
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists!')
                return redirect('student_create')
            
            if Student.objects.filter(registration_number=registration_number).exists():
                messages.error(request, 'Registration number already exists!')
                return redirect('student_create')
            
            # Create user
            user = User.objects.create(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                user_type='STUDENT',
                phone_number=phone_number,
                is_active=True
            )
            user.password = make_password(password)
            user.save()
            
            # Create student
            student = Student.objects.create(
                user=user,
                registration_number=registration_number,
                first_name=first_name,
                last_name=last_name,
                surname=surname,
                programme_id=programme_id,
                current_year=current_year,
                intake_id=intake_id,
                admission_date=admission_date,
                phone=phone,
                email=student_email,
                date_of_birth=date_of_birth if date_of_birth else None,
                address=address,
                parent_name=parent_name,
                parent_phone=parent_phone,
                guardian_name=guardian_name,
                guardian_phone=guardian_phone,
                is_active=True
            )
            
            messages.success(request, f'Student {registration_number} created successfully!')
            return redirect('student_detail', registration_number=registration_number)
            
        except Exception as e:
            messages.error(request, f'Error creating student: {str(e)}')
            return redirect('student_create')
    
    # GET request - show form
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    intakes = Intake.objects.filter(is_active=True).order_by('-intake_date')
    year_levels = Student.YEAR_LEVELS
    
    context = {
        'programmes': programmes,
        'intakes': intakes,
        'year_levels': year_levels,
    }
    
    return render(request, 'admin/student_create.html', context)


# ========================
# UPDATE STUDENT VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def student_update(request, registration_number):
    """Update student information"""
    
    student = get_object_or_404(Student, registration_number=registration_number)
    user = student.user
    
    if request.method == 'POST':
        try:
            # Update user data
            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.phone_number = request.POST.get('phone_number', '')
            
            # Update password if provided
            new_password = request.POST.get('new_password')
            if new_password:
                user.password = make_password(new_password)
            
            user.save()
            
            # Update student data
            student.first_name = request.POST.get('first_name')
            student.last_name = request.POST.get('last_name')
            student.surname = request.POST.get('surname', '')
            student.programme_id = request.POST.get('programme')
            student.current_year = request.POST.get('current_year')
            student.intake_id = request.POST.get('intake')
            student.admission_date = request.POST.get('admission_date')
            student.phone = request.POST.get('phone', '')
            student.email = request.POST.get('student_email', '')
            
            date_of_birth = request.POST.get('date_of_birth')
            student.date_of_birth = date_of_birth if date_of_birth else None
            
            student.address = request.POST.get('address', '')
            student.parent_name = request.POST.get('parent_name', '')
            student.parent_phone = request.POST.get('parent_phone', '')
            student.guardian_name = request.POST.get('guardian_name', '')
            student.guardian_phone = request.POST.get('guardian_phone', '')
            student.is_active = request.POST.get('is_active') == 'on'
            
            student.save()
            
            messages.success(request, f'Student {registration_number} updated successfully!')
            return redirect('student_detail', registration_number=registration_number)
            
        except Exception as e:
            messages.error(request, f'Error updating student: {str(e)}')
    
    # GET request - show form
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    intakes = Intake.objects.filter(is_active=True).order_by('-intake_date')
    year_levels = Student.YEAR_LEVELS
    
    context = {
        'student': student,
        'programmes': programmes,
        'intakes': intakes,
        'year_levels': year_levels,
    }
    
    return render(request, 'admin/student_update.html', context)


# ========================
# DELETE STUDENT VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def student_delete(request, registration_number):
    """Delete a student (soft delete - deactivate)"""
    
    student = get_object_or_404(Student, registration_number=registration_number)
    
    if request.method == 'POST':
        try:
            # Soft delete - deactivate instead of deleting
            student.is_active = False
            student.user.is_active = False
            student.save()
            student.user.save()
            
            messages.success(request, f'Student {registration_number} has been deactivated!')
            return redirect('student_list')
            
        except Exception as e:
            messages.error(request, f'Error deactivating student: {str(e)}')
            return redirect('student_detail', registration_number=registration_number)
    
    context = {
        'student': student,
    }
    
    return render(request, 'admin/student_delete.html', context)


# ========================
# EXPORT STUDENTS VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def student_export(request):
    """Export students to CSV based on filters"""
    
    # Get filter parameters
    programme_id = request.GET.get('programme', '')
    year = request.GET.get('year', '')
    intake_id = request.GET.get('intake', '')
    status = request.GET.get('status', '')
    
    # Get students
    students = Student.objects.select_related(
        'user', 'programme', 'intake'
    ).all()
    
    # Apply filters
    if programme_id:
        students = students.filter(programme_id=programme_id)
    
    if year:
        students = students.filter(current_year=year)
    
    if intake_id:
        students = students.filter(intake_id=intake_id)
    
    if status == 'active':
        students = students.filter(is_active=True)
    elif status == 'inactive':
        students = students.filter(is_active=False)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f'students_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Registration Number',
        'First Name',
        'Last Name',
        'Surname',
        'Email',
        'Phone',
        'Programme',
        'Current Year',
        'Intake',
        'Admission Date',
        'Status',
        'Date of Birth',
        'Address',
        'Parent Name',
        'Parent Phone',
        'Guardian Name',
        'Guardian Phone'
    ])
    
    # Write data
    for student in students:
        writer.writerow([
            student.registration_number,
            student.first_name,
            student.last_name,
            student.surname,
            student.email or '',
            student.phone or '',
            student.programme.name,
            student.get_current_year_display(),
            student.intake.name,
            student.admission_date.strftime('%Y-%m-%d'),
            'Active' if student.is_active else 'Inactive',
            student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
            student.address or '',
            student.parent_name or '',
            student.parent_phone or '',
            student.guardian_name or '',
            student.guardian_phone or ''
        ])
    
    return response


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q, Count
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib.auth.hashers import make_password
import csv
import json
from datetime import datetime

from .models import (
    Lecturer, User, Department, UnitAllocation, 
    Unit, Semester, Programme
)

# Helper function to check if user is admin
def is_admin(user):
    return user.is_authenticated and user.user_type in ['ICT_ADMIN', 'DEAN', 'COD']

# ========================
# LECTURER LIST VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def lecturer_list(request):
    """List all lecturers with filtering and pagination"""
    
    # Get all lecturers
    lecturers = Lecturer.objects.select_related(
        'user', 'department'
    ).all()
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    department_filter = request.GET.get('department', '')
    status_filter = request.GET.get('status', '')
    specialization_filter = request.GET.get('specialization', '')
    
    # Apply filters
    if search_query:
        lecturers = lecturers.filter(
            Q(staff_number__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(user__phone_number__icontains=search_query) |
            Q(specialization__icontains=search_query)
        )
    
    if department_filter:
        lecturers = lecturers.filter(department_id=department_filter)
    
    if status_filter:
        if status_filter == 'active':
            lecturers = lecturers.filter(is_active=True)
        elif status_filter == 'inactive':
            lecturers = lecturers.filter(is_active=False)
    
    if specialization_filter:
        lecturers = lecturers.filter(specialization__icontains=specialization_filter)
    
    # Get statistics
    total_lecturers = lecturers.count()
    active_lecturers = lecturers.filter(is_active=True).count()
    inactive_lecturers = lecturers.filter(is_active=False).count()
    
    # Statistics by department
    dept_stats = lecturers.values('department__name').annotate(count=Count('user')).order_by('-count')
    
    # Pagination
    paginator = Paginator(lecturers, 20)  # 20 lecturers per page
    page = request.GET.get('page', 1)
    
    try:
        lecturers_page = paginator.page(page)
    except PageNotAnInteger:
        lecturers_page = paginator.page(1)
    except EmptyPage:
        lecturers_page = paginator.page(paginator.num_pages)
    
    # Get filter options
    departments = Department.objects.all().order_by('name')
    
    context = {
        'lecturers': lecturers_page,
        'page_obj': lecturers_page,
        'total_lecturers': total_lecturers,
        'active_lecturers': active_lecturers,
        'inactive_lecturers': inactive_lecturers,
        'dept_stats': dept_stats,
        'departments': departments,
        'search_query': search_query,
        'department_filter': department_filter,
        'status_filter': status_filter,
        'specialization_filter': specialization_filter,
    }
    
    return render(request, 'admin/lecturer_list.html', context)


# ========================
# LECTURER DETAIL VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def lecturer_detail(request, staff_number):
    """View detailed information about a lecturer"""
    
    lecturer = get_object_or_404(
        Lecturer.objects.select_related(
            'user', 'department'
        ),
        staff_number=staff_number
    )
    
    # Get unit allocations
    allocations = UnitAllocation.objects.filter(
        lecturer=lecturer,
        is_active=True
    ).select_related(
        'unit', 'semester', 'semester__academic_year'
    ).prefetch_related('programmes').order_by('-semester__start_date')
    
    # Get all unit allocations (including inactive)
    all_allocations = UnitAllocation.objects.filter(
        lecturer=lecturer
    ).select_related(
        'unit', 'semester', 'semester__academic_year'
    ).order_by('-allocated_date')
    
    # Calculate statistics
    total_allocations = all_allocations.count()
    active_allocations = allocations.count()
    
    # Get unique units taught
    units_taught = all_allocations.values('unit__code', 'unit__name').distinct()
    total_units = units_taught.count()
    
    # Get current semester allocations
    current_allocations = allocations.filter(semester__is_current=True)
    
    context = {
        'lecturer': lecturer,
        'allocations': allocations[:10],  # Last 10 allocations
        'current_allocations': current_allocations,
        'total_allocations': total_allocations,
        'active_allocations': active_allocations,
        'total_units': total_units,
        'units_taught': units_taught[:5],  # Top 5 units
    }
    
    return render(request, 'admin/lecturer_detail.html', context)


# ========================
# CREATE LECTURER VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def lecturer_create(request):
    """Create a new lecturer"""
    
    if request.method == 'POST':
        try:
            # Get user data
            username = request.POST.get('username')
            email = request.POST.get('email')
            password = request.POST.get('password')
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            phone_number = request.POST.get('phone_number', '')
            
            # Get lecturer data
            staff_number = request.POST.get('staff_number')
            department_id = request.POST.get('department')
            specialization = request.POST.get('specialization', '')
            office_location = request.POST.get('office_location', '')
            consultation_hours = request.POST.get('consultation_hours', '')
            
            # Validation
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists!')
                return redirect('lecturer_create')
            
            if Lecturer.objects.filter(staff_number=staff_number).exists():
                messages.error(request, 'Staff number already exists!')
                return redirect('lecturer_create')
            
            # Create user
            user = User.objects.create(
                username=username,
                email=email,
                first_name=first_name,
                last_name=last_name,
                user_type='LECTURER',
                phone_number=phone_number,
                is_active=True
            )
            user.password = make_password(password)
            user.save()
            
            # Create lecturer
            lecturer = Lecturer.objects.create(
                user=user,
                staff_number=staff_number,
                department_id=department_id,
                specialization=specialization,
                office_location=office_location,
                consultation_hours=consultation_hours,
                is_active=True
            )
            
            messages.success(request, f'Lecturer {staff_number} created successfully!')
            return redirect('lecturer_detail', staff_number=staff_number)
            
        except Exception as e:
            messages.error(request, f'Error creating lecturer: {str(e)}')
            return redirect('lecturer_create')
    
    # GET request - show form
    departments = Department.objects.all().order_by('name')
    
    context = {
        'departments': departments,
    }
    
    return render(request, 'admin/lecturer_create.html', context)


# ========================
# UPDATE LECTURER VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def lecturer_update(request, staff_number):
    """Update lecturer information"""
    
    lecturer = get_object_or_404(Lecturer, staff_number=staff_number)
    user = lecturer.user
    
    if request.method == 'POST':
        try:
            # Update user data
            user.username = request.POST.get('username')
            user.email = request.POST.get('email')
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.phone_number = request.POST.get('phone_number', '')
            
            # Update password if provided
            new_password = request.POST.get('new_password')
            if new_password:
                user.password = make_password(new_password)
            
            user.save()
            
            # Update lecturer data
            lecturer.department_id = request.POST.get('department')
            lecturer.specialization = request.POST.get('specialization', '')
            lecturer.office_location = request.POST.get('office_location', '')
            lecturer.consultation_hours = request.POST.get('consultation_hours', '')
            lecturer.is_active = request.POST.get('is_active') == 'on'
            
            lecturer.save()
            
            messages.success(request, f'Lecturer {staff_number} updated successfully!')
            return redirect('lecturer_detail', staff_number=staff_number)
            
        except Exception as e:
            messages.error(request, f'Error updating lecturer: {str(e)}')
    
    # GET request - show form
    departments = Department.objects.all().order_by('name')
    
    context = {
        'lecturer': lecturer,
        'departments': departments,
    }
    
    return render(request, 'admin/lecturer_update.html', context)


# ========================
# DELETE LECTURER VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def lecturer_delete(request, staff_number):
    """Delete a lecturer (soft delete - deactivate)"""
    
    lecturer = get_object_or_404(Lecturer, staff_number=staff_number)
    
    if request.method == 'POST':
        try:
            # Soft delete - deactivate instead of deleting
            lecturer.is_active = False
            lecturer.user.is_active = False
            lecturer.save()
            lecturer.user.save()
            
            # Also deactivate all unit allocations
            UnitAllocation.objects.filter(lecturer=lecturer, is_active=True).update(is_active=False)
            
            messages.success(request, f'Lecturer {staff_number} has been deactivated!')
            return redirect('lecturer_list')
            
        except Exception as e:
            messages.error(request, f'Error deactivating lecturer: {str(e)}')
            return redirect('lecturer_detail', staff_number=staff_number)
    
    # Get allocation count for warning
    active_allocations = UnitAllocation.objects.filter(
        lecturer=lecturer, 
        is_active=True
    ).count()
    
    context = {
        'lecturer': lecturer,
        'active_allocations': active_allocations,
    }
    
    return render(request, 'admin/lecturer_delete.html', context)


# ========================
# EXPORT LECTURERS VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def lecturer_export(request):
    """Export lecturers to CSV based on filters"""
    
    # Get filter parameters
    department_id = request.GET.get('department', '')
    status = request.GET.get('status', '')
    specialization = request.GET.get('specialization', '')
    
    # Get lecturers
    lecturers = Lecturer.objects.select_related(
        'user', 'department'
    ).all()
    
    # Apply filters
    if department_id:
        lecturers = lecturers.filter(department_id=department_id)
    
    if status == 'active':
        lecturers = lecturers.filter(is_active=True)
    elif status == 'inactive':
        lecturers = lecturers.filter(is_active=False)
    
    if specialization:
        lecturers = lecturers.filter(specialization__icontains=specialization)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f'lecturers_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Staff Number',
        'First Name',
        'Last Name',
        'Email',
        'Phone',
        'Department',
        'Specialization',
        'Office Location',
        'Consultation Hours',
        'Status',
        'Date Joined'
    ])
    
    # Write data
    for lecturer in lecturers:
        writer.writerow([
            lecturer.staff_number,
            lecturer.user.first_name,
            lecturer.user.last_name,
            lecturer.user.email,
            lecturer.user.phone_number or '',
            lecturer.department.name,
            lecturer.specialization or '',
            lecturer.office_location or '',
            lecturer.consultation_hours or '',
            'Active' if lecturer.is_active else 'Inactive',
            lecturer.user.date_joined.strftime('%Y-%m-%d') if lecturer.user.date_joined else ''
        ])
    
    return response

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Count, Q
from django.views.decorators.http import require_http_methods
import csv
from datetime import datetime
from .models import (
    Programme, Department, ProgrammeUnit, Unit, Student, 
    UnitAllocation, Semester
)


# ========================
# HELPER FUNCTIONS
# ========================

def is_admin(user):
    """Check if user is ICT Admin, Dean, or COD"""
    return user.user_type in ['ICT_ADMIN', 'DEAN', 'COD']


# ========================
# PROGRAMME LIST VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def programme_list(request):
    """List all programmes with filters"""
    
    # Get filter parameters
    department_id = request.GET.get('department', '')
    level = request.GET.get('level', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Get programmes with related data
    programmes = Programme.objects.select_related(
        'department'
    ).annotate(
        student_count=Count('students', filter=Q(students__is_active=True)),
        unit_count=Count('programme_units', distinct=True)
    ).all()
    
    # Apply filters
    if department_id:
        programmes = programmes.filter(department_id=department_id)
    
    if level:
        programmes = programmes.filter(level=level)
    
    if status == 'active':
        programmes = programmes.filter(is_active=True)
    elif status == 'inactive':
        programmes = programmes.filter(is_active=False)
    
    if search:
        programmes = programmes.filter(
            Q(name__icontains=search) |
            Q(code__icontains=search) |
            Q(description__icontains=search)
        )
    
    # Get departments for filter
    departments = Department.objects.all().order_by('name')
    
    # Programme levels for filter
    programme_levels = Programme.PROGRAMME_LEVELS
    
    context = {
        'programmes': programmes,
        'departments': departments,
        'programme_levels': programme_levels,
        'selected_department': department_id,
        'selected_level': level,
        'selected_status': status,
        'search_query': search,
        'total_programmes': programmes.count(),
    }
    
    return render(request, 'admin/programme_list.html', context)


# ========================
# PROGRAMME DETAIL VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def programme_detail(request, programme_code):
    """View detailed programme information with units organized by year and semester"""
    
    programme = get_object_or_404(
        Programme.objects.select_related('department'),
        code=programme_code
    )
    
    # Get all available units in the department
    available_units = Unit.objects.filter(
        department=programme.department
    ).order_by('code')
    
    # Organize programme units by year and semester
    programme_structure = {}
    
    for year in range(1, programme.duration_years + 1):
        programme_structure[year] = {}
        
        for sem in range(1, programme.semesters_per_year + 1):
            # Get units for this year and semester
            units = ProgrammeUnit.objects.filter(
                programme=programme,
                year_level=year,
                semester=sem
            ).select_related('unit').order_by('unit__code')
            
            programme_structure[year][sem] = {
                'units': units,
                'total_units': units.count(),
                'core_units': units.filter(is_mandatory=True).count(),
                'elective_units': units.filter(is_mandatory=False).count(),
                'total_credits': sum(u.unit.credit_hours for u in units)
            }
    
    # Calculate statistics
    total_students = Student.objects.filter(
        programme=programme,
        is_active=True
    ).count()
    
    total_units = ProgrammeUnit.objects.filter(programme=programme).count()
    
    total_credits = sum(
        pu.unit.credit_hours 
        for pu in ProgrammeUnit.objects.filter(programme=programme).select_related('unit')
    )
    
    # Get year levels for the programme
    year_levels = list(range(1, programme.duration_years + 1))
    semester_numbers = list(range(1, programme.semesters_per_year + 1))
    
    context = {
        'programme': programme,
        'programme_structure': programme_structure,
        'available_units': available_units,
        'year_levels': year_levels,
        'semester_numbers': semester_numbers,
        'total_students': total_students,
        'total_units': total_units,
        'total_credits': total_credits,
    }
    
    return render(request, 'admin/programme_detail.html', context)


# ========================
# ADD UNIT TO PROGRAMME (AJAX)
# ========================
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def programme_add_unit(request, programme_code):
    """Add a unit to a programme (AJAX)"""
    
    try:
        programme = get_object_or_404(Programme, code=programme_code)
        
        unit_id = request.POST.get('unit_id')
        year_level = request.POST.get('year_level')
        semester = request.POST.get('semester')
        is_mandatory = request.POST.get('is_mandatory') == 'true'
        
        # Validate inputs
        if not all([unit_id, year_level, semester]):
            return JsonResponse({
                'success': False,
                'message': 'Missing required fields'
            }, status=400)
        
        unit = get_object_or_404(Unit, id=unit_id)
        
        # Check if unit already exists in this programme/year/semester
        existing = ProgrammeUnit.objects.filter(
            programme=programme,
            unit=unit,
            year_level=year_level,
            semester=semester
        ).exists()
        
        if existing:
            return JsonResponse({
                'success': False,
                'message': 'This unit is already assigned to this year and semester'
            }, status=400)
        
        # Create programme unit
        programme_unit = ProgrammeUnit.objects.create(
            programme=programme,
            unit=unit,
            year_level=year_level,
            semester=semester,
            is_mandatory=is_mandatory
        )
        
        return JsonResponse({
            'success': True,
            'message': f'Unit {unit.code} added successfully',
            'data': {
                'id': programme_unit.id,
                'unit_code': unit.code,
                'unit_name': unit.name,
                'credit_hours': unit.credit_hours,
                'is_mandatory': is_mandatory,
                'year_level': year_level,
                'semester': semester
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


# ========================
# UPDATE PROGRAMME UNIT (AJAX)
# ========================
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def programme_update_unit(request, programme_unit_id):
    """Update a programme unit (AJAX)"""
    
    try:
        programme_unit = get_object_or_404(ProgrammeUnit, id=programme_unit_id)
        
        is_mandatory = request.POST.get('is_mandatory') == 'true'
        
        # Update
        programme_unit.is_mandatory = is_mandatory
        programme_unit.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Unit updated successfully',
            'data': {
                'id': programme_unit.id,
                'is_mandatory': is_mandatory
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


# ========================
# DELETE PROGRAMME UNIT (AJAX)
# ========================
@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def programme_delete_unit(request, programme_unit_id):
    """Delete a programme unit (AJAX)"""
    
    try:
        programme_unit = get_object_or_404(ProgrammeUnit, id=programme_unit_id)
        
        # Check if there are students enrolled in this unit
        from .models import UnitEnrollment
        enrollments = UnitEnrollment.objects.filter(
            unit=programme_unit.unit,
            student__programme=programme_unit.programme
        ).count()
        
        if enrollments > 0:
            return JsonResponse({
                'success': False,
                'message': f'Cannot delete. {enrollments} student(s) are enrolled in this unit.'
            }, status=400)
        
        # Store data before deletion
        unit_code = programme_unit.unit.code
        unit_name = programme_unit.unit.name
        
        # Delete
        programme_unit.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Unit {unit_code} removed from programme successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=500)


# ========================
# CREATE PROGRAMME VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def programme_create(request):
    """Create a new programme"""
    
    if request.method == 'POST':
        try:
            # Get form data
            name = request.POST.get('name')
            code = request.POST.get('code')
            level = request.POST.get('level')
            department_id = request.POST.get('department')
            duration_years = request.POST.get('duration_years')
            semesters_per_year = request.POST.get('semesters_per_year')
            description = request.POST.get('description', '')
            
            # Validation
            if Programme.objects.filter(code=code).exists():
                messages.error(request, 'Programme code already exists!')
                return redirect('programme_create')
            
            # Create programme
            programme = Programme.objects.create(
                name=name,
                code=code,
                level=level,
                department_id=department_id,
                duration_years=duration_years,
                semesters_per_year=semesters_per_year,
                description=description,
                is_active=True
            )
            
            messages.success(request, f'Programme {code} created successfully!')
            return redirect('programme_detail', programme_code=code)
            
        except Exception as e:
            messages.error(request, f'Error creating programme: {str(e)}')
            return redirect('programme_create')
    
    # GET request - show form
    departments = Department.objects.all().order_by('name')
    programme_levels = Programme.PROGRAMME_LEVELS
    duration_years = Programme.DURATION_YEARS
    semesters_per_year = Programme.SEMESTERS_PER_YEAR
    
    context = {
        'departments': departments,
        'programme_levels': programme_levels,
        'duration_years': duration_years,
        'semesters_per_year': semesters_per_year,
    }
    
    return render(request, 'admin/programme_create.html', context)


# ========================
# UPDATE PROGRAMME VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def programme_update(request, programme_code):
    """Update programme information"""
    
    programme = get_object_or_404(Programme, code=programme_code)
    
    if request.method == 'POST':
        try:
            # Update programme data
            programme.name = request.POST.get('name')
            programme.level = request.POST.get('level')
            programme.department_id = request.POST.get('department')
            programme.duration_years = request.POST.get('duration_years')
            programme.semesters_per_year = request.POST.get('semesters_per_year')
            programme.description = request.POST.get('description', '')
            programme.is_active = request.POST.get('is_active') == 'on'
            
            programme.save()
            
            messages.success(request, f'Programme {programme_code} updated successfully!')
            return redirect('programme_detail', programme_code=programme_code)
            
        except Exception as e:
            messages.error(request, f'Error updating programme: {str(e)}')
    
    # GET request - show form
    departments = Department.objects.all().order_by('name')
    programme_levels = Programme.PROGRAMME_LEVELS
    duration_years = Programme.DURATION_YEARS
    semesters_per_year = Programme.SEMESTERS_PER_YEAR
    
    context = {
        'programme': programme,
        'departments': departments,
        'programme_levels': programme_levels,
        'duration_years': duration_years,
        'semesters_per_year': semesters_per_year,
    }
    
    return render(request, 'admin/programme_update.html', context)


# ========================
# DELETE PROGRAMME VIEW
# ========================
@login_required
@user_passes_test(is_admin)
def programme_delete(request, programme_code):
    """Delete a programme (soft delete - deactivate)"""
    
    programme = get_object_or_404(Programme, code=programme_code)
    
    if request.method == 'POST':
        try:
            # Soft delete - deactivate instead of deleting
            programme.is_active = False
            programme.save()
            
            messages.success(request, f'Programme {programme_code} has been deactivated!')
            return redirect('programme_list')
            
        except Exception as e:
            messages.error(request, f'Error deactivating programme: {str(e)}')
            return redirect('programme_detail', programme_code=programme_code)
    
    # Get counts for warning
    active_students = Student.objects.filter(
        programme=programme,
        is_active=True
    ).count()
    
    total_units = ProgrammeUnit.objects.filter(programme=programme).count()
    
    context = {
        'programme': programme,
        'active_students': active_students,
        'total_units': total_units,
    }
    
    return render(request, 'admin/programme_delete.html', context)


# ========================
# EXPORT PROGRAMME STRUCTURE
# ========================
@login_required
@user_passes_test(is_admin)
def programme_export_structure(request, programme_code):
    """Export programme structure to CSV"""
    
    programme = get_object_or_404(Programme, code=programme_code)
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f'{programme.code}_structure_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow([
        'Programme',
        'Year Level',
        'Semester',
        'Unit Code',
        'Unit Name',
        'Credit Hours',
        'Type',
        'Department'
    ])
    
    # Write data
    programme_units = ProgrammeUnit.objects.filter(
        programme=programme
    ).select_related('unit', 'unit__department').order_by(
        'year_level', 'semester', 'unit__code'
    )
    
    for pu in programme_units:
        writer.writerow([
            programme.code,
            f'Year {pu.year_level}',
            f'Semester {pu.semester}',
            pu.unit.code,
            pu.unit.name,
            pu.unit.credit_hours,
            'Core' if pu.is_mandatory else 'Elective',
            pu.unit.department.code
        ])
    
    return response

# views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.utils import timezone
from .models import (
    Announcement, Event, EventRegistration, Programme, 
    Venue, User, Student
)
from django.core.paginator import Paginator


# ========================
# ANNOUNCEMENT VIEWS
# ========================

@login_required
def announcement_list(request):
    """List all announcements with filters"""
    announcements = Announcement.objects.select_related('created_by').prefetch_related('target_programmes')
    
    # Search filter
    search_query = request.GET.get('search', '')
    if search_query:
        announcements = announcements.filter(
            Q(title__icontains=search_query) | 
            Q(content__icontains=search_query)
        )
    
    # Priority filter
    priority = request.GET.get('priority', '')
    if priority:
        announcements = announcements.filter(priority=priority)
    
    # Status filter
    status = request.GET.get('status', '')
    if status == 'published':
        announcements = announcements.filter(is_published=True)
    elif status == 'draft':
        announcements = announcements.filter(is_published=False)
    elif status == 'active':
        announcements = announcements.filter(
            is_published=True,
            publish_date__lte=timezone.now()
        ).filter(Q(expiry_date__isnull=True) | Q(expiry_date__gte=timezone.now()))
    elif status == 'expired':
        announcements = announcements.filter(
            expiry_date__lt=timezone.now()
        )
    
    # Order by
    announcements = announcements.order_by('-publish_date')
    
    # Statistics
    total_announcements = Announcement.objects.count()
    active_announcements = Announcement.objects.filter(
        is_published=True,
        publish_date__lte=timezone.now()
    ).filter(Q(expiry_date__isnull=True) | Q(expiry_date__gte=timezone.now())).count()
    urgent_announcements = Announcement.objects.filter(priority='URGENT').count()
    
    # Pagination
    paginator = Paginator(announcements, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'announcements': page_obj,
        'total_announcements': total_announcements,
        'active_announcements': active_announcements,
        'urgent_announcements': urgent_announcements,
        'search_query': search_query,
        'priority_filter': priority,
        'status_filter': status,
        'priority_choices': Announcement.PRIORITY_LEVELS,
    }
    return render(request, 'announcements/announcement_list.html', context)


from django.utils import timezone
@login_required
def announcement_detail(request, pk):
    """View announcement details"""
    announcement = get_object_or_404(
        Announcement.objects.select_related('created_by').prefetch_related('target_programmes'),
        pk=pk
    )
    
    # Get target programmes
    target_programmes = announcement.target_programmes.all()
    
    # Parse target year levels
    target_years = []
    if announcement.target_year_levels:
        target_years = [int(y) for y in announcement.target_year_levels.split(',') if y.strip()]
    
    # Calculate reach (estimated number of students)
    if target_programmes.exists():
        reach = Student.objects.filter(programme__in=target_programmes)
        if target_years:
            reach = reach.filter(current_year__in=target_years)
        reach_count = reach.count()
    else:
        reach_count = Student.objects.filter(is_active=True).count()
    
    # Check if expired
    is_expired = announcement.expiry_date and announcement.expiry_date < timezone.now()
    is_active = (
        announcement.is_published
        and announcement.publish_date <= timezone.now()
        and not is_expired
    )
    
    context = {
        'announcement': announcement,
        'target_programmes': target_programmes,
        'target_years': target_years,
        'reach_count': reach_count,
        'is_expired': is_expired,
        'is_active': is_active,
    }
    return render(request, 'announcements/announcement_detail.html', context)


@login_required
def announcement_create(request):
    """Create new announcement"""
    if request.method == 'POST':
        try:
            # Get form data
            title = request.POST.get('title')
            content = request.POST.get('content')
            priority = request.POST.get('priority', 'NORMAL')
            is_published = request.POST.get('is_published') == 'on'
            publish_date = request.POST.get('publish_date') or timezone.now()
            expiry_date = request.POST.get('expiry_date') or None
            target_year_levels = ','.join(request.POST.getlist('target_year_levels'))
            
            # Handle file upload
            attachments = request.FILES.get('attachments')
            
            # Create announcement
            announcement = Announcement.objects.create(
                title=title,
                content=content,
                created_by=request.user,
                priority=priority,
                is_published=is_published,
                publish_date=publish_date,
                expiry_date=expiry_date,
                target_year_levels=target_year_levels,
                attachments=attachments
            )
            
            # Add target programmes
            target_programmes = request.POST.getlist('target_programmes')
            if target_programmes:
                announcement.target_programmes.set(target_programmes)
            
            messages.success(request, f'Announcement "{title}" created successfully!')
            return redirect('announcement_detail', pk=announcement.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating announcement: {str(e)}')
    
    # Get programmes for selection
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    
    context = {
        'programmes': programmes,
        'priority_choices': Announcement.PRIORITY_LEVELS,
    }
    return render(request, 'announcements/announcement_form.html', context)


@login_required
def announcement_update(request, pk):
    """Update announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    if request.method == 'POST':
        try:
            announcement.title = request.POST.get('title')
            announcement.content = request.POST.get('content')
            announcement.priority = request.POST.get('priority', 'NORMAL')
            announcement.is_published = request.POST.get('is_published') == 'on'
            announcement.publish_date = request.POST.get('publish_date') or announcement.publish_date
            announcement.expiry_date = request.POST.get('expiry_date') or None
            announcement.target_year_levels = ','.join(request.POST.getlist('target_year_levels'))
            
            # Handle file upload
            if request.FILES.get('attachments'):
                announcement.attachments = request.FILES.get('attachments')
            
            announcement.save()
            
            # Update target programmes
            target_programmes = request.POST.getlist('target_programmes')
            if target_programmes:
                announcement.target_programmes.set(target_programmes)
            else:
                announcement.target_programmes.clear()
            
            messages.success(request, f'Announcement "{announcement.title}" updated successfully!')
            return redirect('announcement_detail', pk=announcement.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating announcement: {str(e)}')
    
    # Get programmes for selection
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    selected_programmes = announcement.target_programmes.values_list('id', flat=True)
    
    # Parse target year levels
    target_years = []
    if announcement.target_year_levels:
        target_years = [y.strip() for y in announcement.target_year_levels.split(',') if y.strip()]
    
    context = {
        'announcement': announcement,
        'programmes': programmes,
        'selected_programmes': list(selected_programmes),
        'target_years': target_years,
        'priority_choices': Announcement.PRIORITY_LEVELS,
        'is_update': True,
    }
    return render(request, 'announcements/announcement_form.html', context)


@login_required
def announcement_delete(request, pk):
    """Delete announcement"""
    announcement = get_object_or_404(Announcement, pk=pk)
    
    if request.method == 'POST':
        title = announcement.title
        announcement.delete()
        messages.success(request, f'Announcement "{title}" deleted successfully!')
        return redirect('announcement_list')
    
    context = {
        'announcement': announcement,
    }
    return render(request, 'announcements/announcement_confirm_delete.html', context)


# ========================
# EVENT VIEWS
# ========================

@login_required
def event_list(request):
    """List all events with filters"""
    events = Event.objects.select_related('organizer', 'venue').prefetch_related('target_programmes')
    
    # Search filter
    search_query = request.GET.get('search', '')
    if search_query:
        events = events.filter(
            Q(title__icontains=search_query) | 
            Q(description__icontains=search_query)
        )
    
    # Event type filter
    event_type = request.GET.get('event_type', '')
    if event_type:
        events = events.filter(event_type=event_type)
    
    # Date filter
    date_filter = request.GET.get('date_filter', '')
    today = timezone.now().date()
    if date_filter == 'upcoming':
        events = events.filter(event_date__gte=today)
    elif date_filter == 'past':
        events = events.filter(event_date__lt=today)
    elif date_filter == 'today':
        events = events.filter(event_date=today)
    
    # Published filter
    status = request.GET.get('status', '')
    if status == 'published':
        events = events.filter(is_published=True)
    elif status == 'draft':
        events = events.filter(is_published=False)
    
    # Order by
    events = events.order_by('-event_date', '-start_time')
    
    # Add registration count
    events = events.annotate(registration_count=Count('registrations'))
    
    # Statistics
    total_events = Event.objects.count()
    upcoming_events = Event.objects.filter(event_date__gte=today, is_published=True).count()
    mandatory_events = Event.objects.filter(is_mandatory=True, event_date__gte=today).count()
    
    # Pagination
    paginator = Paginator(events, 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'events': page_obj,
        'total_events': total_events,
        'upcoming_events': upcoming_events,
        'mandatory_events': mandatory_events,
        'search_query': search_query,
        'event_type_filter': event_type,
        'date_filter': date_filter,
        'status_filter': status,
        'event_type_choices': Event.EVENT_TYPES,
    }
    return render(request, 'events/event_list.html', context)


@login_required
def event_detail(request, pk):
    """View event details"""
    event = get_object_or_404(
        Event.objects.select_related('organizer', 'venue').prefetch_related('target_programmes'),
        pk=pk
    )
    
    # Get registrations
    registrations = EventRegistration.objects.filter(event=event).select_related('student__user')
    total_registrations = registrations.count()
    attended_count = registrations.filter(attended=True).count()
    
    # Get target programmes
    target_programmes = event.target_programmes.all()
    
    # Check if event is past
    is_past = event.event_date < timezone.now().date()
    
    # Check if registration is full
    is_full = event.max_attendees and total_registrations >= event.max_attendees
    
    context = {
        'event': event,
        'target_programmes': target_programmes,
        'registrations': registrations[:10],  # Show first 10
        'total_registrations': total_registrations,
        'attended_count': attended_count,
        'is_past': is_past,
        'is_full': is_full,
    }
    return render(request, 'events/event_detail.html', context)


@login_required
def event_create(request):
    """Create new event"""
    if request.method == 'POST':
        try:
            # Get form data
            title = request.POST.get('title')
            description = request.POST.get('description')
            event_type = request.POST.get('event_type')
            venue_id = request.POST.get('venue')
            event_date = request.POST.get('event_date')
            start_time = request.POST.get('start_time')
            end_time = request.POST.get('end_time')
            is_mandatory = request.POST.get('is_mandatory') == 'on'
            registration_required = request.POST.get('registration_required') == 'on'
            is_published = request.POST.get('is_published') == 'on'
            max_attendees = request.POST.get('max_attendees') or None
            
            # Handle file upload
            poster = request.FILES.get('poster')
            
            # Get venue
            venue = get_object_or_404(Venue, id=venue_id) if venue_id else None
            
            # Create event
            event = Event.objects.create(
                title=title,
                description=description,
                event_type=event_type,
                venue=venue,
                event_date=event_date,
                start_time=start_time,
                end_time=end_time,
                organizer=request.user,
                is_mandatory=is_mandatory,
                registration_required=registration_required,
                is_published=is_published,
                max_attendees=max_attendees,
                poster=poster
            )
            
            # Add target programmes
            target_programmes = request.POST.getlist('target_programmes')
            if target_programmes:
                event.target_programmes.set(target_programmes)
            
            messages.success(request, f'Event "{title}" created successfully!')
            return redirect('event_detail', pk=event.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating event: {str(e)}')
    
    # Get venues and programmes for selection
    venues = Venue.objects.filter(is_available=True).order_by('name')
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    
    context = {
        'venues': venues,
        'programmes': programmes,
        'event_type_choices': Event.EVENT_TYPES,
    }
    return render(request, 'events/event_form.html', context)


@login_required
def event_update(request, pk):
    """Update event"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        try:
            event.title = request.POST.get('title')
            event.description = request.POST.get('description')
            event.event_type = request.POST.get('event_type')
            
            venue_id = request.POST.get('venue')
            event.venue = get_object_or_404(Venue, id=venue_id) if venue_id else None
            
            event.event_date = request.POST.get('event_date')
            event.start_time = request.POST.get('start_time')
            event.end_time = request.POST.get('end_time')
            event.is_mandatory = request.POST.get('is_mandatory') == 'on'
            event.registration_required = request.POST.get('registration_required') == 'on'
            event.is_published = request.POST.get('is_published') == 'on'
            event.max_attendees = request.POST.get('max_attendees') or None
            
            # Handle file upload
            if request.FILES.get('poster'):
                event.poster = request.FILES.get('poster')
            
            event.save()
            
            # Update target programmes
            target_programmes = request.POST.getlist('target_programmes')
            if target_programmes:
                event.target_programmes.set(target_programmes)
            else:
                event.target_programmes.clear()
            
            messages.success(request, f'Event "{event.title}" updated successfully!')
            return redirect('event_detail', pk=event.pk)
            
        except Exception as e:
            messages.error(request, f'Error updating event: {str(e)}')
    
    # Get venues and programmes for selection
    venues = Venue.objects.filter(is_available=True).order_by('name')
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    selected_programmes = event.target_programmes.values_list('id', flat=True)
    
    context = {
        'event': event,
        'venues': venues,
        'programmes': programmes,
        'selected_programmes': list(selected_programmes),
        'event_type_choices': Event.EVENT_TYPES,
        'is_update': True,
    }
    return render(request, 'events/event_form.html', context)


@login_required
def event_delete(request, pk):
    """Delete event"""
    event = get_object_or_404(Event, pk=pk)
    
    if request.method == 'POST':
        title = event.title
        event.delete()
        messages.success(request, f'Event "{title}" deleted successfully!')
        return redirect('event_list')
    
    # Get registration count
    registration_count = EventRegistration.objects.filter(event=event).count()
    
    context = {
        'event': event,
        'registration_count': registration_count,
    }
    return render(request, 'events/event_confirm_delete.html', context)

from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Avg, Sum, Q
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
from .models import (
    Student, Programme, Department, UnitEnrollment, 
    SemesterRegistration, FeePayment, Semester, AcademicYear
)
import json


@login_required
def admin_reports(request):
    """
    Comprehensive admin reports dashboard with multiple chart types
    """
    
    # Get current academic year and semester
    current_academic_year = AcademicYear.objects.filter(is_current=True).first()
    current_semester = Semester.objects.filter(is_current=True).first()
    
    # ===========================
    # 1. STUDENT ENROLLMENT BY PROGRAMME (Bar Chart)
    # ===========================
    programme_enrollments = Student.objects.filter(is_active=True).values(
        'programme__code', 'programme__name'
    ).annotate(
        student_count=Count('user')
    ).order_by('-student_count')
    
    programme_labels = [item['programme__code'] for item in programme_enrollments]
    programme_data = [item['student_count'] for item in programme_enrollments]
    
    
    # ===========================
    # 2. STUDENT DISTRIBUTION BY YEAR LEVEL (Pie Chart)
    # ===========================
    year_distribution = Student.objects.filter(is_active=True).values(
        'current_year'
    ).annotate(
        count=Count('user')
    ).order_by('current_year')
    
    year_labels = [f"Year {item['current_year']}" for item in year_distribution]
    year_data = [item['count'] for item in year_distribution]
    
    
    # ===========================
    # 3. ENROLLMENT TRENDS (Line Chart) - Last 6 months
    # ===========================
    six_months_ago = datetime.now() - timedelta(days=180)
    
    enrollment_trends = UnitEnrollment.objects.filter(
        enrollment_date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('enrollment_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    trend_labels = [item['month'].strftime('%b %Y') for item in enrollment_trends]
    trend_data = [item['count'] for item in enrollment_trends]
    
    
    # ===========================
    # 4. ENROLLMENT STATUS DISTRIBUTION (Donut Chart)
    # ===========================
    status_distribution = UnitEnrollment.objects.values('status').annotate(
        count=Count('id')
    ).order_by('-count')
    
    status_labels = [item['status'].replace('_', ' ').title() for item in status_distribution]
    status_data = [item['count'] for item in status_distribution]
    
    
    # ===========================
    # 5. DEPARTMENT STUDENT COUNT (Bar Chart - Horizontal)
    # ===========================
    department_stats = Department.objects.annotate(
        student_count=Count('programmes__students', filter=Q(programmes__students__is_active=True))
    ).order_by('-student_count')[:10]
    
    dept_labels = [dept.code for dept in department_stats]
    dept_data = [dept.student_count for dept in department_stats]
    
    
    # ===========================
    # 6. FEE PAYMENT STATUS (Pie Chart)
    # ===========================
    if current_semester:
        payment_status = FeePayment.objects.filter(
            semester=current_semester
        ).values('status').annotate(
            count=Count('id')
        )
        
        payment_labels = [item['status'].replace('_', ' ').title() for item in payment_status]
        payment_data = [item['count'] for item in payment_status]
    else:
        payment_labels = []
        payment_data = []
    
    
    # ===========================
    # 7. REGISTRATION TRENDS BY SEMESTER (Line Chart)
    # ===========================
    recent_semesters = Semester.objects.all().order_by('-start_date')[:6]
    
    semester_registrations = []
    semester_labels = []
    
    for sem in reversed(recent_semesters):
        reg_count = SemesterRegistration.objects.filter(semester=sem).count()
        semester_registrations.append(reg_count)
        semester_labels.append(str(sem.semester_number))
    
    
    # ===========================
    # 8. PROGRAMME LEVEL DISTRIBUTION (Donut Chart)
    # ===========================
    level_distribution = Programme.objects.filter(is_active=True).values(
        'level'
    ).annotate(
        count=Count('id')
    ).order_by('-count')
    
    level_labels = [item['level'].replace('_', ' ').title() for item in level_distribution]
    level_data = [item['count'] for item in level_distribution]
    
    
    # ===========================
    # SUMMARY STATISTICS
    # ===========================
    total_students = Student.objects.filter(is_active=True).count()
    total_programmes = Programme.objects.filter(is_active=True).count()
    total_departments = Department.objects.count()
    
    if current_semester:
        current_enrollments = UnitEnrollment.objects.filter(
            semester=current_semester,
            status='ENROLLED'
        ).count()
    else:
        current_enrollments = 0
    
    # Calculate completion rate
    total_enrollments = UnitEnrollment.objects.count()
    completed_enrollments = UnitEnrollment.objects.filter(status='COMPLETED').count()
    completion_rate = (completed_enrollments / total_enrollments * 100) if total_enrollments > 0 else 0
    
    
    # ===========================
    # 9. GENDER DISTRIBUTION (Pie Chart) - If you have gender field
    # ===========================
    # Assuming you might add gender field later
    gender_data = [45, 55]  # Placeholder - Male, Female percentages
    gender_labels = ['Male', 'Female']
    
    
    # ===========================
    # 10. MONTHLY REVENUE TREND (Line Chart)
    # ===========================
    if current_academic_year:
        monthly_revenue = FeePayment.objects.filter(
            semester__academic_year=current_academic_year
        ).annotate(
            month=TruncMonth('payment_date')
        ).values('month').annotate(
            total=Sum('amount_paid')
        ).order_by('month')
        
        revenue_labels = [item['month'].strftime('%b %Y') for item in monthly_revenue]
        revenue_data = [float(item['total']) for item in monthly_revenue]
    else:
        revenue_labels = []
        revenue_data = []
    
    
    context = {
        # Summary Stats
        'total_students': total_students,
        'total_programmes': total_programmes,
        'total_departments': total_departments,
        'current_enrollments': current_enrollments,
        'completion_rate': round(completion_rate, 1),
        
        # Chart Data (converted to JSON for JavaScript)
        'programme_labels': json.dumps(programme_labels),
        'programme_data': json.dumps(programme_data),
        
        'year_labels': json.dumps(year_labels),
        'year_data': json.dumps(year_data),
        
        'trend_labels': json.dumps(trend_labels),
        'trend_data': json.dumps(trend_data),
        
        'status_labels': json.dumps(status_labels),
        'status_data': json.dumps(status_data),
        
        'dept_labels': json.dumps(dept_labels),
        'dept_data': json.dumps(dept_data),
        
        'payment_labels': json.dumps(payment_labels),
        'payment_data': json.dumps(payment_data),
        
        'semester_labels': json.dumps(semester_labels),
        'semester_registrations': json.dumps(semester_registrations),
        
        'level_labels': json.dumps(level_labels),
        'level_data': json.dumps(level_data),
        
        'gender_labels': json.dumps(gender_labels),
        'gender_data': json.dumps(gender_data),
        
        'revenue_labels': json.dumps(revenue_labels),
        'revenue_data': json.dumps(revenue_data),
        
        # Additional Info
        'current_semester': current_semester,
        'current_academic_year': current_academic_year,
    }
    
    return render(request, 'admin/reports.html', context)


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from .models import (
    Programme, Semester, Unit, UnitAllocation, Lecturer, Venue,
    TimetableSlot, Department, ProgrammeUnit
)
import json
from datetime import datetime, time


@login_required
def timetable_create(request):
    """
    Main timetable creation view with drag and drop functionality
    """
    # Get current semester or all semesters
    current_semester = Semester.objects.filter(is_current=True).first()
    semesters = Semester.objects.all().order_by('-start_date')[:10]
    
    # Get all active programmes grouped by department
    programmes = Programme.objects.filter(is_active=True).select_related('department').order_by('department', 'name')
    
    # Get all departments
    departments = Department.objects.all().order_by('name')
    
    # Get all venues
    venues = Venue.objects.filter(is_available=True).order_by('building', 'name')
    
    # Get all lecturers
    lecturers = Lecturer.objects.filter(is_active=True).select_related('user', 'department').order_by('staff_number')
    
    # Time slots configuration (8 AM to 6 PM)
    time_slots = [
        {'start': '08:00', 'end': '09:00', 'label': '08:00 - 09:00'},
        {'start': '09:00', 'end': '10:00', 'label': '09:00 - 10:00'},
        {'start': '10:00', 'end': '11:00', 'label': '10:00 - 11:00'},
        {'start': '11:00', 'end': '12:00', 'label': '11:00 - 12:00'},
        {'start': '12:00', 'end': '13:00', 'label': '12:00 - 13:00'},
        {'start': '13:00', 'end': '14:00', 'label': '13:00 - 14:00'},
        {'start': '14:00', 'end': '15:00', 'label': '14:00 - 15:00'},
        {'start': '15:00', 'end': '16:00', 'label': '15:00 - 16:00'},
        {'start': '16:00', 'end': '17:00', 'label': '16:00 - 17:00'},
        {'start': '17:00', 'end': '18:00', 'label': '17:00 - 18:00'},
    ]
    
    # Days of the week
    days_of_week = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
    
    context = {
        'current_semester': current_semester,
        'semesters': semesters,
        'programmes': programmes,
        'departments': departments,
        'venues': venues,
        'lecturers': lecturers,
        'time_slots': time_slots,
        'days_of_week': days_of_week,
        'year_levels': [1, 2, 3, 4],
    }
    
    return render(request, 'admin/timetable_create.html', context)


@login_required
@require_http_methods(["GET"])
def get_programme_units(request, programme_id, year_level):
    """
    Get units for a specific programme and year level
    """
    try:
        programme = get_object_or_404(Programme, id=programme_id)
        
        # Get programme units for the specified year level
        programme_units = ProgrammeUnit.objects.filter(
            programme=programme,
            year_level=year_level
        ).select_related('unit').order_by('semester', 'unit__code')
        
        units_data = []
        for pu in programme_units:
            units_data.append({
                'id': pu.unit.id,
                'code': pu.unit.code,
                'name': pu.unit.name,
                'credit_hours': pu.unit.credit_hours,
                'semester': pu.semester,
                'is_mandatory': pu.is_mandatory,
            })
        
        return JsonResponse({
            'success': True,
            'units': units_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def get_existing_timetable(request, programme_id, year_level, semester_id):
    """
    Get existing timetable slots for a programme, year level, and semester
    """
    try:
        timetable_slots = TimetableSlot.objects.filter(
            programme_id=programme_id,
            year_level=year_level,
            unit_allocation__semester_id=semester_id,
            is_active=True
        ).select_related(
            'unit_allocation__unit',
            'unit_allocation__lecturer__user',
            'venue'
        ).order_by('day_of_week', 'start_time')
        
        slots_data = []
        for slot in timetable_slots:
            slots_data.append({
                'id': slot.id,
                'unit_code': slot.unit_allocation.unit.code,
                'unit_name': slot.unit_allocation.unit.name,
                'unit_id': slot.unit_allocation.unit.id,
                'lecturer_name': slot.unit_allocation.lecturer.user.get_full_name(),
                'lecturer_id': slot.unit_allocation.lecturer.user_id,
                'venue_code': slot.venue.code,
                'venue_name': slot.venue.name,
                'venue_id': slot.venue.id,
                'day_of_week': slot.day_of_week,
                'start_time': slot.start_time.strftime('%H:%M'),
                'end_time': slot.end_time.strftime('%H:%M'),
            })
        
        return JsonResponse({
            'success': True,
            'slots': slots_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def save_timetable_slot(request):
    """
    Save a new timetable slot
    """
    try:
        data = json.loads(request.body)
        
        unit_id = data.get('unit_id')
        lecturer_id = data.get('lecturer_id')
        venue_id = data.get('venue_id')
        programme_id = data.get('programme_id')
        year_level = data.get('year_level')
        semester_id = data.get('semester_id')
        day_of_week = data.get('day_of_week')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        # Validate required fields
        if not all([unit_id, lecturer_id, venue_id, programme_id, year_level, semester_id, day_of_week, start_time, end_time]):
            return JsonResponse({
                'success': False,
                'error': 'All fields are required'
            }, status=400)
        
        # Get objects
        unit = get_object_or_404(Unit, id=unit_id)
        lecturer = get_object_or_404(Lecturer, user_id=lecturer_id)
        venue = get_object_or_404(Venue, id=venue_id)
        programme = get_object_or_404(Programme, id=programme_id)
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Convert time strings to time objects
        start_time_obj = datetime.strptime(start_time, '%H:%M').time()
        end_time_obj = datetime.strptime(end_time, '%H:%M').time()
        
        # Check for conflicts
        # 1. Check if venue is already booked at this time
        venue_conflict = TimetableSlot.objects.filter(
            venue=venue,
            day_of_week=day_of_week,
            unit_allocation__semester=semester,
            is_active=True
        ).filter(
            Q(start_time__lt=end_time_obj, end_time__gt=start_time_obj)
        ).exists()
        
        if venue_conflict:
            return JsonResponse({
                'success': False,
                'error': f'Venue {venue.code} is already booked at this time'
            }, status=400)
        
        # 2. Check if lecturer is already allocated at this time
        lecturer_conflict = TimetableSlot.objects.filter(
            unit_allocation__lecturer=lecturer,
            day_of_week=day_of_week,
            unit_allocation__semester=semester,
            is_active=True
        ).filter(
            Q(start_time__lt=end_time_obj, end_time__gt=start_time_obj)
        ).exists()
        
        if lecturer_conflict:
            return JsonResponse({
                'success': False,
                'error': f'Lecturer is already allocated to another class at this time'
            }, status=400)
        
        # 3. Check if programme/year already has a class at this time
        programme_conflict = TimetableSlot.objects.filter(
            programme=programme,
            year_level=year_level,
            day_of_week=day_of_week,
            unit_allocation__semester=semester,
            is_active=True
        ).filter(
            Q(start_time__lt=end_time_obj, end_time__gt=start_time_obj)
        ).exists()
        
        if programme_conflict:
            return JsonResponse({
                'success': False,
                'error': f'This programme/year already has a class scheduled at this time'
            }, status=400)
        
        # Get or create unit allocation
        unit_allocation, created = UnitAllocation.objects.get_or_create(
            unit=unit,
            lecturer=lecturer,
            semester=semester,
            defaults={'is_active': True}
        )
        
        # Add programme to allocation if not already added
        if not unit_allocation.programmes.filter(id=programme_id).exists():
            unit_allocation.programmes.add(programme)
        
        # Create timetable slot
        timetable_slot = TimetableSlot.objects.create(
            unit_allocation=unit_allocation,
            venue=venue,
            day_of_week=day_of_week,
            start_time=start_time_obj,
            end_time=end_time_obj,
            programme=programme,
            year_level=year_level,
            created_by=request.user,
            is_active=True
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Timetable slot saved successfully',
            'slot': {
                'id': timetable_slot.id,
                'unit_code': unit.code,
                'unit_name': unit.name,
                'unit_id': unit.id,
                'lecturer_name': lecturer.user.get_full_name(),
                'lecturer_id': lecturer.user_id,
                'venue_code': venue.code,
                'venue_name': venue.name,
                'venue_id': venue.id,
                'day_of_week': day_of_week,
                'start_time': start_time,
                'end_time': end_time,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def delete_timetable_slot(request, slot_id):
    """
    Delete a timetable slot
    """
    try:
        slot = get_object_or_404(TimetableSlot, id=slot_id)
        slot.delete()
        
        return JsonResponse({
            'success': True,
            'message': 'Timetable slot deleted successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def update_timetable_slot(request, slot_id):
    """
    Update an existing timetable slot
    """
    try:
        data = json.loads(request.body)
        slot = get_object_or_404(TimetableSlot, id=slot_id)
        
        lecturer_id = data.get('lecturer_id')
        venue_id = data.get('venue_id')
        
        if lecturer_id:
            lecturer = get_object_or_404(Lecturer, user_id=lecturer_id)
            # Update unit allocation with new lecturer
            slot.unit_allocation.lecturer = lecturer
            slot.unit_allocation.save()
        
        if venue_id:
            venue = get_object_or_404(Venue, id=venue_id)
            slot.venue = venue
            slot.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Timetable slot updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)



from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from .models import (
    Programme, Semester, Unit, UnitAllocation, Lecturer, Venue,
    TimetableSlot, Department, ProgrammeUnit
)
import json
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


@login_required
def timetable_view(request):
    """
    View timetable for selected programme, semester, and year
    """
    # Get current semester or all semesters
    current_semester = Semester.objects.filter(is_current=True).first()
    semesters = Semester.objects.all().order_by('-start_date')[:10]
    
    # Get all active programmes grouped by department
    programmes = Programme.objects.filter(is_active=True).select_related('department').order_by('department', 'name')
    
    # Time slots configuration (8 AM to 6 PM)
    time_slots = [
        {'start': '08:00', 'end': '09:00', 'label': '08:00 - 09:00'},
        {'start': '09:00', 'end': '10:00', 'label': '09:00 - 10:00'},
        {'start': '10:00', 'end': '11:00', 'label': '10:00 - 11:00'},
        {'start': '11:00', 'end': '12:00', 'label': '11:00 - 12:00'},
        {'start': '12:00', 'end': '13:00', 'label': '12:00 - 13:00'},
        {'start': '13:00', 'end': '14:00', 'label': '13:00 - 14:00'},
        {'start': '14:00', 'end': '15:00', 'label': '14:00 - 15:00'},
        {'start': '15:00', 'end': '16:00', 'label': '15:00 - 16:00'},
        {'start': '16:00', 'end': '17:00', 'label': '16:00 - 17:00'},
        {'start': '17:00', 'end': '18:00', 'label': '17:00 - 18:00'},
    ]
    
    # Days of the week
    days_of_week = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
    
    context = {
        'current_semester': current_semester,
        'semesters': semesters,
        'programmes': programmes,
        'time_slots': time_slots,
        'days_of_week': days_of_week,
        'year_levels': [1, 2, 3, 4],
    }
    
    return render(request, 'admin/timetable_view.html', context)


@login_required
@require_http_methods(["GET"])
def get_timetable_data(request, programme_id, year_level, semester_id):
    """
    Get timetable data for viewing
    """
    try:
        programme = get_object_or_404(Programme, id=programme_id)
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Get timetable slots
        timetable_slots = TimetableSlot.objects.filter(
            programme_id=programme_id,
            year_level=year_level,
            unit_allocation__semester_id=semester_id,
            is_active=True
        ).select_related(
            'unit_allocation__unit',
            'unit_allocation__lecturer__user',
            'venue'
        ).order_by('day_of_week', 'start_time')
        
        # Time slots for the grid
        time_slots = [
            {'start': '08:00', 'end': '09:00'},
            {'start': '09:00', 'end': '10:00'},
            {'start': '10:00', 'end': '11:00'},
            {'start': '11:00', 'end': '12:00'},
            {'start': '12:00', 'end': '13:00'},
            {'start': '13:00', 'end': '14:00'},
            {'start': '14:00', 'end': '15:00'},
            {'start': '15:00', 'end': '16:00'},
            {'start': '16:00', 'end': '17:00'},
            {'start': '17:00', 'end': '18:00'},
        ]
        
        days_of_week = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
        
        # Create timetable grid
        timetable_grid = {}
        for day in days_of_week:
            timetable_grid[day] = {}
            for slot in time_slots:
                timetable_grid[day][slot['start']] = None
        
        # Fill in the scheduled classes
        for slot in timetable_slots:
            start_time = slot.start_time.strftime('%H:%M')
            if slot.day_of_week in timetable_grid and start_time in timetable_grid[slot.day_of_week]:
                timetable_grid[slot.day_of_week][start_time] = {
                    'unit_code': slot.unit_allocation.unit.code,
                    'unit_name': slot.unit_allocation.unit.name,
                    'lecturer_name': slot.unit_allocation.lecturer.user.get_full_name(),
                    'venue_code': slot.venue.code,
                    'venue_name': slot.venue.name,
                    'start_time': start_time,
                    'end_time': slot.end_time.strftime('%H:%M'),
                }
        
        return JsonResponse({
            'success': True,
            'timetable': timetable_grid,
            'programme': {
                'code': programme.code,
                'name': programme.name,
            },
            'semester': str(semester),
            'year_level': year_level,
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
def export_timetable_excel(request):
    """
    Export timetable to Excel format
    """
    try:
        programme_id = request.GET.get('programme_id')
        year_level = request.GET.get('year_level')
        semester_id = request.GET.get('semester_id')
        
        if not all([programme_id, year_level, semester_id]):
            return HttpResponse('Missing parameters', status=400)
        
        programme = get_object_or_404(Programme, id=programme_id)
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Get timetable slots
        timetable_slots = TimetableSlot.objects.filter(
            programme_id=programme_id,
            year_level=year_level,
            unit_allocation__semester_id=semester_id,
            is_active=True
        ).select_related(
            'unit_allocation__unit',
            'unit_allocation__lecturer__user',
            'venue'
        ).order_by('day_of_week', 'start_time')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Define styles
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        day_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")
        day_font = Font(bold=True, color="FFFFFF", size=11)
        time_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        time_font = Font(bold=True, size=10)
        cell_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )
        
        # Title Section
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "FACULTY OF BUSINESS - TIMETABLE"
        title_cell.font = Font(bold=True, size=16, color="1E293B")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Programme Info
        ws.merge_cells('A2:H2')
        prog_cell = ws['A2']
        prog_cell.value = f"Programme: {programme.code} - {programme.name}"
        prog_cell.font = Font(bold=True, size=12, color="475569")
        prog_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Semester and Year Info
        ws.merge_cells('A3:H3')
        sem_cell = ws['A3']
        sem_cell.value = f"Semester: {semester} | Year: {year_level}"
        sem_cell.font = Font(bold=True, size=11, color="475569")
        sem_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add spacing
        ws.row_dimensions[4].height = 5
        
        # Time slots
        time_slots = [
            '08:00 - 09:00', '09:00 - 10:00', '10:00 - 11:00', '11:00 - 12:00',
            '12:00 - 13:00', '13:00 - 14:00', '14:00 - 15:00', '15:00 - 16:00',
            '16:00 - 17:00', '17:00 - 18:00'
        ]
        
        days_of_week = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY']
        
        # Create timetable grid dictionary
        timetable_grid = {}
        for day in days_of_week:
            timetable_grid[day] = {}
            for slot in time_slots:
                timetable_grid[day][slot] = None
        
        # Fill in the scheduled classes
        for slot in timetable_slots:
            start_time = slot.start_time.strftime('%H:%M')
            end_time = slot.end_time.strftime('%H:%M')
            time_label = f"{start_time} - {end_time}"
            
            if slot.day_of_week in timetable_grid and time_label in timetable_grid[slot.day_of_week]:
                timetable_grid[slot.day_of_week][time_label] = {
                    'unit_code': slot.unit_allocation.unit.code,
                    'unit_name': slot.unit_allocation.unit.name,
                    'lecturer': slot.unit_allocation.lecturer.user.get_full_name(),
                    'venue': slot.venue.code,
                }
        
        # Header row (Days)
        start_row = 5
        ws.cell(row=start_row, column=1, value="Time")
        ws.cell(row=start_row, column=1).fill = header_fill
        ws.cell(row=start_row, column=1).font = header_font
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=start_row, column=1).border = cell_border
        
        for col_idx, day in enumerate(days_of_week, start=2):
            cell = ws.cell(row=start_row, column=col_idx, value=day.title())
            cell.fill = day_fill
            cell.font = day_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = cell_border
        
        # Fill timetable data
        current_row = start_row + 1
        for time_slot in time_slots:
            # Time column
            time_cell = ws.cell(row=current_row, column=1, value=time_slot)
            time_cell.fill = time_fill
            time_cell.font = time_font
            time_cell.alignment = Alignment(horizontal='center', vertical='center')
            time_cell.border = cell_border
            
            # Day columns
            for col_idx, day in enumerate(days_of_week, start=2):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = cell_border
                
                class_info = timetable_grid[day].get(time_slot)
                if class_info:
                    cell_value = f"{class_info['unit_code']}\n{class_info['unit_name']}\n{class_info['lecturer']}\n{class_info['venue']}"
                    cell.value = cell_value
                    cell.font = Font(size=9)
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.fill = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
                else:
                    cell.value = ""
            
            ws.row_dimensions[current_row].height = 60
            current_row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 8):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"Timetable_{programme.code}_Year{year_level}_{semester.academic_year.year_code}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return HttpResponse(f'Error generating Excel: {str(e)}', status=400)
    

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.contrib.auth.hashers import make_password
from django.core.paginator import Paginator
from .models import User, Student, Lecturer, Department, Programme
import json
from datetime import datetime


@login_required
def user_management_view(request):
    """
    Main user management view
    """
    departments = Department.objects.all().order_by('name')
    
    context = {
        'departments': departments,
        'user_types': User.USER_TYPES,
    }
    
    return render(request, 'admin/user_management.html', context)


@login_required
@require_http_methods(["GET"])
def get_users_list(request):
    """
    Get paginated list of users with filtering and search
    """
    try:
        # Get query parameters
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        search = request.GET.get('search', '').strip()
        user_type = request.GET.get('user_type', '')
        is_active = request.GET.get('is_active', '')
        department_id = request.GET.get('department', '')
        
        # Start with all users
        users = User.objects.all()
        
        # Apply search filter
        if search:
            users = users.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone_number__icontains=search)
            )
        
        # Apply user type filter
        if user_type:
            users = users.filter(user_type=user_type)
        
        # Apply active status filter
        if is_active:
            users = users.filter(is_active_user=(is_active == 'true'))
        
        # Apply department filter for students and lecturers
        if department_id:
            users = users.filter(
                Q(student_profile__programme__department_id=department_id) |
                Q(lecturer_profile__department_id=department_id)
            )
        
        # Order by date joined (newest first)
        users = users.order_by('-date_joined')
        
        # Get total count before pagination
        total_count = users.count()
        
        # Paginate
        paginator = Paginator(users, per_page)
        page_obj = paginator.get_page(page)
        
        # Prepare user data
        users_data = []
        for user in page_obj:
            user_data = {
                'id': user.id,
                'username': user.username,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'full_name': user.get_full_name() or user.username,
                'email': user.email,
                'phone_number': user.phone_number or '',
                'user_type': user.user_type,
                'user_type_display': user.get_user_type_display(),
                'is_active': user.is_active_user,
                'is_staff': user.is_staff,
                'is_superuser': user.is_superuser,
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
                'last_login': user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Never',
            }
            
            # Add specific profile info
            if user.user_type == 'STUDENT' and hasattr(user, 'student_profile'):
                student = user.student_profile
                user_data['registration_number'] = student.registration_number
                user_data['programme'] = student.programme.code
                user_data['year'] = student.current_year
            elif user.user_type == 'LECTURER' and hasattr(user, 'lecturer_profile'):
                lecturer = user.lecturer_profile
                user_data['staff_number'] = lecturer.staff_number
                user_data['department'] = lecturer.department.code
            
            users_data.append(user_data)
        
        return JsonResponse({
            'success': True,
            'users': users_data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_count': total_count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'per_page': per_page,
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def get_user_detail(request, user_id):
    """
    Get detailed information about a specific user
    """
    try:
        user = get_object_or_404(User, id=user_id)
        
        user_data = {
            'id': user.id,
            'username': user.username,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'email': user.email,
            'phone_number': user.phone_number or '',
            'user_type': user.user_type,
            'user_type_display': user.get_user_type_display(),
            'is_active': user.is_active_user,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
            'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M:%S'),
            'last_login': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else None,
            'profile_picture': user.profile_picture.url if user.profile_picture else None,
        }
        
        # Add specific profile info
        if user.user_type == 'STUDENT' and hasattr(user, 'student_profile'):
            student = user.student_profile
            user_data['student_profile'] = {
                'registration_number': student.registration_number,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'surname': student.surname,
                'programme_id': student.programme.id,
                'programme_name': str(student.programme),
                'current_year': student.current_year,
                'intake_id': student.intake.id,
                'intake_name': str(student.intake),
                'admission_date': student.admission_date.strftime('%Y-%m-%d'),
                'is_active': student.is_active,
                'phone': student.phone or '',
                'email': student.email or '',
                'date_of_birth': student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else '',
                'address': student.address or '',
                'parent_name': student.parent_name or '',
                'parent_phone': student.parent_phone or '',
                'guardian_name': student.guardian_name or '',
                'guardian_phone': student.guardian_phone or '',
            }
        elif user.user_type == 'LECTURER' and hasattr(user, 'lecturer_profile'):
            lecturer = user.lecturer_profile
            user_data['lecturer_profile'] = {
                'staff_number': lecturer.staff_number,
                'department_id': lecturer.department.id,
                'department_name': str(lecturer.department),
                'specialization': lecturer.specialization,
                'office_location': lecturer.office_location,
                'consultation_hours': lecturer.consultation_hours,
                'is_active': lecturer.is_active,
            }
        
        return JsonResponse({
            'success': True,
            'user': user_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def update_user(request, user_id):
    """
    Update user information
    """
    try:
        user = get_object_or_404(User, id=user_id)
        data = json.loads(request.body)
        
        # Update basic user fields
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'email' in data:
            user.email = data['email']
        if 'phone_number' in data:
            user.phone_number = data['phone_number']
        if 'is_active' in data:
            user.is_active_user = data['is_active']
        if 'is_staff' in data:
            user.is_staff = data['is_staff']
        if 'is_superuser' in data:
            user.is_superuser = data['is_superuser']
        
        user.save()
        
        # Update profile-specific fields
        if user.user_type == 'STUDENT' and hasattr(user, 'student_profile'):
            student = user.student_profile
            if 'student_profile' in data:
                profile_data = data['student_profile']
                if 'first_name' in profile_data:
                    student.first_name = profile_data['first_name']
                if 'last_name' in profile_data:
                    student.last_name = profile_data['last_name']
                if 'surname' in profile_data:
                    student.surname = profile_data['surname']
                if 'phone' in profile_data:
                    student.phone = profile_data['phone']
                if 'email' in profile_data:
                    student.email = profile_data['email']
                if 'address' in profile_data:
                    student.address = profile_data['address']
                if 'date_of_birth' in profile_data:
                    student.date_of_birth = profile_data['date_of_birth']
                if 'parent_name' in profile_data:
                    student.parent_name = profile_data['parent_name']
                if 'parent_phone' in profile_data:
                    student.parent_phone = profile_data['parent_phone']
                if 'guardian_name' in profile_data:
                    student.guardian_name = profile_data['guardian_name']
                if 'guardian_phone' in profile_data:
                    student.guardian_phone = profile_data['guardian_phone']
                if 'is_active' in profile_data:
                    student.is_active = profile_data['is_active']
                
                student.save()
        
        elif user.user_type == 'LECTURER' and hasattr(user, 'lecturer_profile'):
            lecturer = user.lecturer_profile
            if 'lecturer_profile' in data:
                profile_data = data['lecturer_profile']
                if 'specialization' in profile_data:
                    lecturer.specialization = profile_data['specialization']
                if 'office_location' in profile_data:
                    lecturer.office_location = profile_data['office_location']
                if 'consultation_hours' in profile_data:
                    lecturer.consultation_hours = profile_data['consultation_hours']
                if 'is_active' in profile_data:
                    lecturer.is_active = profile_data['is_active']
                
                lecturer.save()
        
        return JsonResponse({
            'success': True,
            'message': 'User updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def change_user_password(request, user_id):
    """
    Change user password
    """
    try:
        user = get_object_or_404(User, id=user_id)
        data = json.loads(request.body)
        
        new_password = data.get('new_password', '').strip()
        
        if not new_password:
            return JsonResponse({
                'success': False,
                'error': 'Password cannot be empty'
            }, status=400)
        
        if len(new_password) < 6:
            return JsonResponse({
                'success': False,
                'error': 'Password must be at least 6 characters long'
            }, status=400)
        
        # Set new password
        user.set_password(new_password)
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Password changed successfully for {user.username}'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def bulk_update_users(request):
    """
    Bulk update user status or type
    """
    try:
        data = json.loads(request.body)
        user_ids = data.get('user_ids', [])
        action = data.get('action', '')
        value = data.get('value', None)
        
        if not user_ids:
            return JsonResponse({
                'success': False,
                'error': 'No users selected'
            }, status=400)
        
        users = User.objects.filter(id__in=user_ids)
        
        if action == 'activate':
            users.update(is_active_user=True)
            message = f'{users.count()} users activated'
        elif action == 'deactivate':
            users.update(is_active_user=False)
            message = f'{users.count()} users deactivated'
        elif action == 'make_staff':
            users.update(is_staff=True)
            message = f'{users.count()} users made staff'
        elif action == 'remove_staff':
            users.update(is_staff=False)
            message = f'{users.count()} users removed from staff'
        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action'
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def delete_user(request, user_id):
    """
    Delete a user (soft delete by deactivating)
    """
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Prevent deleting superuser
        if user.is_superuser:
            return JsonResponse({
                'success': False,
                'error': 'Cannot delete superuser account'
            }, status=400)
        
        # Soft delete
        user.is_active_user = False
        user.is_active = False
        user.save()
        
        return JsonResponse({
            'success': True,
            'message': f'User {user.username} has been deactivated'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["GET"])
def get_programmes_by_department(request, department_id):
    """
    Get programmes for a specific department
    """
    try:
        programmes = Programme.objects.filter(
            department_id=department_id,
            is_active=True
        ).order_by('name')
        
        programmes_data = [
            {
                'id': prog.id,
                'code': prog.code,
                'name': prog.name,
                'level': prog.level
            }
            for prog in programmes
        ]
        
        return JsonResponse({
            'success': True,
            'programmes': programmes_data
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)
    

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.hashers import make_password
from django.db import transaction
from .models import User, Student, Lecturer, Department, Programme, Intake, AcademicYear
import json
from datetime import datetime

def is_admin(user):
    """Check if user is admin/ICT admin"""
    return user.is_superuser or user.user_type == 'ICT_ADMIN'

@login_required
@user_passes_test(is_admin)
def create_user_view(request):
    """
    User creation view - Admin only
    Handles all user types: STUDENT, LECTURER, COD, DEAN, ICT_ADMIN
    """
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get basic user data
                username = request.POST.get('username').strip()
                email = request.POST.get('email').strip()
                first_name = request.POST.get('first_name').strip()
                last_name = request.POST.get('last_name').strip()
                password = request.POST.get('password')
                phone_number = request.POST.get('phone_number', '').strip()
                user_type = request.POST.get('user_type')
                is_staff = request.POST.get('is_staff') == 'on'
                is_superuser = request.POST.get('is_superuser') == 'on'
                
                # Validate required fields
                if not all([username, email, first_name, last_name, password, user_type]):
                    return JsonResponse({
                        'success': False,
                        'error': 'All required fields must be filled'
                    }, status=400)
                
                # Validate user type
                valid_user_types = [choice[0] for choice in User.USER_TYPES]
                if user_type not in valid_user_types:
                    return JsonResponse({
                        'success': False,
                        'error': 'Invalid user type'
                    }, status=400)
                
                # Check if username exists
                if User.objects.filter(username=username).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Username already exists'
                    }, status=400)
                
                # Check if email exists
                if User.objects.filter(email=email).exists():
                    return JsonResponse({
                        'success': False,
                        'error': 'Email already exists'
                    }, status=400)
                
                # Create base user
                user = User.objects.create(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=last_name,
                    phone_number=phone_number,
                    user_type=user_type,
                    is_staff=is_staff,
                    is_superuser=is_superuser,
                    is_active_user=True,
                    is_active=True
                )
                user.set_password(password)
                user.save()
                
                # Create profile based on user type
                if user_type == 'STUDENT':
                    registration_number = request.POST.get('registration_number', '').strip()
                    programme_id = request.POST.get('programme')
                    current_year = request.POST.get('current_year')
                    intake_id = request.POST.get('intake')
                    admission_date = request.POST.get('admission_date')
                    
                    # Validate student required fields
                    if not all([registration_number, programme_id, current_year, intake_id, admission_date]):
                        user.delete()  # Rollback user creation
                        return JsonResponse({
                            'success': False,
                            'error': 'All student fields are required'
                        }, status=400)
                    
                    # Check if registration number exists
                    if Student.objects.filter(registration_number=registration_number).exists():
                        user.delete()
                        return JsonResponse({
                            'success': False,
                            'error': 'Registration number already exists'
                        }, status=400)
                    
                    # Create student profile
                    Student.objects.create(
                        user=user,
                        registration_number=registration_number,
                        first_name=first_name,
                        last_name=last_name,
                        surname=request.POST.get('surname', ''),
                        programme_id=programme_id,
                        current_year=int(current_year),
                        intake_id=intake_id,
                        admission_date=admission_date,
                        phone=request.POST.get('student_phone', ''),
                        email=email,
                        date_of_birth=request.POST.get('date_of_birth') or None,
                        address=request.POST.get('address', ''),
                        parent_name=request.POST.get('parent_name', ''),
                        parent_phone=request.POST.get('parent_phone', ''),
                        guardian_name=request.POST.get('guardian_name', ''),
                        guardian_phone=request.POST.get('guardian_phone', ''),
                        is_active=True
                    )
                    profile_created = "Student profile"
                
                elif user_type in ['LECTURER', 'COD', 'DEAN']:
                    staff_number = request.POST.get('staff_number', '').strip()
                    department_id = request.POST.get('department')
                    
                    # Validate lecturer/staff required fields
                    if not all([staff_number, department_id]):
                        user.delete()
                        return JsonResponse({
                            'success': False,
                            'error': 'Staff number and department are required for academic staff'
                        }, status=400)
                    
                    # Check if staff number exists
                    if Lecturer.objects.filter(staff_number=staff_number).exists():
                        user.delete()
                        return JsonResponse({
                            'success': False,
                            'error': 'Staff number already exists'
                        }, status=400)
                    
                    # Create lecturer profile (COD and DEAN also use Lecturer model)
                    Lecturer.objects.create(
                        user=user,
                        staff_number=staff_number,
                        department_id=department_id,
                        specialization=request.POST.get('specialization', ''),
                        office_location=request.POST.get('office_location', ''),
                        consultation_hours=request.POST.get('consultation_hours', ''),
                        is_active=True
                    )
                    
                    # If COD, update department head
                    if user_type == 'COD':
                        department = Department.objects.get(id=department_id)
                        department.head_of_department = user
                        department.save()
                        profile_created = "Chairman of Department profile"
                    elif user_type == 'DEAN':
                        profile_created = "Dean profile"
                    else:
                        profile_created = "Lecturer profile"
                
                elif user_type == 'ICT_ADMIN':
                    # ICT Admin doesn't need additional profile
                    # Just ensure they have staff and superuser status
                    user.is_staff = True
                    user.is_superuser = True
                    user.save()
                    profile_created = "ICT Administrator account"
                
                return JsonResponse({
                    'success': True,
                    'message': f'{profile_created} created successfully for {username}',
                    'user_id': user.id,
                    'user_type': user_type
                })
                
        except Department.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Selected department does not exist'
            }, status=400)
        except Programme.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Selected programme does not exist'
            }, status=400)
        except Intake.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Selected intake does not exist'
            }, status=400)
        except ValueError as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Error creating user: {str(e)}'
            }, status=500)
    
    # GET request - show form
    context = {
        'departments': Department.objects.all().order_by('name'),
        'programmes': Programme.objects.filter(is_active=True).order_by('name'),
        'intakes': Intake.objects.filter(is_active=True).order_by('-intake_date'),
        'user_types': User.USER_TYPES,
        'year_levels': Student.YEAR_LEVELS,
    }
    return render(request, 'admin/create_user.html', context)


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db.models import Count, Q
from django.utils import timezone
from datetime import timedelta
from django.core.paginator import Paginator
from .models import (
    AuditLog, SecurityEvent, LoginAttempt, UserSession, 
    SystemSettings, BlockedIP, DataExportLog, User
)
import json
from .utils import get_client_ip


def is_admin(user):
    """Check if user is admin/ICT admin"""
    return user.is_superuser or getattr(user, 'user_type', '') == 'ICT_ADMIN'


@login_required
@user_passes_test(is_admin)
def security_dashboard(request):
    """Main security dashboard view"""
    settings = SystemSettings.get_settings()
    
    # Get date range for statistics
    end_date = timezone.now()
    start_date = end_date - timedelta(days=30)
    
    # Basic statistics
    context = {
        'settings': settings,
        'total_users': User.objects.count(),
        'active_sessions': UserSession.objects.filter(is_active=True).count(),
        'failed_logins_today': LoginAttempt.objects.filter(
            success=False,
            timestamp__date=timezone.now().date()
        ).count(),
        'security_events_today': SecurityEvent.objects.filter(
            detected_at__date=timezone.now().date()
        ).count(),
        'blocked_ips': BlockedIP.objects.filter(is_active=True).count(),
        'critical_events': SecurityEvent.objects.filter(
            risk_level='CRITICAL',
            status__in=['DETECTED', 'INVESTIGATING']
        ).count(),
    }
    
    return render(request, 'admin/security_dashboard.html', context)


@login_required
@user_passes_test(is_admin)
def get_realtime_metrics(request):
    """AJAX endpoint for real-time metrics"""
    try:
        now = timezone.now()
        
        # Active users (logged in within last 5 minutes)
        active_users = UserSession.objects.filter(
            is_active=True,
            last_activity__gte=now - timedelta(minutes=5)
        ).count()
        
        # Failed login attempts (last hour)
        failed_logins_hour = LoginAttempt.objects.filter(
            success=False,
            timestamp__gte=now - timedelta(hours=1)
        ).count()
        
        # Security events (last hour)
        security_events_hour = SecurityEvent.objects.filter(
            detected_at__gte=now - timedelta(hours=1)
        ).count()
        
        # Critical events (unresolved)
        critical_events = SecurityEvent.objects.filter(
            risk_level='CRITICAL',
            status__in=['DETECTED', 'INVESTIGATING']
        ).count()
        
        # Recent audit logs (last 10)
        recent_audits = AuditLog.objects.select_related('user').order_by('-timestamp')[:10]
        audit_list = [{
            'user': audit.username,
            'user_type': audit.user_type or 'N/A',
            'action': audit.get_action_type_display(),
            'description': audit.action_description,
            'timestamp': audit.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'ip_address': audit.ip_address or 'N/A',
            'severity': audit.severity
        } for audit in recent_audits]
        
        return JsonResponse({
            'success': True,
            'metrics': {
                'active_users': active_users,
                'failed_logins_hour': failed_logins_hour,
                'security_events_hour': security_events_hour,
                'critical_events': critical_events,
            },
            'recent_audits': audit_list,
            'timestamp': now.strftime('%Y-%m-%d %H:%M:%S')
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_admin)
def get_login_activity_chart(request):
    """AJAX endpoint for login activity chart data"""
    try:
        days = int(request.GET.get('days', 7))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        # Initialize data structures
        dates = []
        successful = []
        failed = []
        
        # Generate date range
        current_date = start_date.date()
        while current_date <= end_date.date():
            dates.append(current_date.strftime('%b %d'))
            
            # Count successful logins for this date
            success_count = LoginAttempt.objects.filter(
                timestamp__date=current_date,
                success=True
            ).count()
            
            # Count failed logins for this date
            fail_count = LoginAttempt.objects.filter(
                timestamp__date=current_date,
                success=False
            ).count()
            
            successful.append(success_count)
            failed.append(fail_count)
            
            current_date += timedelta(days=1)
        
        return JsonResponse({
            'success': True,
            'labels': dates,
            'datasets': [
                {
                    'label': 'Successful Logins',
                    'data': successful,
                    'borderColor': '#10B981',
                    'backgroundColor': 'rgba(16, 185, 129, 0.1)',
                    'tension': 0.4,
                    'fill': True,
                },
                {
                    'label': 'Failed Logins',
                    'data': failed,
                    'borderColor': '#EF4444',
                    'backgroundColor': 'rgba(239, 68, 68, 0.1)',
                    'tension': 0.4,
                    'fill': True,
                }
            ]
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_admin)
def get_user_activity_chart(request):
    """AJAX endpoint for user activity by type"""
    try:
        days = int(request.GET.get('days', 7))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        # Get activity by user type
        user_activity = AuditLog.objects.filter(
            timestamp__gte=start_date
        ).values('user_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        labels = []
        data = []
        colors = []
        
        color_map = {
            'STUDENT': '#3B82F6',
            'LECTURER': '#10B981',
            'COD': '#F59E0B',
            'DEAN': '#8B5CF6',
            'ICT_ADMIN': '#EF4444',
        }
        
        for item in user_activity:
            if item['user_type']:
                labels.append(item['user_type'])
                data.append(item['count'])
                colors.append(color_map.get(item['user_type'], '#64748B'))
        
        # If no data, return empty state
        if not labels:
            labels = ['No Data']
            data = [0]
            colors = ['#E2E8F0']
        
        return JsonResponse({
            'success': True,
            'labels': labels,
            'datasets': [{
                'data': data,
                'backgroundColor': colors,
            }]
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_admin)
def get_security_events_chart(request):
    """AJAX endpoint for security events by risk level"""
    try:
        days = int(request.GET.get('days', 30))
        end_date = timezone.now()
        start_date = end_date - timedelta(days=days)
        
        # Get security events by risk level
        events = SecurityEvent.objects.filter(
            detected_at__gte=start_date
        ).values('risk_level').annotate(
            count=Count('id')
        )
        
        risk_order = ['LOW', 'MEDIUM', 'HIGH', 'CRITICAL']
        labels = []
        data = []
        colors = ['#10B981', '#F59E0B', '#EF4444', '#7C3AED']
        
        for risk in risk_order:
            event = next((e for e in events if e['risk_level'] == risk), None)
            labels.append(risk)
            data.append(event['count'] if event else 0)
        
        return JsonResponse({
            'success': True,
            'labels': labels,
            'datasets': [{
                'label': 'Security Events',
                'data': data,
                'backgroundColor': colors,
            }]
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def toggle_maintenance_mode(request):
    """AJAX endpoint to toggle maintenance mode"""
    try:
        settings = SystemSettings.get_settings()
        data = json.loads(request.body)
        
        enable = data.get('enable', False)
        message = data.get('message', 'System is currently under maintenance.')
        
        settings.maintenance_mode = enable
        settings.maintenance_message = message
        settings.updated_by = request.user
        
        if enable:
            settings.maintenance_started_at = timezone.now()
            settings.maintenance_started_by = request.user
        
        settings.save()
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            user_type=getattr(request.user, 'user_type', ''),
            username=request.user.username,
            action_type='UPDATE',
            action_description=f"Maintenance mode {'enabled' if enable else 'disabled'}",
            model_name='SystemSettings',
            ip_address=get_client_ip(request),
            severity='HIGH'
        )
        
        return JsonResponse({
            'success': True,
            'message': f"Maintenance mode {'enabled' if enable else 'disabled'} successfully",
            'maintenance_mode': enable
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_admin)
@require_http_methods(["POST"])
def update_security_settings(request):
    """AJAX endpoint to update security settings"""
    try:
        settings = SystemSettings.get_settings()
        data = json.loads(request.body)
        
        # Update settings
        settings.max_login_attempts = int(data.get('max_login_attempts', 5))
        settings.lockout_duration_minutes = int(data.get('lockout_duration_minutes', 30))
        settings.session_timeout_minutes = int(data.get('session_timeout_minutes', 60))
        settings.enable_audit_logging = data.get('enable_audit_logging', True)
        settings.log_view_actions = data.get('log_view_actions', False)
        settings.updated_by = request.user
        settings.save()
        
        # Log the action
        AuditLog.objects.create(
            user=request.user,
            user_type=getattr(request.user, 'user_type', ''),
            username=request.user.username,
            action_type='UPDATE',
            action_description="Security settings updated",
            model_name='SystemSettings',
            new_values=data,
            ip_address=get_client_ip(request),
            severity='HIGH'
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Security settings updated successfully'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(is_admin)
def get_audit_logs(request):
    """AJAX endpoint to fetch audit logs with filters"""
    try:
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 50))
        
        # Filters
        user_type = request.GET.get('user_type', '')
        action_type = request.GET.get('action_type', '')
        severity = request.GET.get('severity', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        search = request.GET.get('search', '')
        
        # Build query
        logs = AuditLog.objects.select_related('user').all()
        
        if user_type:
            logs = logs.filter(user_type=user_type)
        if action_type:
            logs = logs.filter(action_type=action_type)
        if severity:
            logs = logs.filter(severity=severity)
        if date_from:
            logs = logs.filter(timestamp__gte=date_from)
        if date_to:
            logs = logs.filter(timestamp__lte=date_to)
        if search:
            logs = logs.filter(
                Q(username__icontains=search) |
                Q(action_description__icontains=search) |
                Q(model_name__icontains=search)
            )
        
        # Paginate
        paginator = Paginator(logs, per_page)
        page_obj = paginator.get_page(page)
        
        # Serialize data
        logs_data = [{
            'id': log.id,
            'username': log.username,
            'user_type': log.user_type or 'N/A',
            'action_type': log.get_action_type_display(),
            'action_description': log.action_description,
            'model_name': log.model_name or 'N/A',
            'ip_address': log.ip_address or 'N/A',
            'timestamp': log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'severity': log.severity,
        } for log in page_obj]
        
        return JsonResponse({
            'success': True,
            'logs': logs_data,
            'pagination': {
                'current_page': page,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            }
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count
from django.utils import timezone
from datetime import datetime
from .models import AcademicYear, Semester
import json


@login_required
def academic_management(request):
    """Main view for academic year and semester management"""
    academic_years = AcademicYear.objects.all().annotate(
        semester_count=Count('semesters')
    )
    
    context = {
        'academic_years': academic_years,
    }
    return render(request, 'admin/academic_management.html', context)


@login_required
@require_http_methods(["GET"])
def get_academic_year(request, year_id):
    """Get single academic year details"""
    try:
        academic_year = AcademicYear.objects.get(id=year_id)
        data = {
            'id': academic_year.id,
            'year_code': academic_year.year_code,
            'start_date': academic_year.start_date.strftime('%Y-%m-%d'),
            'end_date': academic_year.end_date.strftime('%Y-%m-%d'),
            'is_current': academic_year.is_current,
        }
        return JsonResponse({'success': True, 'data': data})
    except AcademicYear.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Academic year not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def create_academic_year(request):
    """Create new academic year"""
    try:
        data = json.loads(request.body)
        
        year_code = data.get('year_code')
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d').date()
        is_current = data.get('is_current', False)
        
        # Validation
        if not year_code or not start_date or not end_date:
            return JsonResponse({'success': False, 'message': 'All fields are required'}, status=400)
        
        if start_date >= end_date:
            return JsonResponse({'success': False, 'message': 'End date must be after start date'}, status=400)
        
        # Check if year_code already exists
        if AcademicYear.objects.filter(year_code=year_code).exists():
            return JsonResponse({'success': False, 'message': 'Academic year code already exists'}, status=400)
        
        # If marking as current, unset other current years
        if is_current:
            AcademicYear.objects.filter(is_current=True).update(is_current=False)
        
        academic_year = AcademicYear.objects.create(
            year_code=year_code,
            start_date=start_date,
            end_date=end_date,
            is_current=is_current
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Academic year created successfully',
            'data': {
                'id': academic_year.id,
                'year_code': academic_year.year_code,
                'start_date': academic_year.start_date.strftime('%Y-%m-%d'),
                'end_date': academic_year.end_date.strftime('%Y-%m-%d'),
                'is_current': academic_year.is_current,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_academic_year(request, year_id):
    """Update existing academic year"""
    try:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        data = json.loads(request.body)
        
        year_code = data.get('year_code')
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d').date()
        is_current = data.get('is_current', False)
        
        # Validation
        if start_date >= end_date:
            return JsonResponse({'success': False, 'message': 'End date must be after start date'}, status=400)
        
        # Check if year_code already exists (excluding current)
        if AcademicYear.objects.filter(year_code=year_code).exclude(id=year_id).exists():
            return JsonResponse({'success': False, 'message': 'Academic year code already exists'}, status=400)
        
        # If marking as current, unset other current years
        if is_current:
            AcademicYear.objects.filter(is_current=True).exclude(id=year_id).update(is_current=False)
        
        academic_year.year_code = year_code
        academic_year.start_date = start_date
        academic_year.end_date = end_date
        academic_year.is_current = is_current
        academic_year.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Academic year updated successfully',
            'data': {
                'id': academic_year.id,
                'year_code': academic_year.year_code,
                'start_date': academic_year.start_date.strftime('%Y-%m-%d'),
                'end_date': academic_year.end_date.strftime('%Y-%m-%d'),
                'is_current': academic_year.is_current,
            }
        })
    except AcademicYear.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Academic year not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_academic_year(request, year_id):
    """Delete academic year"""
    try:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        
        # Check if there are related semesters
        if academic_year.semesters.exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete academic year with existing semesters. Delete semesters first.'
            }, status=400)
        
        academic_year.delete()
        return JsonResponse({'success': True, 'message': 'Academic year deleted successfully'})
    except AcademicYear.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Academic year not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ========================
# SEMESTER MANAGEMENT
# ========================

@login_required
@require_http_methods(["GET"])
def get_semesters(request, year_id):
    """Get all semesters for an academic year"""
    try:
        semesters = Semester.objects.filter(academic_year_id=year_id)
        data = [{
            'id': sem.id,
            'semester_number': sem.semester_number,
            'semester_display': sem.get_semester_number_display(),
            'start_date': sem.start_date.strftime('%Y-%m-%d'),
            'end_date': sem.end_date.strftime('%Y-%m-%d'),
            'registration_deadline': sem.registration_deadline.strftime('%Y-%m-%d'),
            'is_current': sem.is_current,
        } for sem in semesters]
        
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_semester(request, semester_id):
    """Get single semester details"""
    try:
        semester = Semester.objects.get(id=semester_id)
        data = {
            'id': semester.id,
            'academic_year_id': semester.academic_year.id,
            'semester_number': semester.semester_number,
            'start_date': semester.start_date.strftime('%Y-%m-%d'),
            'end_date': semester.end_date.strftime('%Y-%m-%d'),
            'registration_deadline': semester.registration_deadline.strftime('%Y-%m-%d'),
            'is_current': semester.is_current,
        }
        return JsonResponse({'success': True, 'data': data})
    except Semester.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Semester not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def create_semester(request, year_id):
    """Create new semester"""
    try:
        academic_year = get_object_or_404(AcademicYear, id=year_id)
        data = json.loads(request.body)
        
        semester_number = int(data.get('semester_number'))
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d').date()
        registration_deadline = datetime.strptime(data.get('registration_deadline'), '%Y-%m-%d').date()
        is_current = data.get('is_current', False)
        
        # Validation
        if start_date >= end_date:
            return JsonResponse({'success': False, 'message': 'End date must be after start date'}, status=400)
        
        if registration_deadline > start_date:
            return JsonResponse({'success': False, 'message': 'Registration deadline must be before start date'}, status=400)
        
        # Check if semester already exists
        if Semester.objects.filter(academic_year=academic_year, semester_number=semester_number).exists():
            return JsonResponse({'success': False, 'message': 'Semester already exists for this academic year'}, status=400)
        
        # If marking as current, unset other current semesters
        if is_current:
            Semester.objects.filter(is_current=True).update(is_current=False)
        
        semester = Semester.objects.create(
            academic_year=academic_year,
            semester_number=semester_number,
            start_date=start_date,
            end_date=end_date,
            registration_deadline=registration_deadline,
            is_current=is_current
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Semester created successfully',
            'data': {
                'id': semester.id,
                'semester_number': semester.semester_number,
                'semester_display': semester.get_semester_number_display(),
                'start_date': semester.start_date.strftime('%Y-%m-%d'),
                'end_date': semester.end_date.strftime('%Y-%m-%d'),
                'registration_deadline': semester.registration_deadline.strftime('%Y-%m-%d'),
                'is_current': semester.is_current,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_semester(request, semester_id):
    """Update existing semester"""
    try:
        semester = get_object_or_404(Semester, id=semester_id)
        data = json.loads(request.body)
        
        semester_number = int(data.get('semester_number'))
        start_date = datetime.strptime(data.get('start_date'), '%Y-%m-%d').date()
        end_date = datetime.strptime(data.get('end_date'), '%Y-%m-%d').date()
        registration_deadline = datetime.strptime(data.get('registration_deadline'), '%Y-%m-%d').date()
        is_current = data.get('is_current', False)
        
        # Validation
        if start_date >= end_date:
            return JsonResponse({'success': False, 'message': 'End date must be after start date'}, status=400)
        
        if registration_deadline > start_date:
        # No restriction — just allow it to proceed
        # return JsonResponse({'success': False, 'message': 'Registration deadline must be before start date'}, status=400)
          pass

        # Check if semester number already exists (excluding current)
        if Semester.objects.filter(
            academic_year=semester.academic_year,
            semester_number=semester_number
        ).exclude(id=semester_id).exists():
            return JsonResponse({'success': False, 'message': 'Semester number already exists for this academic year'}, status=400)
        
        # If marking as current, unset other current semesters
        if is_current:
            Semester.objects.filter(is_current=True).exclude(id=semester_id).update(is_current=False)
        
        semester.semester_number = semester_number
        semester.start_date = start_date
        semester.end_date = end_date
        semester.registration_deadline = registration_deadline
        semester.is_current = is_current
        semester.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Semester updated successfully',
            'data': {
                'id': semester.id,
                'semester_number': semester.semester_number,
                'semester_display': semester.get_semester_number_display(),
                'start_date': semester.start_date.strftime('%Y-%m-%d'),
                'end_date': semester.end_date.strftime('%Y-%m-%d'),
                'registration_deadline': semester.registration_deadline.strftime('%Y-%m-%d'),
                'is_current': semester.is_current,
            }
        })
    except Semester.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Semester not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_semester(request, semester_id):
    """Delete semester"""
    try:
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Check if there are related records
        if semester.enrollments.exists() or semester.registrations.exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete semester with existing enrollments or registrations.'
            }, status=400)
        
        semester.delete()
        return JsonResponse({'success': True, 'message': 'Semester deleted successfully'})
    except Semester.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Semester not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator
from .models import Unit, Department, ProgrammeUnit, Programme
import json


@login_required
def unit_management(request):
    """Main view for unit management organized by programmes"""
    # Get all programmes with their units
    programmes = Programme.objects.filter(is_active=True).prefetch_related(
        Prefetch(
            'programme_units',
            queryset=ProgrammeUnit.objects.select_related('unit', 'unit__department')
        )
    ).annotate(
        unit_count=Count('programme_units')
    ).order_by('level', 'name')
    
    # Get all departments for the create/edit form
    departments = Department.objects.all().order_by('name')
    
    # Get all units (for assigning to programmes)
    all_units = Unit.objects.select_related('department').order_by('code')
    
    context = {
        'programmes': programmes,
        'departments': departments,
        'all_units': all_units,
    }
    return render(request, 'admin/unit_management.html', context)


@login_required
@require_http_methods(["GET"])
def get_programme_units(request, programme_id):
    """Get all units for a specific programme"""
    try:
        programme_units = ProgrammeUnit.objects.filter(
            programme_id=programme_id
        ).select_related('unit', 'unit__department').order_by('year_level', 'semester', 'unit__code')
        
        data = [{
            'id': pu.id,
            'unit_id': pu.unit.id,
            'unit_code': pu.unit.code,
            'unit_name': pu.unit.name,
            'department': pu.unit.department.name,
            'credit_hours': pu.unit.credit_hours,
            'year_level': pu.year_level,
            'year_display': pu.get_year_level_display(),
            'semester': pu.semester,
            'semester_display': f'Semester {pu.semester}',
            'is_mandatory': pu.is_mandatory,
            'is_core': pu.unit.is_core,
        } for pu in programme_units]
        
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def get_unit(request, unit_id):
    """Get single unit details"""
    try:
        unit = Unit.objects.select_related('department').get(id=unit_id)
        
        # Get prerequisites
        prerequisites = list(unit.prerequisites.values_list('id', flat=True))
        
        data = {
            'id': unit.id,
            'code': unit.code,
            'name': unit.name,
            'description': unit.description,
            'credit_hours': unit.credit_hours,
            'department_id': unit.department.id,
            'is_core': unit.is_core,
            'prerequisites': prerequisites,
        }
        return JsonResponse({'success': True, 'data': data})
    except Unit.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Unit not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def create_unit(request):
    """Create new unit"""
    try:
        data = json.loads(request.body)
        
        code = data.get('code')
        name = data.get('name')
        description = data.get('description', '')
        credit_hours = int(data.get('credit_hours', 3))
        department_id = data.get('department_id')
        is_core = data.get('is_core', True)
        prerequisites = data.get('prerequisites', [])
        
        # Validation
        if not code or not name or not department_id:
            return JsonResponse({'success': False, 'message': 'Code, name, and department are required'}, status=400)
        
        # Check if code already exists
        if Unit.objects.filter(code=code).exists():
            return JsonResponse({'success': False, 'message': 'Unit code already exists'}, status=400)
        
        # Create unit
        unit = Unit.objects.create(
            code=code.upper(),
            name=name,
            description=description,
            credit_hours=credit_hours,
            department_id=department_id,
            is_core=is_core
        )
        
        # Add prerequisites
        if prerequisites:
            unit.prerequisites.set(prerequisites)
        
        return JsonResponse({
            'success': True,
            'message': 'Unit created successfully',
            'data': {
                'id': unit.id,
                'code': unit.code,
                'name': unit.name,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_unit(request, unit_id):
    """Update existing unit"""
    try:
        unit = get_object_or_404(Unit, id=unit_id)
        data = json.loads(request.body)
        
        code = data.get('code')
        name = data.get('name')
        description = data.get('description', '')
        credit_hours = int(data.get('credit_hours', 3))
        department_id = data.get('department_id')
        is_core = data.get('is_core', True)
        prerequisites = data.get('prerequisites', [])
        
        # Check if code already exists (excluding current)
        if Unit.objects.filter(code=code).exclude(id=unit_id).exists():
            return JsonResponse({'success': False, 'message': 'Unit code already exists'}, status=400)
        
        # Update unit
        unit.code = code.upper()
        unit.name = name
        unit.description = description
        unit.credit_hours = credit_hours
        unit.department_id = department_id
        unit.is_core = is_core
        unit.save()
        
        # Update prerequisites
        unit.prerequisites.set(prerequisites)
        
        return JsonResponse({
            'success': True,
            'message': 'Unit updated successfully',
            'data': {
                'id': unit.id,
                'code': unit.code,
                'name': unit.name,
            }
        })
    except Unit.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Unit not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def delete_unit(request, unit_id):
    """Delete unit"""
    try:
        unit = get_object_or_404(Unit, id=unit_id)
        
        # Check if unit is assigned to any programmes
        if unit.offered_in_programmes.exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete unit that is assigned to programmes. Remove from programmes first.'
            }, status=400)
        
        # Check if there are enrollments
        if unit.enrollments.exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot delete unit with existing enrollments.'
            }, status=400)
        
        unit.delete()
        return JsonResponse({'success': True, 'message': 'Unit deleted successfully'})
    except Unit.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Unit not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


# ========================
# PROGRAMME UNIT ASSIGNMENT
# ========================

@login_required
@require_http_methods(["GET"])
def get_programme_unit(request, programme_unit_id):
    """Get single programme unit assignment details"""
    try:
        pu = ProgrammeUnit.objects.select_related('unit', 'programme').get(id=programme_unit_id)
        data = {
            'id': pu.id,
            'programme_id': pu.programme.id,
            'unit_id': pu.unit.id,
            'year_level': pu.year_level,
            'semester': pu.semester,
            'is_mandatory': pu.is_mandatory,
        }
        return JsonResponse({'success': True, 'data': data})
    except ProgrammeUnit.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Programme unit not found'}, status=404)


@login_required
@require_http_methods(["POST"])
def assign_unit_to_programme(request, programme_id):
    """Assign a unit to a programme"""
    try:
        programme = get_object_or_404(Programme, id=programme_id)
        data = json.loads(request.body)
        
        unit_id = data.get('unit_id')
        year_level = int(data.get('year_level'))
        semester = int(data.get('semester'))
        is_mandatory = data.get('is_mandatory', True)
        
        # Validation
        if not unit_id or not year_level or not semester:
            return JsonResponse({'success': False, 'message': 'All fields are required'}, status=400)
        
        # Check if already assigned
        if ProgrammeUnit.objects.filter(
            programme=programme,
            unit_id=unit_id,
            year_level=year_level,
            semester=semester
        ).exists():
            return JsonResponse({'success': False, 'message': 'Unit already assigned to this programme/year/semester'}, status=400)
        
        # Create assignment
        programme_unit = ProgrammeUnit.objects.create(
            programme=programme,
            unit_id=unit_id,
            year_level=year_level,
            semester=semester,
            is_mandatory=is_mandatory
        )
        
        return JsonResponse({
            'success': True,
            'message': 'Unit assigned successfully',
            'data': {
                'id': programme_unit.id,
                'unit_code': programme_unit.unit.code,
                'unit_name': programme_unit.unit.name,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_programme_unit(request, programme_unit_id):
    """Update programme unit assignment"""
    try:
        programme_unit = get_object_or_404(ProgrammeUnit, id=programme_unit_id)
        data = json.loads(request.body)
        
        year_level = int(data.get('year_level'))
        semester = int(data.get('semester'))
        is_mandatory = data.get('is_mandatory', True)
        
        # Check if another assignment exists with same details
        if ProgrammeUnit.objects.filter(
            programme=programme_unit.programme,
            unit=programme_unit.unit,
            year_level=year_level,
            semester=semester
        ).exclude(id=programme_unit_id).exists():
            return JsonResponse({'success': False, 'message': 'Unit already assigned to this year/semester'}, status=400)
        
        # Update assignment
        programme_unit.year_level = year_level
        programme_unit.semester = semester
        programme_unit.is_mandatory = is_mandatory
        programme_unit.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Assignment updated successfully',
        })
    except ProgrammeUnit.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Programme unit not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["DELETE"])
def remove_unit_from_programme(request, programme_unit_id):
    """Remove unit from programme"""
    try:
        programme_unit = get_object_or_404(ProgrammeUnit, id=programme_unit_id)
        
        # Check if there are enrollments
        if programme_unit.unit.enrollments.filter(
            student__programme=programme_unit.programme
        ).exists():
            return JsonResponse({
                'success': False,
                'message': 'Cannot remove unit with existing student enrollments.'
            }, status=400)
        
        programme_unit.delete()
        return JsonResponse({'success': True, 'message': 'Unit removed from programme successfully'})
    except ProgrammeUnit.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Programme unit not found'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def search_units(request):
    """Search units for assignment"""
    try:
        query = request.GET.get('q', '')
        programme_id = request.GET.get('programme_id')
        
        units = Unit.objects.select_related('department').filter(
            Q(code__icontains=query) | Q(name__icontains=query)
        ).order_by('code')[:20]
        
        # Exclude already assigned units if programme_id provided
        if programme_id:
            assigned_unit_ids = ProgrammeUnit.objects.filter(
                programme_id=programme_id
            ).values_list('unit_id', flat=True)
            units = units.exclude(id__in=assigned_unit_ids)
        
        data = [{
            'id': unit.id,
            'code': unit.code,
            'name': unit.name,
            'department': unit.department.name,
            'credit_hours': unit.credit_hours,
            'is_core': unit.is_core,
        } for unit in units]
        
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)


@login_required
def unit_detail(request, unit_id):
    """Detailed view for a single unit"""
    unit = get_object_or_404(Unit.objects.select_related('department'), id=unit_id)
    
    # Get programmes offering this unit
    programme_assignments = ProgrammeUnit.objects.filter(
        unit=unit
    ).select_related('programme', 'programme__department').order_by('programme__name', 'year_level', 'semester')
    
    # Get prerequisites and required_for
    prerequisites = unit.prerequisites.all()
    required_for = unit.required_for.all()
    
    # Get enrollment statistics
    total_enrollments = unit.enrollments.count()
    active_enrollments = unit.enrollments.filter(status='ENROLLED').count()
    
    context = {
        'unit': unit,
        'programme_assignments': programme_assignments,
        'prerequisites': prerequisites,
        'required_for': required_for,
        'total_enrollments': total_enrollments,
        'active_enrollments': active_enrollments,
    }
    return render(request, 'admin/unit_detail.html', context)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import JsonResponse
from .models import GradingScheme, Programme, Department
from .forms import GradingSchemeForm

# ========================
# GRADING SCHEME LIST VIEW
# ========================

@login_required
def grading_scheme_list(request):
    """Display list of all grading schemes with filters"""
    
    # Check user permissions
    if request.user.user_type not in ['ICT_ADMIN', 'DEAN', 'COD']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_dashboard')
    
    # Get filter parameters
    search_query = request.GET.get('search', '')
    programme_filter = request.GET.get('programme', '')
    department_filter = request.GET.get('department', '')
    
    # Base queryset
    schemes = GradingScheme.objects.select_related(
        'programme',
        'programme__department'
    ).all()
    
    # Apply filters
    if search_query:
        schemes = schemes.filter(
            Q(grade__icontains=search_query) |
            Q(programme__name__icontains=search_query) |
            Q(programme__code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if programme_filter:
        schemes = schemes.filter(programme_id=programme_filter)
    
    if department_filter:
        schemes = schemes.filter(programme__department_id=department_filter)
    
    # Order by programme and marks
    schemes = schemes.order_by('programme__code', '-min_marks')
    
    # Pagination
    paginator = Paginator(schemes, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    departments = Department.objects.all().order_by('name')
    
    # Statistics
    total_schemes = schemes.count()
    programmes_with_schemes = schemes.values('programme').distinct().count()
    
    context = {
        'page_obj': page_obj,
        'schemes': page_obj,
        'programmes': programmes,
        'departments': departments,
        'search_query': search_query,
        'programme_filter': programme_filter,
        'department_filter': department_filter,
        'total_schemes': total_schemes,
        'programmes_with_schemes': programmes_with_schemes,
    }
    
    return render(request, 'grading/grading_scheme_list.html', context)


# ========================
# GRADING SCHEME DETAIL VIEW
# ========================

@login_required
def grading_scheme_detail(request, scheme_id):
    """Display detailed information about a grading scheme"""
    
    # Check user permissions
    if request.user.user_type not in ['ICT_ADMIN', 'DEAN', 'COD', 'LECTURER']:
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_dashboard')
    
    scheme = get_object_or_404(
        GradingScheme.objects.select_related('programme', 'programme__department'),
        id=scheme_id
    )
    
    # Get all grading schemes for this programme
    programme_schemes = GradingScheme.objects.filter(
        programme=scheme.programme
    ).order_by('-min_marks')
    
    # Statistics
    total_grades = programme_schemes.count()
    passing_grades = programme_schemes.filter(
        description__icontains='pass'
    ).count()
    
    context = {
        'scheme': scheme,
        'programme_schemes': programme_schemes,
        'total_grades': total_grades,
        'passing_grades': passing_grades,
    }
    
    return render(request, 'grading/grading_scheme_detail.html', context)


# ========================
# GRADING SCHEME CREATE VIEW
# ========================

@login_required
def grading_scheme_create(request):
    """Create a new grading scheme"""
    
    # Check user permissions
    if request.user.user_type not in ['ICT_ADMIN', 'DEAN']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('grading_scheme_list')
    
    if request.method == 'POST':
        form = GradingSchemeForm(request.POST)
        if form.is_valid():
            # Validate mark ranges
            min_marks = form.cleaned_data['min_marks']
            max_marks = form.cleaned_data['max_marks']
            programme = form.cleaned_data['programme']
            grade = form.cleaned_data['grade']
            
            # Check if grade already exists for this programme
            if GradingScheme.objects.filter(programme=programme, grade=grade).exists():
                messages.error(request, f"Grade '{grade}' already exists for {programme.code}.")
                return render(request, 'grading/grading_scheme_form.html', {'form': form})
            
            # Validate min < max
            if min_marks >= max_marks:
                messages.error(request, "Minimum marks must be less than maximum marks.")
                return render(request, 'grading/grading_scheme_form.html', {'form': form})
            
            # Check for overlapping ranges
            overlapping = GradingScheme.objects.filter(
                programme=programme
            ).filter(
                Q(min_marks__lte=max_marks, max_marks__gte=min_marks)
            )
            
            if overlapping.exists():
                messages.error(request, "Mark range overlaps with existing grading scheme.")
                return render(request, 'grading/grading_scheme_form.html', {'form': form})
            
            scheme = form.save()
            messages.success(request, f"Grading scheme '{scheme.grade}' created successfully!")
            return redirect('grading_scheme_detail', scheme_id=scheme.id)
    else:
        form = GradingSchemeForm()
    
    context = {
        'form': form,
        'action': 'Create',
    }
    
    return render(request, 'grading/grading_scheme_form.html', context)


# ========================
# GRADING SCHEME UPDATE VIEW
# ========================

@login_required
def grading_scheme_update(request, scheme_id):
    """Update an existing grading scheme"""
    
    # Check user permissions
    if request.user.user_type not in ['ICT_ADMIN', 'DEAN']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('grading_scheme_list')
    
    scheme = get_object_or_404(GradingScheme, id=scheme_id)
    
    if request.method == 'POST':
        form = GradingSchemeForm(request.POST, instance=scheme)
        if form.is_valid():
            # Validate mark ranges
            min_marks = form.cleaned_data['min_marks']
            max_marks = form.cleaned_data['max_marks']
            programme = form.cleaned_data['programme']
            grade = form.cleaned_data['grade']
            
            # Check if grade already exists for this programme (excluding current)
            if GradingScheme.objects.filter(
                programme=programme, 
                grade=grade
            ).exclude(id=scheme_id).exists():
                messages.error(request, f"Grade '{grade}' already exists for {programme.code}.")
                return render(request, 'grading/grading_scheme_form.html', {
                    'form': form,
                    'scheme': scheme,
                    'action': 'Update'
                })
            
            # Validate min < max
            if min_marks >= max_marks:
                messages.error(request, "Minimum marks must be less than maximum marks.")
                return render(request, 'grading/grading_scheme_form.html', {
                    'form': form,
                    'scheme': scheme,
                    'action': 'Update'
                })
            
            # Check for overlapping ranges (excluding current)
            overlapping = GradingScheme.objects.filter(
                programme=programme
            ).exclude(id=scheme_id).filter(
                Q(min_marks__lte=max_marks, max_marks__gte=min_marks)
            )
            
            if overlapping.exists():
                messages.error(request, "Mark range overlaps with existing grading scheme.")
                return render(request, 'grading/grading_scheme_form.html', {
                    'form': form,
                    'scheme': scheme,
                    'action': 'Update'
                })
            
            scheme = form.save()
            messages.success(request, f"Grading scheme '{scheme.grade}' updated successfully!")
            return redirect('grading_scheme_detail', scheme_id=scheme.id)
    else:
        form = GradingSchemeForm(instance=scheme)
    
    context = {
        'form': form,
        'scheme': scheme,
        'action': 'Update',
    }
    
    return render(request, 'grading/grading_scheme_form.html', context)


# ========================
# GRADING SCHEME DELETE VIEW
# ========================

@login_required
def grading_scheme_delete(request, scheme_id):
    """Delete a grading scheme"""
    
    # Check user permissions
    if request.user.user_type not in ['ICT_ADMIN', 'DEAN']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('grading_scheme_list')
    
    scheme = get_object_or_404(GradingScheme, id=scheme_id)
    
    if request.method == 'POST':
        programme_name = scheme.programme.name
        grade = scheme.grade
        scheme.delete()
        messages.success(request, f"Grading scheme '{grade}' for {programme_name} deleted successfully!")
        return redirect('grading_scheme_list')
    
    context = {
        'scheme': scheme,
    }
    
    return render(request, 'grading/grading_scheme_delete.html', context)


# ========================
# BULK GRADING SCHEME CREATION
# ========================

@login_required
def grading_scheme_bulk_create(request):
    """Create multiple grading schemes for a programme at once"""
    
    # Check user permissions
    if request.user.user_type not in ['ICT_ADMIN', 'DEAN']:
        messages.error(request, "You don't have permission to perform this action.")
        return redirect('grading_scheme_list')
    
    if request.method == 'POST':
        programme_id = request.POST.get('programme')
        programme = get_object_or_404(Programme, id=programme_id)
        
        # Check if grading scheme already exists
        existing_count = GradingScheme.objects.filter(programme=programme).count()
        if existing_count > 0:
            messages.warning(
                request, 
                f"{programme.code} already has {existing_count} grading schemes. "
                "Please delete them first or add schemes individually."
            )
            return redirect('grading_scheme_list')
        
        # Default grading scheme (can be customized)
        default_schemes = [
            {'grade': 'A', 'min_marks': 70, 'max_marks': 100, 'grade_point': 4.0, 'description': 'Distinction'},
            {'grade': 'B', 'min_marks': 60, 'max_marks': 69.99, 'grade_point': 3.0, 'description': 'Credit'},
            {'grade': 'C', 'min_marks': 50, 'max_marks': 59.99, 'grade_point': 2.0, 'description': 'Pass'},
            {'grade': 'D', 'min_marks': 40, 'max_marks': 49.99, 'grade_point': 1.0, 'description': 'Pass'},
            {'grade': 'F', 'min_marks': 0, 'max_marks': 39.99, 'grade_point': 0.0, 'description': 'Fail'},
        ]
        
        created_count = 0
        for scheme_data in default_schemes:
            GradingScheme.objects.create(
                programme=programme,
                **scheme_data
            )
            created_count += 1
        
        messages.success(
            request, 
            f"Successfully created {created_count} grading schemes for {programme.code}!"
        )
        return redirect('grading_scheme_list')
    
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    
    context = {
        'programmes': programmes,
    }
    
    return render(request, 'grading/grading_scheme_bulk_create.html', context)


# ========================
# AJAX: GET PROGRAMME GRADING SCHEMES
# ========================

@login_required
def get_programme_schemes(request, programme_id):
    """AJAX endpoint to get all grading schemes for a programme"""
    
    schemes = GradingScheme.objects.filter(
        programme_id=programme_id
    ).order_by('-min_marks').values(
        'id', 'grade', 'min_marks', 'max_marks', 'grade_point', 'description'
    )
    
    return JsonResponse({
        'schemes': list(schemes)
    })

# views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils import timezone
from .models import (
    FeeStructure, FeePayment, FeeStatement, Student, 
    Programme, AcademicYear, Semester
)
from decimal import Decimal


@login_required
def fee_structure_list(request):
    """Display fee structures organized by academic year and programme"""
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    
    # Get selected filters
    selected_year_id = request.GET.get('academic_year')
    selected_programme_id = request.GET.get('programme')
    
    # Default to current academic year if available
    if not selected_year_id:
        current_year = academic_years.filter(is_current=True).first()
        if current_year:
            selected_year_id = current_year.id
    
    fee_structures = FeeStructure.objects.all()
    
    if selected_year_id:
        fee_structures = fee_structures.filter(academic_year_id=selected_year_id)
    
    if selected_programme_id:
        fee_structures = fee_structures.filter(programme_id=selected_programme_id)
    
    # Organize fee structures by programme and year level
    programme_fees = {}
    for fee in fee_structures.select_related('programme', 'academic_year'):
        prog_key = fee.programme.id
        if prog_key not in programme_fees:
            programme_fees[prog_key] = {
                'programme': fee.programme,
                'academic_year': fee.academic_year,
                'year_levels': {}
            }
        programme_fees[prog_key]['year_levels'][fee.year_level] = fee
    
    context = {
        'academic_years': academic_years,
        'programmes': programmes,
        'programme_fees': programme_fees,
        'selected_year_id': int(selected_year_id) if selected_year_id else None,
        'selected_programme_id': int(selected_programme_id) if selected_programme_id else None,
    }
    
    return render(request, 'fees/fee_structure_list.html', context)


@login_required
def fee_structure_create(request):
    """Create new fee structure"""
    if request.method == 'POST':
        programme_id = request.POST.get('programme')
        academic_year_id = request.POST.get('academic_year')
        year_level = request.POST.get('year_level')
        tuition_fee = Decimal(request.POST.get('tuition_fee', 0))
        examination_fee = Decimal(request.POST.get('examination_fee', 0))
        registration_fee = Decimal(request.POST.get('registration_fee', 0))
        other_fees = Decimal(request.POST.get('other_fees', 0))
        
        total_fee = tuition_fee + examination_fee + registration_fee + other_fees
        
        try:
            FeeStructure.objects.create(
                programme_id=programme_id,
                academic_year_id=academic_year_id,
                year_level=year_level,
                tuition_fee=tuition_fee,
                examination_fee=examination_fee,
                registration_fee=registration_fee,
                other_fees=other_fees,
                total_fee=total_fee
            )
            messages.success(request, 'Fee structure created successfully!')
            return redirect('fee_structure_list')
        except Exception as e:
            messages.error(request, f'Error creating fee structure: {str(e)}')
    
    programmes = Programme.objects.filter(is_active=True)
    academic_years = AcademicYear.objects.all()
    
    context = {
        'programmes': programmes,
        'academic_years': academic_years,
    }
    
    return render(request, 'fees/fee_structure_form.html', context)


@login_required
def fee_structure_update(request, pk):
    """Update existing fee structure"""
    fee_structure = get_object_or_404(FeeStructure, pk=pk)
    
    if request.method == 'POST':
        fee_structure.tuition_fee = Decimal(request.POST.get('tuition_fee', 0))
        fee_structure.examination_fee = Decimal(request.POST.get('examination_fee', 0))
        fee_structure.registration_fee = Decimal(request.POST.get('registration_fee', 0))
        fee_structure.other_fees = Decimal(request.POST.get('other_fees', 0))
        fee_structure.total_fee = (
            fee_structure.tuition_fee + 
            fee_structure.examination_fee + 
            fee_structure.registration_fee + 
            fee_structure.other_fees
        )
        
        try:
            fee_structure.save()
            messages.success(request, 'Fee structure updated successfully!')
            return redirect('fee_structure_list')
        except Exception as e:
            messages.error(request, f'Error updating fee structure: {str(e)}')
    
    context = {
        'fee_structure': fee_structure,
    }
    
    return render(request, 'fees/fee_structure_form.html', context)


@login_required
def fee_structure_delete(request, pk):
    """Delete fee structure"""
    fee_structure = get_object_or_404(FeeStructure, pk=pk)
    
    if request.method == 'POST':
        try:
            fee_structure.delete()
            messages.success(request, 'Fee structure deleted successfully!')
        except Exception as e:
            messages.error(request, f'Error deleting fee structure: {str(e)}')
    
    return redirect('fee_structure_list')


@login_required
def fee_payment_list(request):
    """Display all fee payments with filters"""
    payments = FeePayment.objects.all().select_related(
        'student', 'semester', 'fee_structure'
    ).order_by('-payment_date')
    
    # Filters
    search = request.GET.get('search', '')
    status_filter = request.GET.get('status', '')
    semester_filter = request.GET.get('semester', '')
    
    if search:
        payments = payments.filter(
            Q(student__registration_number__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search) |
            Q(receipt_number__icontains=search)
        )
    
    if status_filter:
        payments = payments.filter(status=status_filter)
    
    if semester_filter:
        payments = payments.filter(semester_id=semester_filter)
    
    # Calculate totals
    total_collected = payments.aggregate(total=Sum('amount_paid'))['total'] or 0
    
    semesters = Semester.objects.all().order_by('-start_date')
    
    context = {
        'payments': payments,
        'semesters': semesters,
        'search': search,
        'status_filter': status_filter,
        'semester_filter': semester_filter,
        'total_collected': total_collected,
        'payment_statuses': FeePayment.PAYMENT_STATUS,
    }
    
    return render(request, 'fees/fee_payment_list.html', context)


@login_required
def fee_payment_create(request):
    """Record a new fee payment"""
    if request.method == 'POST':
        student_id = request.POST.get('student')
        semester_id = request.POST.get('semester')
        amount_paid = Decimal(request.POST.get('amount_paid', 0))
        payment_date = request.POST.get('payment_date')
        payment_method = request.POST.get('payment_method')
        receipt_number = request.POST.get('receipt_number')
        
        try:
            student = Student.objects.get(user_id=student_id)
            semester = Semester.objects.get(id=semester_id)
            
            # Get fee structure
            fee_structure = FeeStructure.objects.get(
                programme=student.programme,
                academic_year=semester.academic_year,
                year_level=student.current_year
            )
            
            # Get or create fee statement
            statement, created = FeeStatement.objects.get_or_create(
                student=student,
                semester=semester,
                defaults={
                    'total_billed': fee_structure.total_fee,
                    'total_paid': 0,
                    'balance': fee_structure.total_fee
                }
            )
            
            # Calculate new totals
            new_total_paid = statement.total_paid + amount_paid
            new_balance = statement.total_billed - new_total_paid
            
            # Determine payment status
            if new_balance <= 0:
                status = 'COMPLETE'
            elif new_balance < statement.total_billed:
                status = 'PARTIAL'
            else:
                status = 'PENDING'
            
            # Create payment record
            FeePayment.objects.create(
                student=student,
                semester=semester,
                fee_structure=fee_structure,
                amount_paid=amount_paid,
                payment_date=payment_date,
                receipt_number=receipt_number,
                payment_method=payment_method,
                status=status,
                balance=new_balance
            )
            
            # Update fee statement
            statement.total_paid = new_total_paid
            statement.balance = new_balance
            statement.can_register = new_balance <= (statement.total_billed * Decimal('0.5'))
            statement.save()
            
            messages.success(request, 'Fee payment recorded successfully!')
            return redirect('fee_payment_list')
            
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
    
    students = Student.objects.filter(is_active=True).select_related('user', 'programme')
    semesters = Semester.objects.all().order_by('-start_date')
    
    context = {
        'students': students,
        'semesters': semesters,
    }
    
    return render(request, 'fees/fee_payment_form.html', context)


@login_required
def fee_payment_detail(request, pk):
    """View detailed information about a payment"""
    payment = get_object_or_404(
        FeePayment.objects.select_related(
            'student', 'student__user', 'semester', 'fee_structure'
        ),
        pk=pk
    )
    
    # Get all payments for this student and semester
    related_payments = FeePayment.objects.filter(
        student=payment.student,
        semester=payment.semester
    ).order_by('-payment_date')
    
    context = {
        'payment': payment,
        'related_payments': related_payments,
    }
    
    return render(request, 'fees/fee_payment_detail.html', context)


@login_required
def fee_statement_list(request):
    """Display fee statements for all students"""
    statements = FeeStatement.objects.all().select_related(
        'student', 'student__user', 'semester'
    ).order_by('-semester__start_date', 'student__registration_number')
    
    # Filters
    search = request.GET.get('search', '')
    semester_filter = request.GET.get('semester', '')
    balance_filter = request.GET.get('balance', '')
    
    if search:
        statements = statements.filter(
            Q(student__registration_number__icontains=search) |
            Q(student__first_name__icontains=search) |
            Q(student__last_name__icontains=search)
        )
    
    if semester_filter:
        statements = statements.filter(semester_id=semester_filter)
    
    if balance_filter == 'with_balance':
        statements = statements.filter(balance__gt=0)
    elif balance_filter == 'cleared':
        statements = statements.filter(balance__lte=0)
    
    # Calculate totals
    totals = statements.aggregate(
        total_billed=Sum('total_billed'),
        total_paid=Sum('total_paid'),
        total_balance=Sum('balance')
    )
    
    semesters = Semester.objects.all().order_by('-start_date')
    
    context = {
        'statements': statements,
        'semesters': semesters,
        'search': search,
        'semester_filter': semester_filter,
        'balance_filter': balance_filter,
        'totals': totals,
    }
    
    return render(request, 'fees/fee_statement_list.html', context)


@login_required
def fee_statement_detail(request, student_id, semester_id):
    """View detailed fee statement for a specific student and semester"""
    student = get_object_or_404(Student, user_id=student_id)
    semester = get_object_or_404(Semester, id=semester_id)
    
    statement = get_object_or_404(
        FeeStatement,
        student=student,
        semester=semester
    )
    
    # Get all payments for this statement
    payments = FeePayment.objects.filter(
        student=student,
        semester=semester
    ).order_by('payment_date')
    
    # Get fee structure
    fee_structure = FeeStructure.objects.filter(
        programme=student.programme,
        academic_year=semester.academic_year,
        year_level=student.current_year
    ).first()
    
    context = {
        'statement': statement,
        'student': student,
        'semester': semester,
        'payments': payments,
        'fee_structure': fee_structure,
    }
    
    return render(request, 'fees/fee_statement_detail.html', context)

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Avg, Sum, Count
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from .models import (
    Student, Programme, AcademicYear, Semester, Unit, 
    UnitEnrollment, StudentMarks, FinalGrade, Department
)


# ========================
# HELPER FUNCTIONS
# ========================

def calculate_gpa(enrollments):
    """Calculate GPA from enrollments with final grades"""
    total_points = 0
    total_credits = 0
    
    for enrollment in enrollments:
        if hasattr(enrollment, 'final_grade') and enrollment.final_grade:
            grade_point = enrollment.final_grade.grade_point
            credits = enrollment.unit.credit_hours
            total_points += (grade_point * credits)
            total_credits += credits
    
    if total_credits > 0:
        return round(total_points / total_credits, 2)
    return 0.0


def get_classification(gpa):
    """Determine degree classification based on GPA"""
    if gpa >= 3.70:
        return "First Class Honours"
    elif gpa >= 3.00:
        return "Second Class Honours (Upper Division)"
    elif gpa >= 2.50:
        return "Second Class Honours (Lower Division)"
    elif gpa >= 2.00:
        return "Pass"
    else:
        return "Fail"


def get_semester_results(student, semester):
    """Get all results for a student in a specific semester"""
    enrollments = UnitEnrollment.objects.filter(
        student=student,
        semester=semester
    ).select_related('unit', 'final_grade').prefetch_related('marks__assessment_component')
    
    results = []
    for enrollment in enrollments:
        unit_data = {
            'enrollment': enrollment,
            'unit': enrollment.unit,
            'marks': [],
            'total_marks': 0,
            'grade': None,
            'grade_point': None
        }
        
        # Get all marks for this enrollment
        marks = StudentMarks.objects.filter(
            enrollment=enrollment
        ).select_related('assessment_component')
        
        for mark in marks:
            unit_data['marks'].append({
                'component': mark.assessment_component.name,
                'type': mark.assessment_component.get_component_type_display(),
                'marks': mark.marks_obtained,
                'max_marks': mark.assessment_component.max_marks,
                'weight': mark.assessment_component.weight_percentage
            })
        
        # Get final grade
        if hasattr(enrollment, 'final_grade') and enrollment.final_grade:
            unit_data['total_marks'] = enrollment.final_grade.total_marks
            unit_data['grade'] = enrollment.final_grade.grade
            unit_data['grade_point'] = enrollment.final_grade.grade_point
        
        results.append(unit_data)
    
    return results


# ========================
# STUDENT PROGRESSION LIST
# ========================

@login_required
def student_progression_list(request):
    """
    Display list of students with progression tracking
    Features: Search, Filter, Export
    """
    # Get filter parameters
    search_query = request.GET.get('search', '')
    programme_id = request.GET.get('programme', '')
    year_level = request.GET.get('year', '')
    academic_year_id = request.GET.get('academic_year', '')
    department_id = request.GET.get('department', '')
    status = request.GET.get('status', '')
    
    # Base queryset
    students = Student.objects.select_related(
        'programme', 'programme__department', 'intake', 'user'
    ).all()
    
    # Apply filters
    if search_query:
        students = students.filter(
            Q(registration_number__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(surname__icontains=search_query)
        )
    
    if programme_id:
        students = students.filter(programme_id=programme_id)
    
    if year_level:
        students = students.filter(current_year=year_level)
    
    if department_id:
        students = students.filter(programme__department_id=department_id)
    
    if status == 'active':
        students = students.filter(is_active=True)
    elif status == 'inactive':
        students = students.filter(is_active=False)
    
    # Calculate progression data for each student
    student_data = []
    for student in students:
        # Get all enrollments with final grades
        enrollments = UnitEnrollment.objects.filter(
            student=student,
            status='COMPLETED'
        ).select_related('final_grade', 'unit', 'semester')
        
        # Calculate GPA
        gpa = calculate_gpa(enrollments)
        
        # Get statistics
        total_units = enrollments.count()
        completed_units = enrollments.filter(final_grade__isnull=False).count()
        failed_units = enrollments.filter(final_grade__grade='F').count()
        
        # Calculate total credits
        total_credits = enrollments.aggregate(
            total=Sum('unit__credit_hours')
        )['total'] or 0
        
        student_data.append({
            'student': student,
            'gpa': gpa,
            'classification': get_classification(gpa),
            'total_units': total_units,
            'completed_units': completed_units,
            'failed_units': failed_units,
            'total_credits': total_credits,
        })
    
    # Pagination
    paginator = Paginator(student_data, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    departments = Department.objects.all().order_by('name')
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    
    context = {
        'page_obj': page_obj,
        'programmes': programmes,
        'departments': departments,
        'academic_years': academic_years,
        'search_query': search_query,
        'selected_programme': programme_id,
        'selected_year': year_level,
        'selected_academic_year': academic_year_id,
        'selected_department': department_id,
        'selected_status': status,
        'total_students': len(student_data),
    }
    
    return render(request, 'students/progression_list.html', context)


# ========================
# STUDENT PROGRESSION DETAIL
# ========================

@login_required
def student_progression_detail(request, registration_number):
    """
    Detailed progression view for individual student
    Shows marks by academic year and semester with classification
    """
    student = get_object_or_404(
        Student.objects.select_related('programme', 'programme__department', 'intake', 'user'),
        registration_number=registration_number
    )
    
    # Get all academic years the student has been enrolled
    enrollments = UnitEnrollment.objects.filter(
        student=student
    ).select_related('semester__academic_year', 'unit', 'final_grade')
    
    # Organize data by academic year and semester
    academic_data = {}
    
    for enrollment in enrollments:
        academic_year = enrollment.semester.academic_year
        semester = enrollment.semester
        
        if academic_year not in academic_data:
            academic_data[academic_year] = {
                'year': academic_year,
                'semesters': {},
                'year_gpa': 0,
                'total_credits': 0
            }
        
        if semester not in academic_data[academic_year]['semesters']:
            academic_data[academic_year]['semesters'][semester] = {
                'semester': semester,
                'results': [],
                'semester_gpa': 0,
                'total_credits': 0
            }
        
        # Get marks for this enrollment
        marks_data = []
        marks = StudentMarks.objects.filter(
            enrollment=enrollment
        ).select_related('assessment_component')
        
        for mark in marks:
            marks_data.append({
                'component': mark.assessment_component.name,
                'type': mark.assessment_component.get_component_type_display(),
                'marks': mark.marks_obtained,
                'max_marks': mark.assessment_component.max_marks,
                'weight': mark.assessment_component.weight_percentage
            })
        
        result_item = {
            'unit': enrollment.unit,
            'marks': marks_data,
            'total_marks': 0,
            'grade': '-',
            'grade_point': 0,
            'credits': enrollment.unit.credit_hours,
            'status': enrollment.get_status_display()
        }
        
        if hasattr(enrollment, 'final_grade') and enrollment.final_grade:
            result_item['total_marks'] = enrollment.final_grade.total_marks
            result_item['grade'] = enrollment.final_grade.grade
            result_item['grade_point'] = enrollment.final_grade.grade_point
        
        academic_data[academic_year]['semesters'][semester]['results'].append(result_item)
    
    # Calculate GPAs for each semester and year
    for year_key, year_data in academic_data.items():
        year_enrollments = []
        for semester_key, semester_data in year_data['semesters'].items():
            # Calculate semester GPA
            semester_enrollments = UnitEnrollment.objects.filter(
                student=student,
                semester=semester_key,
                status='COMPLETED'
            ).select_related('final_grade', 'unit')
            
            semester_gpa = calculate_gpa(semester_enrollments)
            semester_data['semester_gpa'] = semester_gpa
            
            # Calculate semester credits
            semester_credits = semester_enrollments.aggregate(
                total=Sum('unit__credit_hours')
            )['total'] or 0
            semester_data['total_credits'] = semester_credits
            
            year_enrollments.extend(list(semester_enrollments))
        
        # Calculate year GPA
        if year_enrollments:
            year_gpa = calculate_gpa(year_enrollments)
            year_data['year_gpa'] = year_gpa
            
            # Calculate year credits
            year_credits = sum([e.unit.credit_hours for e in year_enrollments])
            year_data['total_credits'] = year_credits
    
    # Calculate overall GPA
    all_enrollments = UnitEnrollment.objects.filter(
        student=student,
        status='COMPLETED'
    ).select_related('final_grade', 'unit')
    
    overall_gpa = calculate_gpa(all_enrollments)
    classification = get_classification(overall_gpa)
    
    # Calculate overall statistics
    total_credits = all_enrollments.aggregate(
        total=Sum('unit__credit_hours')
    )['total'] or 0
    
    completed_units = all_enrollments.filter(final_grade__isnull=False).count()
    failed_units = all_enrollments.filter(final_grade__grade='F').count()
    
    context = {
        'student': student,
        'academic_data': academic_data,
        'overall_gpa': overall_gpa,
        'classification': classification,
        'total_credits': total_credits,
        'completed_units': completed_units,
        'failed_units': failed_units,
        'total_enrollments': all_enrollments.count(),
    }
    
    return render(request, 'students/progression_detail.html', context)


# ========================
# EXPORT FUNCTIONALITY
# ========================

@login_required
def export_student_progression(request):
    """
    Export student progression data to CSV or Excel
    """
    export_format = request.GET.get('format', 'csv')
    programme_id = request.GET.get('programme', '')
    year_level = request.GET.get('year', '')
    academic_year_id = request.GET.get('academic_year', '')
    unit_id = request.GET.get('unit', '')
    
    # Base queryset
    students = Student.objects.select_related(
        'programme', 'programme__department', 'intake'
    ).all()
    
    # Apply filters
    if programme_id:
        students = students.filter(programme_id=programme_id)
    if year_level:
        students = students.filter(current_year=year_level)
    
    # Get enrollments based on filters
    enrollments_filter = Q()
    if academic_year_id:
        enrollments_filter &= Q(semester__academic_year_id=academic_year_id)
    if unit_id:
        enrollments_filter &= Q(unit_id=unit_id)
    
    if export_format == 'csv':
        return export_to_csv(students, enrollments_filter, academic_year_id, unit_id)
    else:
        return export_to_excel(students, enrollments_filter, academic_year_id, unit_id)


def export_to_csv(students, enrollments_filter, academic_year_id, unit_id):
    """Export to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="student_progression.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    if unit_id:
        # Specific unit export
        unit = Unit.objects.get(id=unit_id)
        writer.writerow([
            'Registration Number', 'Student Name', 'Programme', 'Year',
            'Unit Code', 'Unit Name', 'Total Marks', 'Grade', 'Grade Point', 'Status'
        ])
        
        for student in students:
            enrollments = UnitEnrollment.objects.filter(
                student=student,
                unit_id=unit_id
            ).filter(enrollments_filter).select_related('final_grade', 'unit', 'semester')
            
            for enrollment in enrollments:
                writer.writerow([
                    student.registration_number,
                    f"{student.first_name} {student.last_name}",
                    student.programme.code,
                    student.current_year,
                    enrollment.unit.code,
                    enrollment.unit.name,
                    enrollment.final_grade.total_marks if hasattr(enrollment, 'final_grade') and enrollment.final_grade else '-',
                    enrollment.final_grade.grade if hasattr(enrollment, 'final_grade') and enrollment.final_grade else '-',
                    enrollment.final_grade.grade_point if hasattr(enrollment, 'final_grade') and enrollment.final_grade else '-',
                    enrollment.get_status_display()
                ])
    else:
        # General progression export
        writer.writerow([
            'Registration Number', 'Student Name', 'Programme', 'Year',
            'GPA', 'Classification', 'Total Credits', 'Completed Units', 'Failed Units'
        ])
        
        for student in students:
            enrollments = UnitEnrollment.objects.filter(
                student=student,
                status='COMPLETED'
            ).filter(enrollments_filter).select_related('final_grade', 'unit')
            
            gpa = calculate_gpa(enrollments)
            classification = get_classification(gpa)
            total_credits = enrollments.aggregate(total=Sum('unit__credit_hours'))['total'] or 0
            completed_units = enrollments.filter(final_grade__isnull=False).count()
            failed_units = enrollments.filter(final_grade__grade='F').count()
            
            writer.writerow([
                student.registration_number,
                f"{student.first_name} {student.last_name}",
                student.programme.code,
                student.current_year,
                gpa,
                classification,
                total_credits,
                completed_units,
                failed_units
            ])
    
    return response


def export_to_excel(students, enrollments_filter, academic_year_id, unit_id):
    """Export to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Progression"
    
    # Styling
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    if unit_id:
        # Specific unit export
        unit = Unit.objects.get(id=unit_id)
        headers = [
            'Registration Number', 'Student Name', 'Programme', 'Year',
            'Unit Code', 'Unit Name', 'Total Marks', 'Grade', 'Grade Point', 'Status'
        ]
        
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for student in students:
            enrollments = UnitEnrollment.objects.filter(
                student=student,
                unit_id=unit_id
            ).filter(enrollments_filter).select_related('final_grade', 'unit', 'semester')
            
            for enrollment in enrollments:
                ws.append([
                    student.registration_number,
                    f"{student.first_name} {student.last_name}",
                    student.programme.code,
                    student.current_year,
                    enrollment.unit.code,
                    enrollment.unit.name,
                    float(enrollment.final_grade.total_marks) if hasattr(enrollment, 'final_grade') and enrollment.final_grade else '-',
                    enrollment.final_grade.grade if hasattr(enrollment, 'final_grade') and enrollment.final_grade else '-',
                    float(enrollment.final_grade.grade_point) if hasattr(enrollment, 'final_grade') and enrollment.final_grade else '-',
                    enrollment.get_status_display()
                ])
    else:
        # General progression export
        headers = [
            'Registration Number', 'Student Name', 'Programme', 'Year',
            'GPA', 'Classification', 'Total Credits', 'Completed Units', 'Failed Units'
        ]
        
        ws.append(headers)
        
        # Style header
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for student in students:
            enrollments = UnitEnrollment.objects.filter(
                student=student,
                status='COMPLETED'
            ).filter(enrollments_filter).select_related('final_grade', 'unit')
            
            gpa = calculate_gpa(enrollments)
            classification = get_classification(gpa)
            total_credits = enrollments.aggregate(total=Sum('unit__credit_hours'))['total'] or 0
            completed_units = enrollments.filter(final_grade__isnull=False).count()
            failed_units = enrollments.filter(final_grade__grade='F').count()
            
            ws.append([
                student.registration_number,
                f"{student.first_name} {student.last_name}",
                student.programme.code,
                student.current_year,
                gpa,
                classification,
                total_credits,
                completed_units,
                failed_units
            ])
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    response = HttpResponse(
        excel_file.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="student_progression.xlsx"'
    
    return response

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db import transaction
from .models import (
    Student, UnitEnrollment, Unit, Semester, Programme, 
    AcademicYear, ProgrammeUnit, SemesterRegistration
)
from datetime import datetime


# ========================
# ENROLLMENT LIST VIEW
# ========================

@login_required
def enrollment_list(request):
    """
    Display all enrollments with filtering options
    """
    # Get filter parameters
    search_query = request.GET.get('search', '')
    programme_id = request.GET.get('programme', '')
    semester_id = request.GET.get('semester', '')
    unit_id = request.GET.get('unit', '')
    status = request.GET.get('status', '')
    year_level = request.GET.get('year', '')
    
    # Base queryset
    enrollments = UnitEnrollment.objects.select_related(
        'student', 'student__programme', 'unit', 'semester', 'semester__academic_year'
    ).all().order_by('-enrollment_date')
    
    # Apply filters
    if search_query:
        enrollments = enrollments.filter(
            Q(student__registration_number__icontains=search_query) |
            Q(student__first_name__icontains=search_query) |
            Q(student__last_name__icontains=search_query) |
            Q(unit__code__icontains=search_query) |
            Q(unit__name__icontains=search_query)
        )
    
    if programme_id:
        enrollments = enrollments.filter(student__programme_id=programme_id)
    
    if semester_id:
        enrollments = enrollments.filter(semester_id=semester_id)
    
    if unit_id:
        enrollments = enrollments.filter(unit_id=unit_id)
    
    if status:
        enrollments = enrollments.filter(status=status)
    
    if year_level:
        enrollments = enrollments.filter(student__current_year=year_level)
    
    # Pagination
    paginator = Paginator(enrollments, 25)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    semesters = Semester.objects.select_related('academic_year').order_by('-academic_year__start_date', '-semester_number')
    units = Unit.objects.all().order_by('code')
    
    # Get statistics
    total_enrollments = enrollments.count()
    enrolled_count = enrollments.filter(status='ENROLLED').count()
    completed_count = enrollments.filter(status='COMPLETED').count()
    dropped_count = enrollments.filter(status='DROPPED').count()
    
    context = {
        'page_obj': page_obj,
        'programmes': programmes,
        'semesters': semesters,
        'units': units,
        'search_query': search_query,
        'selected_programme': programme_id,
        'selected_semester': semester_id,
        'selected_unit': unit_id,
        'selected_status': status,
        'selected_year': year_level,
        'total_enrollments': total_enrollments,
        'enrolled_count': enrolled_count,
        'completed_count': completed_count,
        'dropped_count': dropped_count,
    }
    
    return render(request, 'enrollments/enrollment_list.html', context)


# ========================
# SINGLE ENROLLMENT
# ========================

@login_required
def single_enrollment(request):
    """
    Enroll a single student in units
    """
    if request.method == 'POST':
        student_id = request.POST.get('student')
        semester_id = request.POST.get('semester')
        unit_ids = request.POST.getlist('units')
        
        try:
            student = Student.objects.get(user_id=student_id)
            semester = Semester.objects.get(id=semester_id)
            
            enrolled_count = 0
            already_enrolled = []
            errors = []
            
            with transaction.atomic():
                for unit_id in unit_ids:
                    unit = Unit.objects.get(id=unit_id)
                    
                    # Check if already enrolled
                    existing = UnitEnrollment.objects.filter(
                        student=student,
                        unit=unit,
                        semester=semester
                    ).first()
                    
                    if existing:
                        already_enrolled.append(unit.code)
                        continue
                    
                    # Create enrollment
                    UnitEnrollment.objects.create(
                        student=student,
                        unit=unit,
                        semester=semester,
                        status='ENROLLED'
                    )
                    enrolled_count += 1
                
                # Update semester registration
                sem_reg, created = SemesterRegistration.objects.get_or_create(
                    student=student,
                    semester=semester,
                    defaults={'status': 'REGISTERED', 'units_enrolled': 0}
                )
                sem_reg.units_enrolled = UnitEnrollment.objects.filter(
                    student=student,
                    semester=semester,
                    status='ENROLLED'
                ).count()
                sem_reg.save()
            
            if enrolled_count > 0:
                messages.success(request, f'Successfully enrolled {student.registration_number} in {enrolled_count} unit(s).')
            
            if already_enrolled:
                messages.warning(request, f'Already enrolled in: {", ".join(already_enrolled)}')
            
            return redirect('enrollment_list')
            
        except Exception as e:
            messages.error(request, f'Error enrolling student: {str(e)}')
            return redirect('single_enrollment')
    
    # GET request
    students = Student.objects.filter(is_active=True).select_related('programme').order_by('registration_number')
    semesters = Semester.objects.select_related('academic_year').order_by('-academic_year__start_date')
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    
    context = {
        'students': students,
        'semesters': semesters,
        'programmes': programmes,
    }
    
    return render(request, 'enrollments/single_enrollment.html', context)


# ========================
# BULK ENROLLMENT
# ========================

@login_required
def bulk_enrollment(request):
    """
    Enroll multiple students in multiple units at once
    """
    if request.method == 'POST':
        programme_id = request.POST.get('programme')
        year_level = request.POST.get('year_level')
        semester_id = request.POST.get('semester')
        unit_ids = request.POST.getlist('units')
        
        if not all([programme_id, year_level, semester_id, unit_ids]):
            messages.error(request, 'Please fill in all required fields.')
            return redirect('bulk_enrollment')
        
        try:
            programme = Programme.objects.get(id=programme_id)
            semester = Semester.objects.get(id=semester_id)
            units = Unit.objects.filter(id__in=unit_ids)
            
            # Get all students matching criteria
            students = Student.objects.filter(
                programme=programme,
                current_year=year_level,
                is_active=True
            )
            
            if not students.exists():
                messages.warning(request, 'No students found matching the criteria.')
                return redirect('bulk_enrollment')
            
            enrolled_count = 0
            skipped_count = 0
            student_count = students.count()
            
            with transaction.atomic():
                for student in students:
                    # Check or create semester registration
                    sem_reg, created = SemesterRegistration.objects.get_or_create(
                        student=student,
                        semester=semester,
                        defaults={'status': 'REGISTERED', 'units_enrolled': 0}
                    )
                    
                    for unit in units:
                        # Check if already enrolled
                        existing = UnitEnrollment.objects.filter(
                            student=student,
                            unit=unit,
                            semester=semester
                        ).first()
                        
                        if existing:
                            skipped_count += 1
                            continue
                        
                        # Create enrollment
                        UnitEnrollment.objects.create(
                            student=student,
                            unit=unit,
                            semester=semester,
                            status='ENROLLED'
                        )
                        enrolled_count += 1
                    
                    # Update semester registration unit count
                    sem_reg.units_enrolled = UnitEnrollment.objects.filter(
                        student=student,
                        semester=semester,
                        status='ENROLLED'
                    ).count()
                    sem_reg.save()
            
            messages.success(
                request, 
                f'Bulk enrollment completed! Enrolled {student_count} students in {len(units)} units. '
                f'Total enrollments created: {enrolled_count}. Skipped (already enrolled): {skipped_count}.'
            )
            return redirect('enrollment_list')
            
        except Exception as e:
            messages.error(request, f'Error during bulk enrollment: {str(e)}')
            return redirect('bulk_enrollment')
    
    # GET request
    programmes = Programme.objects.filter(is_active=True).order_by('name')
    semesters = Semester.objects.select_related('academic_year').order_by('-academic_year__start_date')
    academic_years = AcademicYear.objects.all().order_by('-start_date')
    
    context = {
        'programmes': programmes,
        'semesters': semesters,
        'academic_years': academic_years,
    }
    
    return render(request, 'enrollments/bulk_enrollment.html', context)


# ========================
# AJAX ENDPOINTS
# ========================

@login_required
def get_programme_units(request):
    """
    Get units for a specific programme and year level
    AJAX endpoint
    """
    programme_id = request.GET.get('programme_id')
    year_level = request.GET.get('year_level')
    semester_number = request.GET.get('semester_number')
    
    if not all([programme_id, year_level]):
        return JsonResponse({'units': []})
    
    # Get programme units
    programme_units = ProgrammeUnit.objects.filter(
        programme_id=programme_id,
        year_level=year_level
    ).select_related('unit')
    
    # Filter by semester if provided
    if semester_number:
        programme_units = programme_units.filter(semester=semester_number)
    
    units_data = []
    for pu in programme_units:
        units_data.append({
            'id': pu.unit.id,
            'code': pu.unit.code,
            'name': pu.unit.name,
            'credits': pu.unit.credit_hours,
            'is_mandatory': pu.is_mandatory,
            'semester': pu.semester
        })
    
    return JsonResponse({'units': units_data})


@login_required
def get_students_count(request):
    """
    Get count of students matching criteria for bulk enrollment
    AJAX endpoint
    """
    programme_id = request.GET.get('programme_id')
    year_level = request.GET.get('year_level')
    
    if not all([programme_id, year_level]):
        return JsonResponse({'count': 0, 'students': []})
    
    students = Student.objects.filter(
        programme_id=programme_id,
        current_year=year_level,
        is_active=True
    ).select_related('programme')
    
    students_data = []
    for student in students[:50]:  # Limit to 50 for preview
        students_data.append({
            'registration_number': student.registration_number,
            'name': f"{student.first_name} {student.last_name}",
            'programme': student.programme.code
        })
    
    return JsonResponse({
        'count': students.count(),
        'students': students_data,
        'has_more': students.count() > 50
    })


@login_required
def check_enrollment_conflicts(request):
    """
    Check if a student is already enrolled in selected units
    AJAX endpoint
    """
    student_id = request.GET.get('student_id')
    semester_id = request.GET.get('semester_id')
    unit_ids = request.GET.getlist('unit_ids[]')
    
    if not all([student_id, semester_id, unit_ids]):
        return JsonResponse({'conflicts': []})
    
    conflicts = []
    existing_enrollments = UnitEnrollment.objects.filter(
        student_id=student_id,
        semester_id=semester_id,
        unit_id__in=unit_ids
    ).select_related('unit')
    
    for enrollment in existing_enrollments:
        conflicts.append({
            'unit_code': enrollment.unit.code,
            'unit_name': enrollment.unit.name,
            'status': enrollment.get_status_display()
        })
    
    return JsonResponse({'conflicts': conflicts})


# ========================
# ENROLLMENT ACTIONS
# ========================

@login_required
def drop_enrollment(request, enrollment_id):
    """
    Drop a unit enrollment
    """
    if request.method == 'POST':
        try:
            enrollment = get_object_or_404(UnitEnrollment, id=enrollment_id)
            
            # Check if enrollment can be dropped
            if enrollment.status == 'COMPLETED':
                messages.error(request, 'Cannot drop a completed enrollment.')
                return redirect('enrollment_list')
            
            enrollment.status = 'DROPPED'
            enrollment.save()
            
            # Update semester registration count
            sem_reg = SemesterRegistration.objects.filter(
                student=enrollment.student,
                semester=enrollment.semester
            ).first()
            
            if sem_reg:
                sem_reg.units_enrolled = UnitEnrollment.objects.filter(
                    student=enrollment.student,
                    semester=enrollment.semester,
                    status='ENROLLED'
                ).count()
                sem_reg.save()
            
            messages.success(request, f'Successfully dropped {enrollment.unit.code} for {enrollment.student.registration_number}.')
            
        except Exception as e:
            messages.error(request, f'Error dropping enrollment: {str(e)}')
    
    return redirect('enrollment_list')


@login_required
def delete_enrollment(request, enrollment_id):
    """
    Delete an enrollment (admin only)
    """
    if request.method == 'POST':
        try:
            enrollment = get_object_or_404(UnitEnrollment, id=enrollment_id)
            student = enrollment.student
            semester = enrollment.semester
            unit_code = enrollment.unit.code
            reg_number = enrollment.student.registration_number
            
            enrollment.delete()
            
            # Update semester registration count
            sem_reg = SemesterRegistration.objects.filter(
                student=student,
                semester=semester
            ).first()
            
            if sem_reg:
                sem_reg.units_enrolled = UnitEnrollment.objects.filter(
                    student=student,
                    semester=semester,
                    status='ENROLLED'
                ).count()
                sem_reg.save()
            
            messages.success(request, f'Successfully deleted enrollment: {unit_code} for {reg_number}.')
            
        except Exception as e:
            messages.error(request, f'Error deleting enrollment: {str(e)}')
    
    return redirect('enrollment_list')


@login_required
def enrollment_detail(request, enrollment_id):
    """
    View detailed information about an enrollment
    """
    enrollment = get_object_or_404(
        UnitEnrollment.objects.select_related(
            'student', 'student__programme', 'unit', 'semester', 'semester__academic_year'
        ),
        id=enrollment_id
    )
    
    # Get marks for this enrollment
    marks = enrollment.marks.select_related('assessment_component', 'entered_by').all()
    
    # Get final grade if exists
    final_grade = None
    if hasattr(enrollment, 'final_grade'):
        final_grade = enrollment.final_grade
    
    context = {
        'enrollment': enrollment,
        'marks': marks,
        'final_grade': final_grade,
    }
    
    return render(request, 'enrollments/enrollment_detail.html', context)


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q, Count, Prefetch
from django.views.decorators.http import require_http_methods
from django.contrib import messages
from django.core.paginator import Paginator
from django.db import transaction
from .models import (
    UnitAllocation, Unit, Lecturer, Semester, Programme,
    AcademicYear, Department
)
import json


@login_required
def unit_allocation_list(request):
    """Main view for unit allocation management"""
    # Get current semester or latest
    current_semester = Semester.objects.filter(is_current=True).first()
    if not current_semester:
        current_semester = Semester.objects.order_by('-start_date').first()
    
    # Get filter parameters
    semester_id = request.GET.get('semester', current_semester.id if current_semester else None)
    department_id = request.GET.get('department', '')
    lecturer_id = request.GET.get('lecturer', '')
    unit_id = request.GET.get('unit', '')
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')
    
    # Base queryset with prefetch
    allocations = UnitAllocation.objects.select_related(
        'unit', 'lecturer', 'lecturer__user', 'lecturer__department', 'semester'
    ).prefetch_related('programmes')
    
    # Apply filters
    if semester_id:
        allocations = allocations.filter(semester_id=semester_id)
    
    if department_id:
        allocations = allocations.filter(lecturer__department_id=department_id)
    
    if lecturer_id:
        allocations = allocations.filter(lecturer_id=lecturer_id)
    
    if unit_id:
        allocations = allocations.filter(unit_id=unit_id)
    
    if status:
        allocations = allocations.filter(is_active=(status == 'active'))
    
    if search:
        allocations = allocations.filter(
            Q(unit__code__icontains=search) |
            Q(unit__name__icontains=search) |
            Q(lecturer__staff_number__icontains=search) |
            Q(lecturer__user__first_name__icontains=search) |
            Q(lecturer__user__last_name__icontains=search)
        )
    
    # Order by date
    allocations = allocations.order_by('-allocated_date')
    
    # Pagination
    paginator = Paginator(allocations, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Get filter options
    semesters = Semester.objects.order_by('-start_date')
    departments = Department.objects.order_by('name')
    
    # Statistics
    total_allocations = allocations.count()
    active_allocations = allocations.filter(is_active=True).count()
    inactive_allocations = allocations.filter(is_active=False).count()
    
    context = {
        'allocations': page_obj,
        'semesters': semesters,
        'departments': departments,
        'current_semester': current_semester,
        'selected_semester': semester_id,
        'selected_department': department_id,
        'selected_lecturer': lecturer_id,
        'selected_unit': unit_id,
        'selected_status': status,
        'search_query': search,
        'total_allocations': total_allocations,
        'active_allocations': active_allocations,
        'inactive_allocations': inactive_allocations,
    }
    
    return render(request, 'admin/unit_allocation_list.html', context)


@login_required
@require_http_methods(["POST"])
def unit_allocation_create(request):
    """Create new unit allocation via AJAX"""
    try:
        data = json.loads(request.body)
        
        unit_id = data.get('unit_id')
        lecturer_id = data.get('lecturer_id')
        semester_id = data.get('semester_id')
        programme_ids = data.get('programme_ids', [])
        
        # Validate required fields
        if not all([unit_id, lecturer_id, semester_id]):
            return JsonResponse({
                'success': False,
                'message': 'Unit, Lecturer, and Semester are required.'
            }, status=400)
        
        # Get objects
        unit = get_object_or_404(Unit, id=unit_id)
        lecturer = get_object_or_404(Lecturer, user_id=lecturer_id)
        semester = get_object_or_404(Semester, id=semester_id)
        
        # Check if allocation already exists
        existing = UnitAllocation.objects.filter(
            unit=unit,
            lecturer=lecturer,
            semester=semester
        ).first()
        
        if existing:
            return JsonResponse({
                'success': False,
                'message': f'This unit is already allocated to {lecturer.user.get_full_name()} for this semester.'
            }, status=400)
        
        # Create allocation
        with transaction.atomic():
            allocation = UnitAllocation.objects.create(
                unit=unit,
                lecturer=lecturer,
                semester=semester,
                is_active=True
            )
            
            # Add programmes
            if programme_ids:
                programmes = Programme.objects.filter(id__in=programme_ids)
                allocation.programmes.set(programmes)
        
        return JsonResponse({
            'success': True,
            'message': 'Unit allocated successfully!',
            'allocation': {
                'id': allocation.id,
                'unit_code': unit.code,
                'unit_name': unit.name,
                'lecturer_name': lecturer.user.get_full_name(),
                'semester': str(semester),
                'allocated_date': allocation.allocated_date.strftime('%d %b %Y %H:%M')
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid JSON data.'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def unit_allocation_bulk_create(request):
    """Create multiple unit allocations at once"""
    try:
        data = json.loads(request.body)
        
        unit_ids = data.get('unit_ids', [])
        lecturer_id = data.get('lecturer_id')
        semester_id = data.get('semester_id')
        programme_ids = data.get('programme_ids', [])
        
        if not unit_ids or not lecturer_id or not semester_id:
            return JsonResponse({
                'success': False,
                'message': 'Units, Lecturer, and Semester are required.'
            }, status=400)
        
        lecturer = get_object_or_404(Lecturer, user_id=lecturer_id)
        semester = get_object_or_404(Semester, id=semester_id)
        programmes = Programme.objects.filter(id__in=programme_ids)
        
        created_count = 0
        skipped_count = 0
        skipped_units = []
        
        with transaction.atomic():
            for unit_id in unit_ids:
                unit = get_object_or_404(Unit, id=unit_id)
                
                # Check if already allocated
                existing = UnitAllocation.objects.filter(
                    unit=unit,
                    lecturer=lecturer,
                    semester=semester
                ).exists()
                
                if existing:
                    skipped_count += 1
                    skipped_units.append(unit.code)
                    continue
                
                # Create allocation
                allocation = UnitAllocation.objects.create(
                    unit=unit,
                    lecturer=lecturer,
                    semester=semester,
                    is_active=True
                )
                
                if programme_ids:
                    allocation.programmes.set(programmes)
                
                created_count += 1
        
        message = f'{created_count} unit(s) allocated successfully.'
        if skipped_count > 0:
            message += f' {skipped_count} unit(s) skipped (already allocated): {", ".join(skipped_units)}'
        
        return JsonResponse({
            'success': True,
            'message': message,
            'created_count': created_count,
            'skipped_count': skipped_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def unit_allocation_update(request, allocation_id):
    """Update unit allocation"""
    try:
        allocation = get_object_or_404(UnitAllocation, id=allocation_id)
        data = json.loads(request.body)
        
        programme_ids = data.get('programme_ids', [])
        is_active = data.get('is_active', allocation.is_active)
        
        with transaction.atomic():
            allocation.is_active = is_active
            allocation.save()
            
            if programme_ids:
                programmes = Programme.objects.filter(id__in=programme_ids)
                allocation.programmes.set(programmes)
        
        return JsonResponse({
            'success': True,
            'message': 'Allocation updated successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def unit_allocation_delete(request, allocation_id):
    """Delete unit allocation"""
    try:
        allocation = get_object_or_404(UnitAllocation, id=allocation_id)
        unit_code = allocation.unit.code
        
        allocation.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Allocation for {unit_code} deleted successfully!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
@require_http_methods(["POST"])
def unit_allocation_toggle_status(request, allocation_id):
    """Toggle allocation active status"""
    try:
        allocation = get_object_or_404(UnitAllocation, id=allocation_id)
        allocation.is_active = not allocation.is_active
        allocation.save()
        
        status = 'activated' if allocation.is_active else 'deactivated'
        
        return JsonResponse({
            'success': True,
            'message': f'Allocation {status} successfully!',
            'is_active': allocation.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


# AJAX API Endpoints for autocomplete and data fetching

@login_required
def api_search_units(request):
    """Search units for autocomplete"""
    query = request.GET.get('q', '')
    department_id = request.GET.get('department', '')
    semester_id = request.GET.get('semester', '')
    
    units = Unit.objects.filter(
        Q(code__icontains=query) | Q(name__icontains=query)
    )
    
    if department_id:
        units = units.filter(department_id=department_id)
    
    # Exclude already allocated units if semester is provided
    if semester_id:
        allocated_unit_ids = UnitAllocation.objects.filter(
            semester_id=semester_id,
            is_active=True
        ).values_list('unit_id', flat=True)
        # Don't exclude, just mark as allocated
    
    units = units[:20]  # Limit results
    
    results = []
    for unit in units:
        is_allocated = False
        if semester_id:
            is_allocated = UnitAllocation.objects.filter(
                unit=unit,
                semester_id=semester_id,
                is_active=True
            ).exists()
        
        results.append({
            'id': unit.id,
            'code': unit.code,
            'name': unit.name,
            'credit_hours': unit.credit_hours,
            'department': unit.department.name,
            'is_allocated': is_allocated
        })
    
    return JsonResponse({'results': results})


@login_required
def api_search_lecturers(request):
    """Search lecturers for autocomplete"""
    query = request.GET.get('q', '')
    department_id = request.GET.get('department', '')
    
    lecturers = Lecturer.objects.select_related('user', 'department').filter(
        Q(user__first_name__icontains=query) |
        Q(user__last_name__icontains=query) |
        Q(staff_number__icontains=query),
        is_active=True
    )
    
    if department_id:
        lecturers = lecturers.filter(department_id=department_id)
    
    lecturers = lecturers[:20]
    
    results = [{
        'id': lecturer.user_id,
        'staff_number': lecturer.staff_number,
        'name': lecturer.user.get_full_name(),
        'department': lecturer.department.name,
        'specialization': lecturer.specialization or 'N/A'
    } for lecturer in lecturers]
    
    return JsonResponse({'results': results})


@login_required
def api_get_programmes(request):
    """Get programmes for a department"""
    department_id = request.GET.get('department', '')
    
    programmes = Programme.objects.filter(is_active=True)
    
    if department_id:
        programmes = programmes.filter(department_id=department_id)
    
    programmes = programmes.order_by('name')
    
    results = [{
        'id': prog.id,
        'code': prog.code,
        'name': prog.name,
        'level': prog.get_level_display()
    } for prog in programmes]
    
    return JsonResponse({'results': results})


@login_required
def api_check_allocation(request):
    """Check if a unit is already allocated"""
    unit_id = request.GET.get('unit_id')
    semester_id = request.GET.get('semester_id')
    
    if not unit_id or not semester_id:
        return JsonResponse({'allocated': False})
    
    allocation = UnitAllocation.objects.filter(
        unit_id=unit_id,
        semester_id=semester_id,
        is_active=True
    ).select_related('lecturer', 'lecturer__user').first()
    
    if allocation:
        return JsonResponse({
            'allocated': True,
            'lecturer_name': allocation.lecturer.user.get_full_name(),
            'lecturer_staff_number': allocation.lecturer.staff_number,
            'allocation_id': allocation.id
        })
    
    return JsonResponse({'allocated': False})


@login_required
def api_lecturer_allocations(request, lecturer_id):
    """Get all allocations for a lecturer"""
    semester_id = request.GET.get('semester')
    
    allocations = UnitAllocation.objects.filter(
        lecturer_id=lecturer_id
    ).select_related('unit', 'semester').prefetch_related('programmes')
    
    if semester_id:
        allocations = allocations.filter(semester_id=semester_id)
    
    allocations = allocations.order_by('-allocated_date')
    
    results = [{
        'id': alloc.id,
        'unit_code': alloc.unit.code,
        'unit_name': alloc.unit.name,
        'semester': str(alloc.semester),
        'programmes': [{'code': p.code, 'name': p.name} for p in alloc.programmes.all()],
        'is_active': alloc.is_active,
        'allocated_date': alloc.allocated_date.strftime('%d %b %Y')
    } for alloc in allocations]
    
    return JsonResponse({'results': results})


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from .models import Department, User, Programme, Unit, Lecturer
import json

@login_required
def department_list(request):
    """List all departments with filters and search"""

    # Filters
    search = request.GET.get('search', '')
    sort_by = request.GET.get('sort', 'name')

    # Query departments with annotations
    departments = (
        Department.objects.annotate(
            programme_count=Count('programmes', distinct=True),
            unit_count=Count('units', distinct=True),
            lecturer_count=Count('lecturers', distinct=True),
        )
        .select_related('head_of_department', 'head_of_department__lecturer_profile')
    )

    # Search filter
    if search:
        departments = departments.filter(
            Q(name__icontains=search)
            | Q(code__icontains=search)
            | Q(description__icontains=search)
        )

    # Sorting
    valid_sort_fields = ['name', 'code', '-created_at', 'created_at']
    if sort_by in valid_sort_fields:
        departments = departments.order_by(sort_by)
    else:
        departments = departments.order_by('name')

    # Pagination
    paginator = Paginator(departments, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Summary statistics
    total_departments = Department.objects.count()
    with_heads = Department.objects.exclude(head_of_department__isnull=True).count()
    without_heads = Department.objects.filter(head_of_department__isnull=True).count()

    context = {
        'departments': page_obj,
        'search_query': search,
        'sort_by': sort_by,
        'total_departments': total_departments,
        'with_heads': with_heads,
        'without_heads': without_heads,
    }

    return render(request, 'department/department_list.html', context)



@login_required
def department_detail(request, code):
    """View department details"""

    department = get_object_or_404(
        Department.objects.annotate(
            programme_count=Count('programmes', distinct=True),
            unit_count=Count('units', distinct=True),
            lecturer_count=Count('lecturers', distinct=True)
        ).select_related('head_of_department'),
        code=code
    )

    # Related data
    programmes = department.programmes.filter(is_active=True).order_by('level', 'name')
    all_units = department.units.order_by('code')
    units = all_units[:10]  # ✅ only slice for display
    lecturers = (
        department.lecturers.filter(is_active=True)
        .select_related('user')
        .order_by('user__first_name')
    )

    # Programme statistics
    programme_by_level = programmes.values('level').annotate(count=Count('id')).order_by('level')

    # Unit statistics (✅ count from full queryset)
    core_units = all_units.filter(is_core=True).count()
    elective_units = all_units.filter(is_core=False).count()
    total_units = all_units.count()

    context = {
        'department': department,
        'programmes': programmes,
        'units': units,
        'lecturers': lecturers,
        'programme_by_level': programme_by_level,
        'core_units': core_units,
        'elective_units': elective_units,
        'total_units': total_units,
    }

    return render(request, 'department/department_detail.html', context)

@login_required
def department_create(request):
    """Create new department"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        head_of_department_id = request.POST.get('head_of_department', '')
        
        # Validation
        if not name or not code:
            messages.error(request, 'Department name and code are required.')
            return redirect('department_create')
        
        # Check if code already exists
        if Department.objects.filter(code=code).exists():
            messages.error(request, f'Department with code "{code}" already exists.')
            return redirect('department_create')
        
        try:
            with transaction.atomic():
                department = Department.objects.create(
                    name=name,
                    code=code,
                    description=description
                )
                
                # Assign head if provided
                if head_of_department_id:
                    try:
                        head = User.objects.get(id=head_of_department_id, user_type='COD')
                        department.head_of_department = head
                        department.save()
                    except User.DoesNotExist:
                        pass
                
                messages.success(request, f'Department "{name}" created successfully!')
                return redirect('department_detail', code=department.code)
                
        except Exception as e:
            messages.error(request, f'Error creating department: {str(e)}')
            return redirect('department_create')
    
    # GET request - show form
    # Get available CODs (not already assigned as heads)
    assigned_head_ids = Department.objects.exclude(
        head_of_department__isnull=True
    ).values_list('head_of_department_id', flat=True)
    
    available_cods = User.objects.filter(
        user_type='COD',
        is_active_user=True
    ).exclude(id__in=assigned_head_ids).order_by('first_name', 'last_name')
    
    context = {
        'available_cods': available_cods,
    }
    
    return render(request, 'admin/department_form.html', context)


@login_required
def department_update(request, code):
    """Update department"""
    department = get_object_or_404(Department, code=code)
    
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        new_code = request.POST.get('code', '').strip().upper()
        description = request.POST.get('description', '').strip()
        head_of_department_id = request.POST.get('head_of_department', '')
        
        # Validation
        if not name or not new_code:
            messages.error(request, 'Department name and code are required.')
            return redirect('department_update', code=code)
        
        # Check if new code already exists (excluding current department)
        if new_code != code and Department.objects.filter(code=new_code).exists():
            messages.error(request, f'Department with code "{new_code}" already exists.')
            return redirect('department_update', code=code)
        
        try:
            with transaction.atomic():
                department.name = name
                department.code = new_code
                department.description = description
                
                # Update head
                if head_of_department_id:
                    try:
                        head = User.objects.get(id=head_of_department_id, user_type='COD')
                        department.head_of_department = head
                    except User.DoesNotExist:
                        department.head_of_department = None
                else:
                    department.head_of_department = None
                
                department.save()
                
                messages.success(request, f'Department "{name}" updated successfully!')
                return redirect('department_detail', code=department.code)
                
        except Exception as e:
            messages.error(request, f'Error updating department: {str(e)}')
            return redirect('department_update', code=code)
    
    # GET request - show form
    # Get available CODs
    assigned_head_ids = Department.objects.exclude(
        head_of_department__isnull=True
    ).exclude(id=department.id).values_list('head_of_department_id', flat=True)
    
    available_cods = User.objects.filter(
        user_type='COD',
        is_active_user=True
    ).exclude(id__in=assigned_head_ids).order_by('first_name', 'last_name')
    
    context = {
        'department': department,
        'available_cods': available_cods,
        'is_update': True,
    }
    
    return render(request, 'department/department_form.html', context)


@login_required
def department_delete(request, code):
    """Delete department"""
    department = get_object_or_404(Department, code=code)
    
    if request.method == 'POST':
        # Check if department has related data
        has_programmes = department.programmes.exists()
        has_units = department.units.exists()
        has_lecturers = department.lecturers.exists()
        
        if has_programmes or has_units or has_lecturers:
            messages.error(
                request,
                f'Cannot delete department "{department.name}". '
                'It has associated programmes, units, or lecturers. '
                'Please remove or reassign them first.'
            )
            return redirect('department_detail', code=code)
        
        try:
            dept_name = department.name
            department.delete()
            messages.success(request, f'Department "{dept_name}" deleted successfully!')
            return redirect('department_list')
        except Exception as e:
            messages.error(request, f'Error deleting department: {str(e)}')
            return redirect('department_detail', code=code)
    
    # GET request - show confirmation page
    # Count related objects
    programme_count = department.programmes.count()
    unit_count = department.units.count()
    lecturer_count = department.lecturers.count()
    
    context = {
        'department': department,
        'programme_count': programme_count,
        'unit_count': unit_count,
        'lecturer_count': lecturer_count,
        'can_delete': programme_count == 0 and unit_count == 0 and lecturer_count == 0,
    }
    
    return render(request, 'department/department_confirm_delete.html', context)


# AJAX API Endpoints

@login_required
def api_search_departments(request):
    """Search departments for autocomplete"""
    query = request.GET.get('q', '')
    
    departments = Department.objects.filter(
        Q(name__icontains=query) | Q(code__icontains=query)
    ).order_by('name')[:20]
    
    results = [{
        'id': dept.id,
        'code': dept.code,
        'name': dept.name,
        'head': dept.head_of_department.get_full_name() if dept.head_of_department else 'No Head'
    } for dept in departments]
    
    return JsonResponse({'results': results})


@login_required
def api_search_cods(request):
    """Search CODs (Chairmen of Department) for autocomplete"""
    query = request.GET.get('q', '')
    exclude_assigned = request.GET.get('exclude_assigned', 'true') == 'true'
    current_dept_id = request.GET.get('current_dept', '')
    
    cods = User.objects.filter(
        user_type='COD',
        is_active_user=True
    ).filter(
        Q(first_name__icontains=query) |
        Q(last_name__icontains=query) |
        Q(username__icontains=query)
    )
    
    if exclude_assigned:
        # Exclude CODs already assigned as heads
        assigned_ids = Department.objects.exclude(
            head_of_department__isnull=True
        )
        
        if current_dept_id:
            assigned_ids = assigned_ids.exclude(id=current_dept_id)
        
        assigned_ids = assigned_ids.values_list('head_of_department_id', flat=True)
        cods = cods.exclude(id__in=assigned_ids)
    
    cods = cods.order_by('first_name', 'last_name')[:20]
    
    results = [{
        'id': cod.id,
        'username': cod.username,
        'name': cod.get_full_name(),
        'email': cod.email
    } for cod in cods]
    
    return JsonResponse({'results': results})


@login_required
@require_http_methods(["POST"])
def api_department_quick_update(request, code):
    """Quick update department via AJAX"""
    try:
        department = get_object_or_404(Department, code=code)
        data = json.loads(request.body)
        
        field = data.get('field')
        value = data.get('value')
        
        if field == 'name':
            department.name = value
        elif field == 'description':
            department.description = value
        elif field == 'head_of_department':
            if value:
                head = get_object_or_404(User, id=value, user_type='COD')
                department.head_of_department = head
            else:
                department.head_of_department = None
        else:
            return JsonResponse({
                'success': False,
                'message': 'Invalid field'
            }, status=400)
        
        department.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Department updated successfully'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)


@login_required
def api_department_stats(request, code):
    """Get department statistics"""
    department = get_object_or_404(Department, code=code)
    
    programmes = department.programmes.all()
    units = department.units.all()
    lecturers = department.lecturers.filter(is_active=True)
    
    # Programme breakdown
    programme_levels = programmes.values('level').annotate(
        count=Count('id')
    ).order_by('level')
    
    # Unit breakdown
    core_units = units.filter(is_core=True).count()
    elective_units = units.filter(is_core=False).count()
    
    stats = {
        'total_programmes': programmes.count(),
        'active_programmes': programmes.filter(is_active=True).count(),
        'total_units': units.count(),
        'core_units': core_units,
        'elective_units': elective_units,
        'total_lecturers': lecturers.count(),
        'programme_levels': list(programme_levels),
    }
    
    return JsonResponse(stats)


from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Q, Count, Prefetch
from datetime import datetime, timedelta
import calendar
import json

from .models import (
    Event, Announcement, Semester, UnitEnrollment, 
    TimetableSlot, SemesterRegistration, Student,
    AcademicYear, Programme
)


@login_required
def admin_academic_calendar(request):
    """
    Admin Academic Calendar View - Shows all activities in a calendar format
    """
    # Check if user is admin/staff
    if not request.user.user_type in ['ICT_ADMIN', 'DEAN', 'COD']:
        return render(request, 'error.html', {
            'message': 'You do not have permission to access this page.'
        })
    
    # Get month and year from query params or use current
    today = timezone.now().date()
    month = int(request.GET.get('month', today.month))
    year = int(request.GET.get('year', today.year))
    
    # Validate month and year
    if month < 1 or month > 12:
        month = today.month
    if year < 2020 or year > 2030:
        year = today.year
    
    # Get month name
    month_name = calendar.month_name[month]
    
    # Generate calendar matrix
    cal = calendar.monthcalendar(year, month)
    
    # Get date range for the month
    first_day = datetime(year, month, 1).date()
    if month == 12:
        last_day = datetime(year + 1, 1, 1).date() - timedelta(days=1)
    else:
        last_day = datetime(year, month + 1, 1).date() - timedelta(days=1)
    
    # Fetch all data for the month
    calendar_data = {}
    
    # 1. Get Events
    events = Event.objects.filter(
        event_date__gte=first_day,
        event_date__lte=last_day,
        is_published=True
    ).select_related('venue', 'organizer').prefetch_related('target_programmes')
    
    for event in events:
        date_key = event.event_date.strftime('%Y-%m-%d')
        if date_key not in calendar_data:
            calendar_data[date_key] = {
                'events': [],
                'announcements': [],
                'semesters': [],
                'registrations': [],
                'classes': []
            }
        
        # Determine event icon based on type
        icon_map = {
            'SEMINAR': 'bi-chat-dots',
            'WORKSHOP': 'bi-tools',
            'CONFERENCE': 'bi-people',
            'MEETING': 'bi-calendar2-check',
            'ORIENTATION': 'bi-compass',
            'EXAMINATION': 'bi-file-earmark-text',
            'OTHER': 'bi-calendar-event'
        }
        
        calendar_data[date_key]['events'].append({
            'id': event.id,
            'title': event.title,
            'type': event.get_event_type_display(),
            'time': f"{event.start_time.strftime('%H:%M')} - {event.end_time.strftime('%H:%M')}",
            'venue': event.venue.name if event.venue else 'TBA',
            'is_mandatory': event.is_mandatory,
            'organizer': event.organizer.get_full_name(),
            'icon': icon_map.get(event.event_type, 'bi-calendar-event'),
            'description': event.description[:100] + '...' if len(event.description) > 100 else event.description,
            'attendees': event.registrations.count() if event.registration_required else None,
            'max_attendees': event.max_attendees
        })
    
    # 2. Get Announcements (published within or expiring in this month)
    announcements = Announcement.objects.filter(
        Q(publish_date__gte=first_day, publish_date__lte=last_day) |
        Q(expiry_date__gte=first_day, expiry_date__lte=last_day),
        is_published=True
    ).select_related('created_by').prefetch_related('target_programmes')
    
    for announcement in announcements:
        # Add to publish date
        publish_date_key = announcement.publish_date.date().strftime('%Y-%m-%d')
        if first_day <= announcement.publish_date.date() <= last_day:
            if publish_date_key not in calendar_data:
                calendar_data[publish_date_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'registrations': [],
                    'classes': []
                }
            
            calendar_data[publish_date_key]['announcements'].append({
                'id': announcement.id,
                'title': announcement.title,
                'priority': announcement.get_priority_display(),
                'author': announcement.created_by.get_full_name(),
                'content_preview': announcement.content[:150] + '...' if len(announcement.content) > 150 else announcement.content,
                'expiry_date': announcement.expiry_date.strftime('%Y-%m-%d') if announcement.expiry_date else None,
                'programmes': ', '.join([p.code for p in announcement.target_programmes.all()]) or 'All Programmes'
            })
        
        # Add marker for expiry date if different from publish date
        if announcement.expiry_date and first_day <= announcement.expiry_date.date() <= last_day:
            expiry_date_key = announcement.expiry_date.date().strftime('%Y-%m-%d')
            if expiry_date_key != publish_date_key:
                if expiry_date_key not in calendar_data:
                    calendar_data[expiry_date_key] = {
                        'events': [],
                        'announcements': [],
                        'semesters': [],
                        'registrations': [],
                        'classes': []
                    }
                
                calendar_data[expiry_date_key]['announcements'].append({
                    'id': announcement.id,
                    'title': f"{announcement.title} (Expires)",
                    'priority': announcement.get_priority_display(),
                    'author': announcement.created_by.get_full_name(),
                    'content_preview': announcement.content[:150] + '...' if len(announcement.content) > 150 else announcement.content,
                    'is_expiry': True,
                    'programmes': ', '.join([p.code for p in announcement.target_programmes.all()]) or 'All Programmes'
                })
    
    # 3. Get Semester Information (start dates, end dates, registration deadlines)
    semesters = Semester.objects.filter(
        Q(start_date__gte=first_day, start_date__lte=last_day) |
        Q(end_date__gte=first_day, end_date__lte=last_day) |
        Q(registration_deadline__gte=first_day, registration_deadline__lte=last_day)
    ).select_related('academic_year')
    
    for semester in semesters:
        # Start date
        if first_day <= semester.start_date <= last_day:
            start_key = semester.start_date.strftime('%Y-%m-%d')
            if start_key not in calendar_data:
                calendar_data[start_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'registrations': [],
                    'classes': []
                }
            
            calendar_data[start_key]['semesters'].append({
                'id': semester.id,
                'title': f"{semester} - Starts",
                'semester': str(semester),
                'type': 'start',
                'is_current': semester.is_current,
                'academic_year': semester.academic_year.year_code
            })
        
        # End date
        if first_day <= semester.end_date <= last_day:
            end_key = semester.end_date.strftime('%Y-%m-%d')
            if end_key not in calendar_data:
                calendar_data[end_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'registrations': [],
                    'classes': []
                }
            
            calendar_data[end_key]['semesters'].append({
                'id': semester.id,
                'title': f"{semester} - Ends",
                'semester': str(semester),
                'type': 'end',
                'is_current': semester.is_current,
                'academic_year': semester.academic_year.year_code
            })
        
        # Registration deadline
        if first_day <= semester.registration_deadline <= last_day:
            reg_key = semester.registration_deadline.strftime('%Y-%m-%d')
            if reg_key not in calendar_data:
                calendar_data[reg_key] = {
                    'events': [],
                    'announcements': [],
                    'semesters': [],
                    'registrations': [],
                    'classes': []
                }
            
            calendar_data[reg_key]['registrations'].append({
                'id': semester.id,
                'title': f"{semester} - Registration Deadline",
                'semester': str(semester),
                'type': 'deadline',
                'is_current': semester.is_current,
                'academic_year': semester.academic_year.year_code,
                'registered_count': semester.registrations.filter(status='REGISTERED').count(),
                'total_students': Student.objects.filter(is_active=True).count()
            })
    
    # 4. Get Daily Class Count (from timetable)
    # For each day of the week in the month, count scheduled classes
    for week in cal:
        for day in week:
            if day == 0:
                continue
            
            date_obj = datetime(year, month, day).date()
            date_key = date_obj.strftime('%Y-%m-%d')
            day_name = date_obj.strftime('%A').upper()
            
            # Count timetable slots for this day
            timetable_slots = TimetableSlot.objects.filter(
                day_of_week=day_name,
                is_active=True,
                unit_allocation__semester__is_current=True
            ).select_related(
                'unit_allocation__unit',
                'unit_allocation__lecturer__user',
                'venue',
                'programme'
            ).distinct()
            
            if timetable_slots.exists():
                if date_key not in calendar_data:
                    calendar_data[date_key] = {
                        'events': [],
                        'announcements': [],
                        'semesters': [],
                        'registrations': [],
                        'classes': []
                    }
                
                for slot in timetable_slots:
                    calendar_data[date_key]['classes'].append({
                        'id': slot.id,
                        'unit': slot.unit_allocation.unit.code,
                        'unit_name': slot.unit_allocation.unit.name,
                        'time': f"{slot.start_time.strftime('%H:%M')} - {slot.end_time.strftime('%H:%M')}",
                        'venue': slot.venue.code,
                        'venue_name': slot.venue.name,
                        'lecturer': slot.unit_allocation.lecturer.user.get_full_name(),
                        'programme': slot.programme.code,
                        'year_level': slot.year_level
                    })
    
    # Get statistics
    current_semester = Semester.objects.filter(is_current=True).first()
    stats = {
        'total_events': Event.objects.filter(
            event_date__gte=first_day,
            event_date__lte=last_day,
            is_published=True
        ).count(),
        'total_announcements': Announcement.objects.filter(
            publish_date__gte=first_day,
            publish_date__lte=last_day,
            is_published=True
        ).count(),
        'active_semester': str(current_semester) if current_semester else 'No Active Semester',
        'registered_students': SemesterRegistration.objects.filter(
            semester=current_semester,
            status='REGISTERED'
        ).count() if current_semester else 0,
        'total_students': Student.objects.filter(is_active=True).count(),
    }
    
    # Get upcoming events (next 7 days from selected month)
    upcoming_events = Event.objects.filter(
        event_date__gte=today,
        event_date__lte=today + timedelta(days=7),
        is_published=True
    ).order_by('event_date', 'start_time')[:5]
    
    context = {
        'month': month,
        'year': year,
        'month_name': month_name,
        'calendar': cal,
        'today': today,
        'calendar_data': calendar_data,
        'calendar_data_json': json.dumps(calendar_data),
        'stats': stats,
        'upcoming_events': upcoming_events,
        'current_semester': current_semester,
    }
    
    return render(request, 'admin/admin_academic_calendar.html', context)


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import ensure_csrf_cookie
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils import timezone
from django.db import transaction
import json
import traceback

from .models import Venue, TimetableSlot, User


@login_required
@ensure_csrf_cookie
def venue_list(request):
    """
    Main venue management page
    """
    # Check if user has permission
    if not request.user.user_type in ['ICT_ADMIN', 'DEAN', 'COD']:
        return render(request, 'error.html', {
            'message': 'You do not have permission to access this page.'
        })
    
    # Get statistics
    stats = {
        'total_venues': Venue.objects.count(),
        'available_venues': Venue.objects.filter(is_available=True).count(),
        'lecture_halls': Venue.objects.filter(venue_type='LECTURE_HALL').count(),
        'labs': Venue.objects.filter(venue_type='LAB').count(),
        'venues_with_projectors': Venue.objects.filter(has_projector=True).count(),
        'venues_with_computers': Venue.objects.filter(has_computers=True).count(),
    }
    
    context = {
        'stats': stats,
        'venue_types': Venue.VENUE_TYPES,
    }
    
    return render(request, 'admin/venue_management.html', context)


@login_required
@require_http_methods(["GET"])
def venue_list_ajax(request):
    """
    AJAX endpoint to get paginated venue list with filters
    """
    try:
        # Get filter parameters with defaults
        search_query = request.GET.get('search', '').strip()
        venue_type = request.GET.get('venue_type', '').strip()
        availability = request.GET.get('availability', '').strip()
        building = request.GET.get('building', '').strip()
        
        try:
            page = int(request.GET.get('page', 1))
            if page < 1:
                page = 1
        except (ValueError, TypeError):
            page = 1
            
        try:
            per_page = int(request.GET.get('per_page', 10))
            if per_page < 1 or per_page > 100:
                per_page = 10
        except (ValueError, TypeError):
            per_page = 10
        
        # Base queryset
        venues = Venue.objects.all()
        
        # Apply filters
        if search_query:
            venues = venues.filter(
                Q(name__icontains=search_query) |
                Q(code__icontains=search_query) |
                Q(building__icontains=search_query)
            )
        
        if venue_type:
            venues = venues.filter(venue_type=venue_type)
        
        if availability:
            is_available = availability.lower() == 'true'
            venues = venues.filter(is_available=is_available)
        
        if building:
            venues = venues.filter(building__icontains=building)
        
        # Annotate with timetable count
        venues = venues.annotate(
            timetable_count=Count('timetable_slots', distinct=True)
        )
        
        # Order by building, then name
        venues = venues.order_by('building', 'name')
        
        # Get total count before pagination
        total_count = venues.count()
        
        # Pagination
        paginator = Paginator(venues, per_page)
        
        # Handle invalid page numbers
        if page > paginator.num_pages:
            page = paginator.num_pages if paginator.num_pages > 0 else 1
            
        page_obj = paginator.get_page(page)
        
        # Serialize data
        venues_data = []
        for venue in page_obj.object_list:
            try:
                venues_data.append({
                    'id': venue.id,
                    'name': venue.name,
                    'code': venue.code,
                    'venue_type': venue.get_venue_type_display(),
                    'venue_type_code': venue.venue_type,
                    'capacity': venue.capacity,
                    'building': venue.building,
                    'floor': venue.floor if venue.floor else '-',
                    'has_projector': venue.has_projector,
                    'has_computers': venue.has_computers,
                    'is_available': venue.is_available,
                    'timetable_count': venue.timetable_count,
                })
            except Exception as e:
                print(f"Error serializing venue {venue.id}: {str(e)}")
                continue
        
        return JsonResponse({
            'success': True,
            'venues': venues_data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': total_count,
                'has_previous': page_obj.has_previous(),
                'has_next': page_obj.has_next(),
                'per_page': per_page,
            }
        })
        
    except Exception as e:
        print("Error in venue_list_ajax:")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def venue_detail_ajax(request, venue_id):
    """
    AJAX endpoint to get venue details
    """
    try:
        venue = get_object_or_404(Venue, id=venue_id)
        
        # Get timetable slots for this venue
        timetable_slots = TimetableSlot.objects.filter(
            venue=venue,
            is_active=True,
            unit_allocation__semester__is_current=True
        ).select_related(
            'unit_allocation__unit',
            'unit_allocation__lecturer__user',
            'programme'
        ).order_by('day_of_week', 'start_time')
        
        # Group by day
        timetable_by_day = {}
        for slot in timetable_slots:
            day = slot.get_day_of_week_display()
            if day not in timetable_by_day:
                timetable_by_day[day] = []
            
            timetable_by_day[day].append({
                'id': slot.id,
                'unit_code': slot.unit_allocation.unit.code,
                'unit_name': slot.unit_allocation.unit.name,
                'lecturer': slot.unit_allocation.lecturer.user.get_full_name(),
                'programme': slot.programme.code,
                'year_level': slot.year_level,
                'start_time': slot.start_time.strftime('%H:%M'),
                'end_time': slot.end_time.strftime('%H:%M'),
            })
        
        venue_data = {
            'id': venue.id,
            'name': venue.name,
            'code': venue.code,
            'venue_type': venue.get_venue_type_display(),
            'venue_type_code': venue.venue_type,
            'capacity': venue.capacity,
            'building': venue.building,
            'floor': venue.floor if venue.floor else '',
            'has_projector': venue.has_projector,
            'has_computers': venue.has_computers,
            'is_available': venue.is_available,
            'timetable': timetable_by_day,
            'total_slots': timetable_slots.count(),
        }
        
        return JsonResponse({
            'success': True,
            'venue': venue_data
        })
        
    except Venue.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Venue not found'
        }, status=404)
    except Exception as e:
        print(f"Error in venue_detail_ajax: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["POST"])
def venue_create_ajax(request):
    """
    AJAX endpoint to create a new venue
    """
    try:
        data = json.loads(request.body)
        
        # Validate required fields
        required_fields = ['name', 'code', 'venue_type', 'capacity', 'building']
        for field in required_fields:
            if not data.get(field):
                return JsonResponse({
                    'success': False,
                    'error': f'{field.replace("_", " ").title()} is required'
                }, status=400)
        
        # Check if code already exists
        if Venue.objects.filter(code=data['code'].upper()).exists():
            return JsonResponse({
                'success': False,
                'error': f'Venue code "{data["code"]}" already exists'
            }, status=400)
        
        # Validate capacity
        try:
            capacity = int(data['capacity'])
            if capacity < 1:
                raise ValueError
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'error': 'Capacity must be a positive number'
            }, status=400)
        
        # Create venue
        with transaction.atomic():
            venue = Venue.objects.create(
                name=data['name'].strip(),
                code=data['code'].strip().upper(),
                venue_type=data['venue_type'],
                capacity=capacity,
                building=data['building'].strip(),
                floor=data.get('floor', '').strip(),
                has_projector=bool(data.get('has_projector', False)),
                has_computers=bool(data.get('has_computers', False)),
                is_available=bool(data.get('is_available', True)),
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Venue created successfully',
            'venue': {
                'id': venue.id,
                'name': venue.name,
                'code': venue.code,
                'venue_type': venue.get_venue_type_display(),
                'capacity': venue.capacity,
                'building': venue.building,
            }
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        print(f"Error in venue_create_ajax: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["PUT", "PATCH"])
def venue_update_ajax(request, venue_id):
    """
    AJAX endpoint to update a venue
    """
    try:
        venue = get_object_or_404(Venue, id=venue_id)
        data = json.loads(request.body)
        
        # Check if code is being changed and already exists
        if 'code' in data and data['code'].upper() != venue.code:
            if Venue.objects.filter(code=data['code'].upper()).exclude(id=venue_id).exists():
                return JsonResponse({
                    'success': False,
                    'error': f'Venue code "{data["code"]}" already exists'
                }, status=400)
        
        # Validate capacity if provided
        if 'capacity' in data:
            try:
                capacity = int(data['capacity'])
                if capacity < 1:
                    raise ValueError
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'error': 'Capacity must be a positive number'
                }, status=400)
        
        # Update venue
        with transaction.atomic():
            if 'name' in data:
                venue.name = data['name'].strip()
            if 'code' in data:
                venue.code = data['code'].strip().upper()
            if 'venue_type' in data:
                venue.venue_type = data['venue_type']
            if 'capacity' in data:
                venue.capacity = int(data['capacity'])
            if 'building' in data:
                venue.building = data['building'].strip()
            if 'floor' in data:
                venue.floor = data['floor'].strip()
            if 'has_projector' in data:
                venue.has_projector = bool(data['has_projector'])
            if 'has_computers' in data:
                venue.has_computers = bool(data['has_computers'])
            if 'is_available' in data:
                venue.is_available = bool(data['is_available'])
            
            venue.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Venue updated successfully',
            'venue': {
                'id': venue.id,
                'name': venue.name,
                'code': venue.code,
                'venue_type': venue.get_venue_type_display(),
                'capacity': venue.capacity,
                'building': venue.building,
                'is_available': venue.is_available,
            }
        })
        
    except Venue.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Venue not found'
        }, status=404)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON data'
        }, status=400)
    except Exception as e:
        print(f"Error in venue_update_ajax: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["DELETE"])
def venue_delete_ajax(request, venue_id):
    """
    AJAX endpoint to delete a venue
    """
    try:
        venue = get_object_or_404(Venue, id=venue_id)
        
        # Check if venue has timetable slots
        timetable_count = venue.timetable_slots.filter(is_active=True).count()
        if timetable_count > 0:
            return JsonResponse({
                'success': False,
                'error': f'Cannot delete venue. It has {timetable_count} active timetable slot(s). Please remove or reassign them first.'
            }, status=400)
        
        venue_name = venue.name
        venue.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Venue "{venue_name}" deleted successfully'
        })
        
    except Venue.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Venue not found'
        }, status=404)
    except Exception as e:
        print(f"Error in venue_delete_ajax: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["POST"])
def venue_toggle_availability_ajax(request, venue_id):
    """
    AJAX endpoint to toggle venue availability
    """
    try:
        venue = get_object_or_404(Venue, id=venue_id)
        venue.is_available = not venue.is_available
        venue.save()
        
        status = "available" if venue.is_available else "unavailable"
        
        return JsonResponse({
            'success': True,
            'message': f'Venue marked as {status}',
            'is_available': venue.is_available
        })
        
    except Venue.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Venue not found'
        }, status=404)
    except Exception as e:
        print(f"Error in venue_toggle_availability_ajax: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def venue_buildings_ajax(request):
    """
    AJAX endpoint to get unique building names for filter
    """
    try:
        buildings = Venue.objects.values_list('building', flat=True).distinct().order_by('building')
        
        return JsonResponse({
            'success': True,
            'buildings': list(buildings)
        })
        
    except Exception as e:
        print(f"Error in venue_buildings_ajax: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)


@login_required
@require_http_methods(["GET"])
def venue_export_ajax(request):
    """
    AJAX endpoint to export venues data
    """
    try:
        # Get all venues
        venues = Venue.objects.all().order_by('building', 'name')
        
        venues_data = []
        for venue in venues:
            venues_data.append({
                'code': venue.code,
                'name': venue.name,
                'type': venue.get_venue_type_display(),
                'capacity': venue.capacity,
                'building': venue.building,
                'floor': venue.floor if venue.floor else '-',
                'projector': 'Yes' if venue.has_projector else 'No',
                'computers': 'Yes' if venue.has_computers else 'No',
                'available': 'Yes' if venue.is_available else 'No',
            })
        
        return JsonResponse({
            'success': True,
            'venues': venues_data,
            'total': len(venues_data)
        })
        
    except Exception as e:
        print(f"Error in venue_export_ajax: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'An error occurred: {str(e)}'
        }, status=400)